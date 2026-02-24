"""Shared helpers for financial modeling tools."""
from __future__ import annotations

import logging
import math
import uuid
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


# Lazy import to avoid circular imports at module load time
def _get_excel_models_dir() -> Path:
    from config.settings import EXCEL_MODELS_DIR

    return EXCEL_MODELS_DIR


# ── Excel styling constants ──────────────────────────────────────────────────

_FILL_HEADER = PatternFill("solid", fgColor="1C2030") if EXCEL_AVAILABLE else None
_FILL_INPUT = PatternFill("solid", fgColor="FFFDE7") if EXCEL_AVAILABLE else None
_FILL_FORMULA = PatternFill("solid", fgColor="E3F2FD") if EXCEL_AVAILABLE else None
_FILL_RESULT = PatternFill("solid", fgColor="E8F5E9") if EXCEL_AVAILABLE else None

_FONT_HEADER = Font(bold=True, color="FFFFFF") if EXCEL_AVAILABLE else None
_FONT_RESULT = Font(bold=True) if EXCEL_AVAILABLE else None
_ALIGN_CENTER = Alignment(horizontal="center") if EXCEL_AVAILABLE else None


# ── Excel styling functions ──────────────────────────────────────────────────


def _style_header(cell):
    cell.fill = _FILL_HEADER
    cell.font = _FONT_HEADER
    cell.alignment = _ALIGN_CENTER


def _style_input(cell):
    cell.fill = _FILL_INPUT


def _style_formula(cell):
    cell.fill = _FILL_FORMULA


def _style_result(cell):
    cell.fill = _FILL_RESULT
    cell.font = _FONT_RESULT


def _auto_width(ws, min_width: int = 12):
    """Set column widths to max content length."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _save_excel(wb, title: str) -> str | None:
    """Save workbook to data/models/{uuid}.xlsx, return file_id or None."""
    if not EXCEL_AVAILABLE:
        return None
    try:
        models_dir = _get_excel_models_dir()
        models_dir.mkdir(parents=True, exist_ok=True)
        file_id = str(uuid.uuid4())
        path = models_dir / f"{file_id}.xlsx"
        wb.save(str(path))
        return file_id
    except Exception as exc:
        logger.warning("Excel save failed: %s", exc)
        return None


# ── Math helpers (numpy_financial fallbacks) ─────────────────────────────────


def _pmt(rate: float, nper: int, pv: float) -> float:
    """PMT: periodic payment for a fully-amortizing loan.

    Formula: PMT = PV * r * (1+r)^n / ((1+r)^n - 1)
    """
    if rate == 0:
        return pv / nper
    factor = (1 + rate) ** nper
    return pv * rate * factor / (factor - 1)


def _irr(cash_flows: list[float], max_iter: int = 100, tol: float = 1e-6) -> float | None:
    """Compute IRR via Newton-Raphson. Returns per-period rate or None on failure."""
    rate = 0.1
    for _ in range(max_iter):
        npv = sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))
        dnpv = sum(-t * cf / (1 + rate) ** (t + 1) for t, cf in enumerate(cash_flows))
        if abs(dnpv) < 1e-12:
            break
        new_rate = rate - npv / dnpv
        if abs(new_rate - rate) < tol:
            return new_rate
        rate = new_rate
    return rate if abs(sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))) < 0.01 else None


# ── Black-Scholes-Merton helpers ─────────────────────────────────────────────


def _norm_cdf(x: float) -> float:
    """Standard normal CDF using math.erf."""
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def _norm_pdf(x: float) -> float:
    """Standard normal PDF."""
    return math.exp(-x * x / 2.0) / math.sqrt(2.0 * math.pi)


def _bsm_d1_d2(
    spot: float, strike: float, t: float, r: float, sigma: float, q: float = 0.0,
) -> tuple[float, float]:
    """Compute d1 and d2 for Black-Scholes-Merton."""
    d1 = (math.log(spot / strike) + (r - q + sigma * sigma / 2.0) * t) / (sigma * math.sqrt(t))
    d2 = d1 - sigma * math.sqrt(t)
    return d1, d2


def _bsm_price(
    spot: float, strike: float, t: float, r: float, sigma: float,
    option_type: str = "call", q: float = 0.0,
) -> float:
    """Return BSM option price (call or put)."""
    d1, d2 = _bsm_d1_d2(spot, strike, t, r, sigma, q)
    if option_type == "call":
        return spot * math.exp(-q * t) * _norm_cdf(d1) - strike * math.exp(-r * t) * _norm_cdf(d2)
    else:
        return strike * math.exp(-r * t) * _norm_cdf(-d2) - spot * math.exp(-q * t) * _norm_cdf(-d1)


def _bsm_vega_raw(
    spot: float, strike: float, t: float, r: float, sigma: float, q: float = 0.0,
) -> float:
    """Return BSM vega (raw, not per-1%)."""
    d1, _ = _bsm_d1_d2(spot, strike, t, r, sigma, q)
    return spot * math.exp(-q * t) * _norm_pdf(d1) * math.sqrt(t)


# ── Portfolio variance helper ────────────────────────────────────────────────


def _portfolio_variance(
    weights: list[float],
    volatilities: list[float],
    correlation_matrix: list[list[float]],
) -> float:
    """Compute portfolio variance from weights, volatilities, and correlation matrix.

    σ²_p = Σᵢ Σⱼ wᵢ wⱼ σᵢ σⱼ ρᵢⱼ
    """
    n = len(weights)
    var = 0.0
    for i in range(n):
        for j in range(n):
            var += weights[i] * weights[j] * volatilities[i] * volatilities[j] * correlation_matrix[i][j]
    return var
