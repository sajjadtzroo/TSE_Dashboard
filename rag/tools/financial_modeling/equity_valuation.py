"""IB-Grade Equity Valuation Workbook — 10-sheet interconnected model.

Sheets: Assumptions, Income Statement, Balance Sheet, Cash Flow,
Beta & WACC, DCF, Comps, Valuation Summary, Sensitivity, Model Checks.
"""

from __future__ import annotations

import json
import logging

from rag.tools.financial_modeling._fm_helpers import (
    EXCEL_AVAILABLE,
    _auto_width,
    _save_excel,
    _style_check_fail,
    _style_check_pass,
    _style_formula,
    _style_header,
    _style_input,
    _style_result,
    _style_section_header,
)

if EXCEL_AVAILABLE:
    import openpyxl
    from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ── Cross-sheet cell reference map ──────────────────────────────────────────
# Centralised constants to prevent cross-sheet formula typos.

_R = {
    # Assumptions sheet — Market Data block
    "A_PRICE": "Assumptions!$C$5",
    "A_SHARES": "Assumptions!$C$6",
    "A_DILUTED": "Assumptions!$C$7",
    "A_MCAP": "Assumptions!$C$8",
    "A_HIGH52": "Assumptions!$C$9",
    "A_LOW52": "Assumptions!$C$10",
    "A_DIV": "Assumptions!$C$11",
    # Assumptions — WACC block
    "A_RF": "Assumptions!$C$14",
    "A_ERP": "Assumptions!$C$15",
    "A_BETA": "Assumptions!$C$16",
    "A_KE": "Assumptions!$C$17",
    "A_KD": "Assumptions!$C$18",
    "A_TAX": "Assumptions!$C$19",
    "A_WD": "Assumptions!$C$20",
    "A_WE": "Assumptions!$C$21",
    "A_WACC": "Assumptions!$C$22",
    # Assumptions — Projection rows (row 26 = header, 27-34 = data)
    "A_PROJ_ROW_START": 27,  # first data row in projection block
    "A_PROJ_REV_GR": 27,
    "A_PROJ_GM": 28,
    "A_PROJ_RD": 29,
    "A_PROJ_SGA": 30,
    "A_PROJ_TAX": 31,
    "A_PROJ_DA": 32,
    "A_PROJ_CAPEX": 33,
    "A_PROJ_NWC": 34,
    # Assumptions — Terminal value
    "A_TG": "Assumptions!$C$37",
    "A_EXIT_MULT": "Assumptions!$C$38",
    # Income Statement rows
    "IS_REVENUE": 4,
    "IS_REV_GR": 5,
    "IS_COGS": 6,
    "IS_GP": 7,
    "IS_GM": 8,
    "IS_RD": 9,
    "IS_SGA": 10,
    "IS_OPEX": 11,
    "IS_EBIT": 12,
    "IS_EBIT_M": 13,
    "IS_OTHER": 14,
    "IS_PBT": 15,
    "IS_TAX": 16,
    "IS_ETR": 17,
    "IS_NI": 18,
    "IS_NI_M": 19,
    "IS_EBITDA": 20,
    "IS_DA": 21,
}

# Number of projection years = 5
_N_PROJ = 5


# ── Sheet builders ──────────────────────────────────────────────────────────


def _build_assumptions_sheet(
    wb,
    company_name,
    ticker,
    currency,
    current_price,
    shares_outstanding,
    diluted_shares,
    high_52w,
    low_52w,
    annual_dividend,
    beta_raw,
    risk_free_rate,
    equity_risk_premium,
    cost_of_debt,
    tax_rate,
    debt_weight,
    projection_years,
    terminal_growth,
    exit_ev_ebitda,
):
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = "1C2030"

    # Title
    ws["A1"] = f"Equity Valuation — {company_name} ({ticker})"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws["A2"] = f"Currency: {currency}"

    # ── Market Data ──
    _style_section_header(ws.cell(row=4, column=1, value="Market Data"))
    labels = [
        ("Share Price", current_price),
        ("Shares Outstanding (M)", shares_outstanding),
        ("Diluted Shares (M)", diluted_shares),
        ("Market Cap (M)", None),  # formula
        ("52-Week High", high_52w),
        ("52-Week Low", low_52w),
        ("Annual Dividend / Share", annual_dividend),
    ]
    for i, (lbl, val) in enumerate(labels, start=5):
        ws.cell(row=i, column=2, value=lbl)
        c = ws.cell(row=i, column=3)
        if lbl == "Market Cap (M)":
            c.value = "=C5*C6"
            _style_formula(c)
        else:
            c.value = val
            _style_input(c)

    # ── WACC Components ──
    _style_section_header(ws.cell(row=13, column=1, value="Cost of Capital"))
    equity_weight = round(1 - debt_weight, 6)
    ke = round(risk_free_rate + beta_raw * equity_risk_premium, 6)
    wacc_items = [
        ("Risk-Free Rate (Rf)", risk_free_rate),
        ("Equity Risk Premium", equity_risk_premium),
        ("Beta (selected)", beta_raw),
        ("Cost of Equity (Ke)", None),  # formula
        ("Cost of Debt (Kd)", cost_of_debt),
        ("Tax Rate", tax_rate),
        ("Debt Weight (Wd)", debt_weight),
        ("Equity Weight (We)", None),  # formula
        ("WACC", None),  # formula
    ]
    for i, (lbl, val) in enumerate(wacc_items, start=14):
        ws.cell(row=i, column=2, value=lbl)
        c = ws.cell(row=i, column=3)
        if lbl == "Cost of Equity (Ke)":
            c.value = "=C14+C16*C15"
            _style_formula(c)
        elif lbl == "Equity Weight (We)":
            c.value = "=1-C20"
            _style_formula(c)
        elif lbl == "WACC":
            c.value = "=C21*C17+C20*C18*(1-C19)"
            _style_formula(c)
        else:
            c.value = val
            _style_input(c)

    # Format as percentages
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    for r in (14, 15, 17, 18, 19, 20, 21, 22):
        ws.cell(row=r, column=3).number_format = FORMAT_PERCENTAGE_00

    # ── Projection Assumptions ──
    _style_section_header(ws.cell(row=25, column=1, value="Projection Assumptions"))
    proj_labels = [
        "Year",
        "Revenue Growth",
        "Gross Margin",
        "R&D % Revenue",
        "SG&A % Revenue",
        "Tax Rate",
        "D&A % Revenue",
        "CapEx % Revenue",
        "ΔNWC % Revenue",
    ]
    for i, lbl in enumerate(proj_labels):
        c = ws.cell(row=26 + i, column=2, value=lbl)
        if i == 0:
            _style_header(c)

    for j, yr in enumerate(projection_years):
        col = j + 3  # C, D, E, F, G
        ws.cell(row=26, column=col, value=yr.get("year", j + 1))
        _style_header(ws.cell(row=26, column=col))

        mapping = [
            ("rev_growth", _R["A_PROJ_REV_GR"]),
            ("gross_margin", _R["A_PROJ_GM"]),
            ("rd_pct", _R["A_PROJ_RD"]),
            ("sga_pct", _R["A_PROJ_SGA"]),
            ("tax_rate", _R["A_PROJ_TAX"]),
            ("da_pct", _R["A_PROJ_DA"]),
            ("capex_pct", _R["A_PROJ_CAPEX"]),
            ("nwc_pct", _R["A_PROJ_NWC"]),
        ]
        for k, (key, _row_id) in enumerate(mapping):
            c = ws.cell(row=27 + k, column=col, value=yr.get(key, 0))
            _style_input(c)
            c.number_format = FORMAT_PERCENTAGE_00

    # ── Terminal Value ──
    _style_section_header(ws.cell(row=36, column=1, value="Terminal Value"))
    ws.cell(row=37, column=2, value="Terminal Growth Rate")
    _style_input(ws.cell(row=37, column=3, value=terminal_growth))
    ws.cell(row=37, column=3).number_format = FORMAT_PERCENTAGE_00
    ws.cell(row=38, column=2, value="Exit EV/EBITDA Multiple")
    _style_input(ws.cell(row=38, column=3, value=exit_ev_ebitda))

    _auto_width(ws)
    return ws


def _build_income_statement_sheet(wb, historical_income, projection_years, n_hist):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_properties.tabColor = "2196F3"

    # Row labels
    row_labels = [
        ("Revenue", _R["IS_REVENUE"]),
        ("Revenue Growth %", _R["IS_REV_GR"]),
        ("COGS", _R["IS_COGS"]),
        ("Gross Profit", _R["IS_GP"]),
        ("Gross Margin %", _R["IS_GM"]),
        ("R&D", _R["IS_RD"]),
        ("SG&A", _R["IS_SGA"]),
        ("Total OpEx", _R["IS_OPEX"]),
        ("EBIT", _R["IS_EBIT"]),
        ("EBIT Margin %", _R["IS_EBIT_M"]),
        ("Other Income/Expense", _R["IS_OTHER"]),
        ("Pre-Tax Income", _R["IS_PBT"]),
        ("Tax", _R["IS_TAX"]),
        ("ETR %", _R["IS_ETR"]),
        ("Net Income", _R["IS_NI"]),
        ("Net Margin %", _R["IS_NI_M"]),
        ("EBITDA", _R["IS_EBITDA"]),
        ("D&A", _R["IS_DA"]),
    ]

    # Header row
    _style_header(ws.cell(row=3, column=1, value="Item"))
    for i, lbl_row in enumerate(row_labels):
        ws.cell(row=lbl_row[1], column=1, value=lbl_row[0])

    # ── Historical columns ──
    for j, hist in enumerate(historical_income):
        col = j + 2  # B, C, D, ...
        cl = get_column_letter(col)
        yr_label = hist.get("year", f"FY-{n_hist - j}")
        _style_header(ws.cell(row=3, column=col, value=yr_label))

        rev = hist.get("revenue", 0)
        cogs = hist.get("cogs", 0)
        gp = hist.get("gross_profit", rev - cogs)
        rd = hist.get("rd", 0)
        sga = hist.get("sga", 0)
        opex = hist.get("total_opex", rd + sga)
        ebit = hist.get("ebit", gp - opex)
        other = hist.get("other_income", 0)
        pbt = hist.get("pbt", ebit + other)
        tax = hist.get("tax", 0)
        ni = hist.get("net_income", pbt - tax)
        da = hist.get("da", 0)
        ebitda = hist.get("ebitda", ebit + da)

        values = {
            _R["IS_REVENUE"]: rev,
            _R["IS_COGS"]: cogs,
            _R["IS_GP"]: gp,
            _R["IS_RD"]: rd,
            _R["IS_SGA"]: sga,
            _R["IS_OPEX"]: opex,
            _R["IS_EBIT"]: ebit,
            _R["IS_OTHER"]: other,
            _R["IS_PBT"]: pbt,
            _R["IS_TAX"]: tax,
            _R["IS_NI"]: ni,
            _R["IS_EBITDA"]: ebitda,
            _R["IS_DA"]: da,
        }
        for row, val in values.items():
            _style_input(ws.cell(row=row, column=col, value=val))

        # Derived % rows (formulas even for historical for consistency)
        prev_col = get_column_letter(col - 1) if j > 0 else None
        if j > 0:
            ws.cell(row=_R["IS_REV_GR"], column=col,
                    value=f"=({cl}{_R['IS_REVENUE']}-{prev_col}{_R['IS_REVENUE']})/{prev_col}{_R['IS_REVENUE']}")
        else:
            ws.cell(row=_R["IS_REV_GR"], column=col, value="")
        _style_formula(ws.cell(row=_R["IS_GM"], column=col,
                               value=f"=IF({cl}{_R['IS_REVENUE']}<>0,{cl}{_R['IS_GP']}/{cl}{_R['IS_REVENUE']},0)"))
        _style_formula(ws.cell(row=_R["IS_EBIT_M"], column=col,
                               value=f"=IF({cl}{_R['IS_REVENUE']}<>0,{cl}{_R['IS_EBIT']}/{cl}{_R['IS_REVENUE']},0)"))
        _style_formula(ws.cell(row=_R["IS_ETR"], column=col,
                               value=f"=IF({cl}{_R['IS_PBT']}<>0,{cl}{_R['IS_TAX']}/{cl}{_R['IS_PBT']},0)"))
        _style_formula(ws.cell(row=_R["IS_NI_M"], column=col,
                               value=f"=IF({cl}{_R['IS_REVENUE']}<>0,{cl}{_R['IS_NI']}/{cl}{_R['IS_REVENUE']},0)"))

    # ── Projected columns ──
    last_hist_col = n_hist + 1  # 1-indexed column of last historical year
    for j, yr in enumerate(projection_years):
        col = n_hist + 2 + j
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        # Assumption column for this projection year
        a_cl = get_column_letter(j + 3)  # Assumptions projection starts at col C

        yr_label = yr.get("year", f"FY+{j + 1}")
        _style_header(ws.cell(row=3, column=col, value=yr_label))

        # Revenue = prev_revenue * (1 + growth)
        _style_formula(ws.cell(
            row=_R["IS_REVENUE"], column=col,
            value=f"={prev_cl}{_R['IS_REVENUE']}*(1+Assumptions!{a_cl}{_R['A_PROJ_REV_GR']})"))

        # Revenue Growth %
        _style_formula(ws.cell(
            row=_R["IS_REV_GR"], column=col,
            value=f"=Assumptions!{a_cl}{_R['A_PROJ_REV_GR']}"))

        # COGS = Revenue * (1 - GM)
        _style_formula(ws.cell(
            row=_R["IS_COGS"], column=col,
            value=f"={cl}{_R['IS_REVENUE']}*(1-Assumptions!{a_cl}{_R['A_PROJ_GM']})"))

        # Gross Profit = Revenue - COGS
        _style_formula(ws.cell(
            row=_R["IS_GP"], column=col,
            value=f"={cl}{_R['IS_REVENUE']}-{cl}{_R['IS_COGS']}"))

        # GM%
        _style_formula(ws.cell(
            row=_R["IS_GM"], column=col,
            value=f"=Assumptions!{a_cl}{_R['A_PROJ_GM']}"))

        # R&D = Revenue * R&D%
        _style_formula(ws.cell(
            row=_R["IS_RD"], column=col,
            value=f"={cl}{_R['IS_REVENUE']}*Assumptions!{a_cl}{_R['A_PROJ_RD']}"))

        # SG&A = Revenue * SGA%
        _style_formula(ws.cell(
            row=_R["IS_SGA"], column=col,
            value=f"={cl}{_R['IS_REVENUE']}*Assumptions!{a_cl}{_R['A_PROJ_SGA']}"))

        # Total OpEx = R&D + SG&A
        _style_formula(ws.cell(
            row=_R["IS_OPEX"], column=col,
            value=f"={cl}{_R['IS_RD']}+{cl}{_R['IS_SGA']}"))

        # EBIT = GP - OpEx
        _style_formula(ws.cell(
            row=_R["IS_EBIT"], column=col,
            value=f"={cl}{_R['IS_GP']}-{cl}{_R['IS_OPEX']}"))

        # EBIT Margin
        _style_formula(ws.cell(
            row=_R["IS_EBIT_M"], column=col,
            value=f"=IF({cl}{_R['IS_REVENUE']}<>0,{cl}{_R['IS_EBIT']}/{cl}{_R['IS_REVENUE']},0)"))

        # Other Income = 0
        ws.cell(row=_R["IS_OTHER"], column=col, value=0)

        # PBT = EBIT + Other
        _style_formula(ws.cell(
            row=_R["IS_PBT"], column=col,
            value=f"={cl}{_R['IS_EBIT']}+{cl}{_R['IS_OTHER']}"))

        # Tax = PBT * tax_rate
        _style_formula(ws.cell(
            row=_R["IS_TAX"], column=col,
            value=f"=MAX(0,{cl}{_R['IS_PBT']}*Assumptions!{a_cl}{_R['A_PROJ_TAX']})"))

        # ETR%
        _style_formula(ws.cell(
            row=_R["IS_ETR"], column=col,
            value=f"=IF({cl}{_R['IS_PBT']}<>0,{cl}{_R['IS_TAX']}/{cl}{_R['IS_PBT']},0)"))

        # Net Income
        _style_result(ws.cell(
            row=_R["IS_NI"], column=col,
            value=f"={cl}{_R['IS_PBT']}-{cl}{_R['IS_TAX']}"))

        # NI Margin
        _style_formula(ws.cell(
            row=_R["IS_NI_M"], column=col,
            value=f"=IF({cl}{_R['IS_REVENUE']}<>0,{cl}{_R['IS_NI']}/{cl}{_R['IS_REVENUE']},0)"))

        # D&A = Revenue * DA%
        _style_formula(ws.cell(
            row=_R["IS_DA"], column=col,
            value=f"={cl}{_R['IS_REVENUE']}*Assumptions!{a_cl}{_R['A_PROJ_DA']}"))

        # EBITDA = EBIT + D&A
        _style_formula(ws.cell(
            row=_R["IS_EBITDA"], column=col,
            value=f"={cl}{_R['IS_EBIT']}+{cl}{_R['IS_DA']}"))

    # Percentage format for margin rows
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    total_cols = n_hist + 1 + _N_PROJ
    for row in (_R["IS_REV_GR"], _R["IS_GM"], _R["IS_EBIT_M"], _R["IS_ETR"], _R["IS_NI_M"]):
        for col in range(2, total_cols + 1):
            ws.cell(row=row, column=col).number_format = FORMAT_PERCENTAGE_00

    _auto_width(ws)
    return ws


def _build_balance_sheet_sheet(wb, historical_balance_sheet):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_properties.tabColor = "4CAF50"

    if not historical_balance_sheet:
        ws["A1"] = "No balance sheet data provided"
        return ws

    row_labels = [
        (5, "Cash & Equivalents"),
        (6, "Short-Term Investments"),
        (7, "Accounts Receivable"),
        (8, "Inventory"),
        (9, "Other Current Assets"),
        (10, "Total Current Assets"),
        (11, ""),
        (12, "Long-Term Investments"),
        (13, "PP&E, Net"),
        (14, "Other Non-Current Assets"),
        (15, "Total Assets"),
        (16, ""),
        (17, "Accounts Payable"),
        (18, "Short-Term Debt"),
        (19, "Other Current Liabilities"),
        (20, "Total Current Liabilities"),
        (21, ""),
        (22, "Long-Term Debt"),
        (23, "Other Non-Current Liabilities"),
        (24, "Total Liabilities"),
        (25, ""),
        (26, "Total Equity"),
        (27, "Total Liabilities & Equity"),
        (28, ""),
        (29, "Balance Check"),
    ]

    _style_section_header(ws.cell(row=3, column=1, value="Balance Sheet"))
    _style_header(ws.cell(row=4, column=1, value="Item"))

    for row, lbl in row_labels:
        ws.cell(row=row, column=1, value=lbl)

    for j, bs in enumerate(historical_balance_sheet):
        col = j + 2
        cl = get_column_letter(col)
        yr_label = bs.get("year", f"FY-{len(historical_balance_sheet) - j}")
        _style_header(ws.cell(row=4, column=col, value=yr_label))

        direct = {
            5: bs.get("cash", 0),
            6: bs.get("st_investments", 0),
            7: bs.get("accounts_receivable", 0),
            8: bs.get("inventory", 0),
            9: bs.get("other_current_assets", 0),
            12: bs.get("lt_investments", 0),
            13: bs.get("ppe_net", 0),
            14: bs.get("other_noncurrent_assets", 0),
            17: bs.get("accounts_payable", 0),
            18: bs.get("st_debt", 0),
            19: bs.get("other_current_liabilities", 0),
            22: bs.get("lt_debt", 0),
            23: bs.get("other_noncurrent_liabilities", 0),
            26: bs.get("total_equity", 0),
        }
        for row, val in direct.items():
            _style_input(ws.cell(row=row, column=col, value=val))

        # Formulas for totals
        _style_formula(ws.cell(row=10, column=col, value=f"=SUM({cl}5:{cl}9)"))
        _style_formula(ws.cell(row=15, column=col, value=f"={cl}10+{cl}12+{cl}13+{cl}14"))
        _style_formula(ws.cell(row=20, column=col, value=f"=SUM({cl}17:{cl}19)"))
        _style_formula(ws.cell(row=24, column=col, value=f"={cl}20+{cl}22+{cl}23"))
        _style_formula(ws.cell(row=27, column=col, value=f"={cl}24+{cl}26"))

        # Balance check
        c = ws.cell(row=29, column=col,
                     value=f'=IF(ABS({cl}15-{cl}27)<0.01,"\u2705 Balanced","\u274c Unbalanced")')
        _style_formula(c)

    _auto_width(ws)
    return ws


def _build_cash_flow_sheet(wb, historical_cash_flow, n_hist):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_properties.tabColor = "FF9800"

    if not historical_cash_flow:
        ws["A1"] = "No cash flow data provided"
        return ws

    _style_section_header(ws.cell(row=3, column=1, value="Cash Flow Statement"))
    _style_header(ws.cell(row=4, column=1, value="Item"))

    row_labels = [
        (5, "--- Operating Activities ---"),
        (6, "Net Income"),
        (7, "D&A"),
        (8, "Stock-Based Compensation"),
        (9, "Working Capital Changes"),
        (10, "Other Operating"),
        (11, "Cash from Operations (CFO)"),
        (12, ""),
        (13, "--- Investing Activities ---"),
        (14, "Capital Expenditures"),
        (15, "Other Investing"),
        (16, "Cash from Investing (CFI)"),
        (17, ""),
        (18, "--- Financing Activities ---"),
        (19, "Dividends Paid"),
        (20, "Share Buybacks"),
        (21, "Net Debt Issuance"),
        (22, "Other Financing"),
        (23, "Cash from Financing (CFF)"),
        (24, ""),
        (25, "Net Change in Cash"),
        (26, ""),
        (27, "Free Cash Flow (FCF)"),
        (28, "FCF Margin %"),
    ]

    for row, lbl in row_labels:
        ws.cell(row=row, column=1, value=lbl)

    for j, cf in enumerate(historical_cash_flow):
        col = j + 2
        cl = get_column_letter(col)
        yr_label = cf.get("year", f"FY-{len(historical_cash_flow) - j}")
        _style_header(ws.cell(row=4, column=col, value=yr_label))

        direct = {
            6: cf.get("net_income", 0),
            7: cf.get("da", 0),
            8: cf.get("sbc", 0),
            9: cf.get("wc_changes", 0),
            10: cf.get("other_operating", 0),
            14: cf.get("capex", 0),
            15: cf.get("other_investing", 0),
            19: cf.get("dividends", 0),
            20: cf.get("buybacks", 0),
            21: cf.get("net_debt_issuance", 0),
            22: cf.get("other_financing", 0),
        }
        for row, val in direct.items():
            _style_input(ws.cell(row=row, column=col, value=val))

        # Totals
        _style_formula(ws.cell(row=11, column=col, value=f"=SUM({cl}6:{cl}10)"))
        _style_formula(ws.cell(row=16, column=col, value=f"={cl}14+{cl}15"))
        _style_formula(ws.cell(row=23, column=col, value=f"=SUM({cl}19:{cl}22)"))
        _style_formula(ws.cell(row=25, column=col, value=f"={cl}11+{cl}16+{cl}23"))

        # FCF = CFO - CapEx (CapEx usually negative, so CFO + CapEx)
        _style_result(ws.cell(row=27, column=col, value=f"={cl}11+{cl}14"))

        # FCF Margin = FCF / Revenue (cross-sheet)
        is_cl = get_column_letter(col)  # same position in IS
        _style_formula(ws.cell(
            row=28, column=col,
            value=f"=IF('Income Statement'!{is_cl}{_R['IS_REVENUE']}<>0,"
                  f"{cl}27/'Income Statement'!{is_cl}{_R['IS_REVENUE']},0)"))

    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    for col in range(2, len(historical_cash_flow) + 2):
        ws.cell(row=28, column=col).number_format = FORMAT_PERCENTAGE_00

    _auto_width(ws)
    return ws


def _build_beta_wacc_sheet(
    wb,
    beta_raw,
    blume_beta,
    bottom_up_beta,
    fama_french_beta,
    market_implied_beta,
    risk_free_rate,
    equity_risk_premium,
    cost_of_debt,
    tax_rate,
    debt_weight,
):
    ws = wb.create_sheet("Beta & WACC")
    ws.sheet_properties.tabColor = "9C27B0"

    _style_section_header(ws.cell(row=1, column=1, value="Beta & WACC Sensitivity"))

    # Headers
    headers = ["Method", "Beta", "Ke", "WACC", "Note"]
    for i, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=3, column=i, value=h))

    equity_weight = 1 - debt_weight

    methods = [
        ("Raw Regression", beta_raw, "Observed beta from regression"),
        ("Blume Adjusted", blume_beta, "2/3 × Raw + 1/3 × 1.0"),
        ("Bottom-Up", bottom_up_beta, "Industry unlevered, re-levered"),
        ("Fama-French", fama_french_beta, "3-factor model beta"),
        ("Market-Implied", market_implied_beta, "Reverse-engineered from price"),
    ]

    for i, (name, beta, note) in enumerate(methods, start=4):
        ws.cell(row=i, column=1, value=name)
        if beta is not None:
            _style_input(ws.cell(row=i, column=2, value=beta))
            ke = risk_free_rate + beta * equity_risk_premium
            wacc = equity_weight * ke + debt_weight * cost_of_debt * (1 - tax_rate)
            _style_formula(ws.cell(row=i, column=3, value=ke))
            _style_formula(ws.cell(row=i, column=4, value=wacc))
        else:
            ws.cell(row=i, column=2, value="N/A")
            ws.cell(row=i, column=3, value="N/A")
            ws.cell(row=i, column=4, value="N/A")
        ws.cell(row=i, column=5, value=note)

    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    for r in range(4, 9):
        for c in (3, 4):
            ws.cell(row=r, column=c).number_format = FORMAT_PERCENTAGE_00

    # CAPM breakdown
    _style_section_header(ws.cell(row=10, column=1, value="CAPM Breakdown"))
    ws.cell(row=11, column=1, value="Ke = Rf + \u03b2 \u00d7 ERP")
    ws.cell(row=11, column=2, value=f"= {risk_free_rate:.2%} + {beta_raw:.2f} \u00d7 {equity_risk_premium:.2%}")
    ke_selected = risk_free_rate + beta_raw * equity_risk_premium
    _style_result(ws.cell(row=11, column=4, value=ke_selected))
    ws.cell(row=11, column=4).number_format = FORMAT_PERCENTAGE_00

    # WACC breakdown
    _style_section_header(ws.cell(row=13, column=1, value="WACC Breakdown"))
    ws.cell(row=14, column=1, value="WACC = We\u00d7Ke + Wd\u00d7Kd\u00d7(1-T)")
    wacc_selected = equity_weight * ke_selected + debt_weight * cost_of_debt * (1 - tax_rate)
    _style_result(ws.cell(row=14, column=4, value=wacc_selected))
    ws.cell(row=14, column=4).number_format = FORMAT_PERCENTAGE_00

    _auto_width(ws)
    return ws


def _build_dcf_sheet(
    wb,
    n_hist,
    projection_years,
    net_debt,
    shares_outstanding,
    tax_rate,
    wacc,
    terminal_growth,
    exit_ev_ebitda,
):
    ws = wb.create_sheet("DCF")
    ws.sheet_properties.tabColor = "009688"

    _style_section_header(ws.cell(row=1, column=1, value="Discounted Cash Flow Valuation"))

    # Headers
    headers_row = 3
    _style_header(ws.cell(row=headers_row, column=1, value="Item"))
    for j in range(_N_PROJ):
        yr = projection_years[j].get("year", f"FY+{j + 1}")
        _style_header(ws.cell(row=headers_row, column=j + 2, value=yr))

    # IS column offset: projected columns start at (n_hist + 2)
    is_col_start = n_hist + 2

    # Build FCFF from IS cross-references
    row_labels = [
        (4, "Revenue"),
        (5, "EBIT"),
        (6, "Tax Rate"),
        (7, "NOPAT [EBIT×(1-T)]"),
        (8, "D&A"),
        (9, "CapEx"),
        (10, "\u0394NWC"),
        (11, "FCFF"),
        (12, ""),
        (13, "Discount Factor"),
        (14, "PV of FCFF"),
    ]
    for row, lbl in row_labels:
        ws.cell(row=row, column=1, value=lbl)

    for j in range(_N_PROJ):
        col = j + 2
        cl = get_column_letter(col)
        is_cl = get_column_letter(is_col_start + j)
        a_cl = get_column_letter(j + 3)  # Assumptions projection column

        # Revenue (cross-ref IS)
        _style_formula(ws.cell(
            row=4, column=col,
            value=f"='Income Statement'!{is_cl}{_R['IS_REVENUE']}"))

        # EBIT (cross-ref IS)
        _style_formula(ws.cell(
            row=5, column=col,
            value=f"='Income Statement'!{is_cl}{_R['IS_EBIT']}"))

        # Tax Rate (from assumptions)
        _style_formula(ws.cell(
            row=6, column=col,
            value=f"=Assumptions!{a_cl}{_R['A_PROJ_TAX']}"))

        # NOPAT = EBIT * (1 - Tax)
        _style_formula(ws.cell(
            row=7, column=col,
            value=f"={cl}5*(1-{cl}6)"))

        # D&A (cross-ref IS)
        _style_formula(ws.cell(
            row=8, column=col,
            value=f"='Income Statement'!{is_cl}{_R['IS_DA']}"))

        # CapEx = Revenue * CapEx%
        _style_formula(ws.cell(
            row=9, column=col,
            value=f"={cl}4*Assumptions!{a_cl}{_R['A_PROJ_CAPEX']}"))

        # ΔNWC = Revenue * NWC%
        _style_formula(ws.cell(
            row=10, column=col,
            value=f"={cl}4*Assumptions!{a_cl}{_R['A_PROJ_NWC']}"))

        # FCFF = NOPAT + D&A - CapEx - ΔNWC
        _style_result(ws.cell(
            row=11, column=col,
            value=f"={cl}7+{cl}8-{cl}9-{cl}10"))

        # Discount factor = (1+WACC)^year
        _style_formula(ws.cell(
            row=13, column=col,
            value=f"=(1+{_R['A_WACC']})^{j + 1}"))

        # PV FCFF
        _style_formula(ws.cell(
            row=14, column=col,
            value=f"={cl}11/{cl}13"))

    # Format tax row as pct
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    for j in range(_N_PROJ):
        ws.cell(row=6, column=j + 2).number_format = FORMAT_PERCENTAGE_00

    # ── Terminal Value section ──
    _style_section_header(ws.cell(row=16, column=1, value="Terminal Value"))

    last_fcff_cl = get_column_letter(_N_PROJ + 1)

    tv_rows = [
        (17, "Sum of PV FCFFs"),
        (18, ""),
        (19, "--- Gordon Growth Method ---"),
        (20, "Terminal FCFF"),
        (21, "Terminal Value (Gordon)"),
        (22, "PV of Terminal Value (Gordon)"),
        (23, ""),
        (24, "--- Exit Multiple Method ---"),
        (25, "Terminal EBITDA"),
        (26, "Terminal Value (Exit)"),
        (27, "PV of Terminal Value (Exit)"),
        (28, ""),
        (29, "--- Equity Bridge ---"),
        (30, "EV (Gordon)"),
        (31, "EV (Exit Multiple)"),
        (32, "Blended EV"),
        (33, "Less: Net Debt"),
        (34, "Equity Value"),
        (35, "Shares Outstanding (M)"),
        (36, "Implied Price (Gordon)"),
        (37, "Implied Price (Exit)"),
        (38, "Implied Price (Blended)"),
        (39, ""),
        (40, "Current Price"),
        (41, "Upside/Downside (Blended)"),
    ]
    for row, lbl in tv_rows:
        ws.cell(row=row, column=1, value=lbl)

    # Sum PV FCFFs
    first_cl = get_column_letter(2)
    last_pv_cl = get_column_letter(_N_PROJ + 1)
    _style_formula(ws.cell(row=17, column=2, value=f"=SUM({first_cl}14:{last_pv_cl}14)"))

    # Gordon Growth
    _style_formula(ws.cell(row=20, column=2,
                           value=f"={last_fcff_cl}11*(1+{_R['A_TG']})"))
    _style_formula(ws.cell(row=21, column=2,
                           value=f"=B20/({_R['A_WACC']}-{_R['A_TG']})"))
    _style_formula(ws.cell(row=22, column=2,
                           value=f"=B21/(1+{_R['A_WACC']})^{_N_PROJ}"))

    # Exit Multiple
    last_ebitda_is_cl = get_column_letter(is_col_start + _N_PROJ - 1)
    _style_formula(ws.cell(row=25, column=2,
                           value=f"='Income Statement'!{last_ebitda_is_cl}{_R['IS_EBITDA']}"))
    _style_formula(ws.cell(row=26, column=2,
                           value=f"=B25*{_R['A_EXIT_MULT']}"))
    _style_formula(ws.cell(row=27, column=2,
                           value=f"=B26/(1+{_R['A_WACC']})^{_N_PROJ}"))

    # Equity bridge
    _style_formula(ws.cell(row=30, column=2, value="=B17+B22"))  # EV Gordon
    _style_formula(ws.cell(row=31, column=2, value="=B17+B27"))  # EV Exit
    _style_formula(ws.cell(row=32, column=2, value="=(B30+B31)/2"))  # Blended EV
    _style_input(ws.cell(row=33, column=2, value=net_debt))
    _style_formula(ws.cell(row=34, column=2, value="=B32-B33"))  # Equity value
    ws.cell(row=35, column=2, value=f"={_R['A_DILUTED']}")
    _style_formula(ws.cell(row=35, column=2))
    _style_result(ws.cell(row=36, column=2, value="=(B30-B33)/B35"))  # Gordon price
    _style_result(ws.cell(row=37, column=2, value="=(B31-B33)/B35"))  # Exit price
    _style_result(ws.cell(row=38, column=2, value="=B34/B35"))  # Blended price

    ws.cell(row=40, column=2, value=f"={_R['A_PRICE']}")
    _style_formula(ws.cell(row=41, column=2, value="=(B38-B40)/B40"))
    ws.cell(row=41, column=2).number_format = FORMAT_PERCENTAGE_00

    _auto_width(ws)
    return ws


def _build_comps_sheet(wb, peers, n_hist):
    ws = wb.create_sheet("Comps")
    ws.sheet_properties.tabColor = "FFC107"

    _style_section_header(ws.cell(row=1, column=1, value="Comparable Companies Analysis"))

    headers = [
        "Company", "Ticker", "EV (M)", "EV/Revenue",
        "EV/EBITDA", "P/E", "Rev Growth %",
        "EBITDA Margin %", "NI Margin %",
    ]
    for i, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=3, column=i, value=h))

    if not peers:
        ws.cell(row=4, column=1, value="No peer data provided")
        _auto_width(ws)
        return ws

    n_peers = len(peers)
    for j, peer in enumerate(peers):
        row = j + 4
        ws.cell(row=row, column=1, value=peer.get("name", ""))
        ws.cell(row=row, column=2, value=peer.get("ticker", ""))
        _style_input(ws.cell(row=row, column=3, value=peer.get("ev", 0)))
        _style_input(ws.cell(row=row, column=4, value=peer.get("ev_revenue", 0)))
        _style_input(ws.cell(row=row, column=5, value=peer.get("ev_ebitda", 0)))
        _style_input(ws.cell(row=row, column=6, value=peer.get("pe", 0)))
        _style_input(ws.cell(row=row, column=7, value=peer.get("rev_growth", 0)))
        _style_input(ws.cell(row=row, column=8, value=peer.get("ebitda_margin", 0)))
        _style_input(ws.cell(row=row, column=9, value=peer.get("ni_margin", 0)))

    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    for j in range(n_peers):
        for c in (7, 8, 9):
            ws.cell(row=j + 4, column=c).number_format = FORMAT_PERCENTAGE_00

    # Median row
    med_row = n_peers + 4
    _style_section_header(ws.cell(row=med_row, column=1, value="Peer Median"))
    first_data = 4
    last_data = n_peers + 3
    for c in range(3, 10):
        cl = get_column_letter(c)
        _style_result(ws.cell(
            row=med_row, column=c,
            value=f"=MEDIAN({cl}{first_data}:{cl}{last_data})"))

    # Subject company row
    subj_row = med_row + 2
    _style_section_header(ws.cell(row=subj_row, column=1, value="Subject Company"))
    # EV from Assumptions market cap (simplified: market_cap + net_debt proxy)
    ws.cell(row=subj_row, column=3, value=f"={_R['A_MCAP']}")
    _style_formula(ws.cell(row=subj_row, column=3))

    # Last historical IS column for subject company ratios
    is_cl = get_column_letter(n_hist + 1)

    # EV/Revenue = EV / Revenue
    _style_formula(ws.cell(
        row=subj_row, column=4,
        value=f"=IF('Income Statement'!{is_cl}{_R['IS_REVENUE']}<>0,"
              f"{get_column_letter(3)}{subj_row}/'Income Statement'!{is_cl}{_R['IS_REVENUE']},0)"))

    # EV/EBITDA
    _style_formula(ws.cell(
        row=subj_row, column=5,
        value=f"=IF('Income Statement'!{is_cl}{_R['IS_EBITDA']}<>0,"
              f"{get_column_letter(3)}{subj_row}/'Income Statement'!{is_cl}{_R['IS_EBITDA']},0)"))

    # P/E
    _style_formula(ws.cell(
        row=subj_row, column=6,
        value=f"=IF('Income Statement'!{is_cl}{_R['IS_NI']}<>0,"
              f"{_R['A_MCAP']}/'Income Statement'!{is_cl}{_R['IS_NI']},0)"))

    # Comps-implied prices
    impl_row = subj_row + 2
    _style_section_header(ws.cell(row=impl_row, column=1, value="Comps-Implied Prices"))

    # Implied from EV/EBITDA
    ws.cell(row=impl_row + 1, column=1, value="Implied Price (EV/EBITDA)")
    # = (Median EV/EBITDA × Subject EBITDA - Net Debt) / Shares
    last_is_proj_cl = get_column_letter(n_hist + 1 + _N_PROJ)
    med_ev_ebitda = f"{get_column_letter(5)}{med_row}"
    _style_result(ws.cell(
        row=impl_row + 1, column=2,
        value=f"=({med_ev_ebitda}*'Income Statement'!{last_is_proj_cl}{_R['IS_EBITDA']}"
              f"-DCF!B33)/{_R['A_DILUTED']}"))

    # Implied from P/E
    ws.cell(row=impl_row + 2, column=1, value="Implied Price (P/E)")
    med_pe = f"{get_column_letter(6)}{med_row}"
    _style_result(ws.cell(
        row=impl_row + 2, column=2,
        value=f"={med_pe}*'Income Statement'!{last_is_proj_cl}{_R['IS_NI']}/{_R['A_DILUTED']}"))

    _auto_width(ws)
    return ws, med_row, impl_row


def _build_valuation_summary_sheet(wb, current_price, high_52w, low_52w, bull_price, bear_price,
                                   bull_description, bear_description, risks, catalysts,
                                   comps_med_row, comps_impl_row):
    ws = wb.create_sheet("Valuation Summary")
    ws.sheet_properties.tabColor = "E91E63"

    _style_section_header(ws.cell(row=1, column=1, value="Valuation Summary — Football Field"))

    # Headers
    field_headers = ["Method", "Low", "Base", "High"]
    for i, h in enumerate(field_headers, start=1):
        _style_header(ws.cell(row=3, column=i, value=h))

    # Football field rows (Low = 85% of base, High = 115%)
    methods = [
        (4, "DCF (Gordon Growth)", "=DCF!B36*0.85", "=DCF!B36", "=DCF!B36*1.15"),
        (5, "DCF (Exit Multiple)", "=DCF!B37*0.85", "=DCF!B37", "=DCF!B37*1.15"),
        (6, "EV/EBITDA Comps",
         f"=Comps!B{comps_impl_row + 1}*0.85",
         f"=Comps!B{comps_impl_row + 1}",
         f"=Comps!B{comps_impl_row + 1}*1.15"),
        (7, "P/E Comps",
         f"=Comps!B{comps_impl_row + 2}*0.85",
         f"=Comps!B{comps_impl_row + 2}",
         f"=Comps!B{comps_impl_row + 2}*1.15"),
        (8, "52-Week Range",
         f"={_R['A_LOW52']}" if low_52w else "=0",
         f"={_R['A_PRICE']}",
         f"={_R['A_HIGH52']}" if high_52w else "=0"),
    ]
    for row, name, low, base, high in methods:
        ws.cell(row=row, column=1, value=name)
        _style_formula(ws.cell(row=row, column=2, value=low))
        _style_formula(ws.cell(row=row, column=3, value=base))
        _style_formula(ws.cell(row=row, column=4, value=high))

    # Current price
    ws.cell(row=9, column=1, value="Current Market Price")
    _style_input(ws.cell(row=9, column=3, value=f"={_R['A_PRICE']}"))

    # ── Weighted valuation ──
    _style_section_header(ws.cell(row=11, column=1, value="Weighted Valuation"))
    weight_rows = [
        (12, "Method", "Weight", "Price"),
        (13, "DCF (Blended)", 0.50, "=DCF!B38"),
        (14, "EV/EBITDA Comps", 0.25, f"=Comps!B{comps_impl_row + 1}"),
        (15, "P/E Comps", 0.25, f"=Comps!B{comps_impl_row + 2}"),
    ]
    for row, m, w, p in weight_rows:
        ws.cell(row=row, column=1, value=m)
        if row == 12:
            _style_header(ws.cell(row=row, column=1))
            _style_header(ws.cell(row=row, column=2, value=w))
            _style_header(ws.cell(row=row, column=3, value=p))
        else:
            _style_input(ws.cell(row=row, column=2, value=w))
            _style_formula(ws.cell(row=row, column=3, value=p))

    # Weighted price
    ws.cell(row=17, column=1, value="Weighted Target Price")
    _style_result(ws.cell(row=17, column=3,
                          value="=B13*C13+B14*C14+B15*C15"))

    ws.cell(row=18, column=1, value="Upside/Downside")
    _style_formula(ws.cell(row=18, column=3,
                           value=f"=(C17-{_R['A_PRICE']})/{_R['A_PRICE']}"))
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    ws.cell(row=18, column=3).number_format = FORMAT_PERCENTAGE_00

    # Recommendation
    ws.cell(row=20, column=1, value="Recommendation")
    _style_result(ws.cell(row=20, column=3,
                          value='=IF(C18>0.15,"BUY",IF(C18<-0.15,"SELL","HOLD"))'))

    # ── Scenario Analysis ──
    _style_section_header(ws.cell(row=22, column=1, value="Scenario Analysis"))
    _style_header(ws.cell(row=23, column=1, value="Scenario"))
    _style_header(ws.cell(row=23, column=2, value="Target Price"))
    _style_header(ws.cell(row=23, column=3, value="Description"))

    if bull_price is not None:
        ws.cell(row=24, column=1, value="Bull")
        _style_input(ws.cell(row=24, column=2, value=bull_price))
        ws.cell(row=24, column=3, value=bull_description)

    ws.cell(row=25, column=1, value="Base")
    _style_formula(ws.cell(row=25, column=2, value="=C17"))
    ws.cell(row=25, column=3, value="Weighted average of all methods")

    if bear_price is not None:
        ws.cell(row=26, column=1, value="Bear")
        _style_input(ws.cell(row=26, column=2, value=bear_price))
        ws.cell(row=26, column=3, value=bear_description)

    # ── Risks & Catalysts ──
    _style_section_header(ws.cell(row=28, column=1, value="Key Risks"))
    if risks:
        for i, line in enumerate(risks.split("\n")[:5]):
            ws.cell(row=29 + i, column=1, value=line.strip())

    _style_section_header(ws.cell(row=35, column=1, value="Catalysts"))
    if catalysts:
        for i, line in enumerate(catalysts.split("\n")[:5]):
            ws.cell(row=36 + i, column=1, value=line.strip())

    _auto_width(ws)
    return ws


def _build_sensitivity_sheet(
    wb,
    projection_years,
    historical_income,
    wacc,
    terminal_growth,
    net_debt,
    shares_outstanding,
    tax_rate,
):
    """Build WACC vs Terminal Growth sensitivity matrix (Python-computed values)."""
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "607D8B"

    _style_section_header(ws.cell(row=1, column=1, value="Sensitivity: WACC vs Terminal Growth"))

    # Compute base FCFF projections from inputs for sensitivity
    last_hist = historical_income[-1] if historical_income else {}
    base_rev = last_hist.get("revenue", 0)
    fcff_list = []
    rev = base_rev
    for yr in projection_years:
        rev = rev * (1 + yr.get("rev_growth", 0))
        gm = yr.get("gross_margin", 0.4)
        rd = yr.get("rd_pct", 0)
        sga = yr.get("sga_pct", 0)
        da_pct = yr.get("da_pct", 0)
        capex_pct = yr.get("capex_pct", 0)
        nwc_pct = yr.get("nwc_pct", 0)
        t = yr.get("tax_rate", tax_rate)

        gp = rev * gm
        opex = rev * (rd + sga)
        ebit = gp - opex
        nopat = ebit * (1 - t)
        da = rev * da_pct
        capex = rev * capex_pct
        nwc = rev * nwc_pct
        fcff = nopat + da - capex - nwc
        fcff_list.append(fcff)

    last_fcff = fcff_list[-1] if fcff_list else 0
    n = len(fcff_list)

    # WACC range
    wacc_steps = [round(wacc + delta, 4) for delta in [-0.03, -0.02, -0.01, 0, 0.01, 0.02, 0.03]]
    # TG range
    tg_steps = [round(terminal_growth + delta, 4) for delta in [-0.02, -0.01, 0, 0.01, 0.02]]

    # Headers
    _style_header(ws.cell(row=3, column=1, value="WACC \\ TG"))
    for ci, tg in enumerate(tg_steps):
        _style_header(ws.cell(row=3, column=ci + 2, value=f"{tg:.1%}"))

    for ri, w in enumerate(wacc_steps):
        _style_header(ws.cell(row=ri + 4, column=1, value=f"{w:.1%}"))
        for ci, tg in enumerate(tg_steps):
            if w <= tg:
                ws.cell(row=ri + 4, column=ci + 2, value="N/A")
                continue
            # PV of projection FCFFs
            pv_fcff = sum(f / (1 + w) ** (t + 1) for t, f in enumerate(fcff_list))
            tv = last_fcff * (1 + tg) / (w - tg)
            pv_tv = tv / (1 + w) ** n
            ev = pv_fcff + pv_tv
            eq = ev - net_debt
            price = round(eq / shares_outstanding, 2) if shares_outstanding else 0
            c = ws.cell(row=ri + 4, column=ci + 2, value=price)
            # Highlight base case
            if abs(w - wacc) < 0.0001 and abs(tg - terminal_growth) < 0.0001:
                _style_result(c)

    _auto_width(ws)
    return ws


def _build_model_checks_sheet(
    wb,
    wacc,
    terminal_growth,
    projection_years,
    historical_income,
    exit_ev_ebitda,
    debt_weight,
):
    ws = wb.create_sheet("Model Checks")
    ws.sheet_properties.tabColor = "F44336"

    _style_section_header(ws.cell(row=1, column=1, value="Model Integrity Checks"))

    headers = ["Check", "Status", "Value", "Threshold"]
    for i, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=3, column=i, value=h))

    checks = []

    # 1. WACC in reasonable range
    wacc_ok = 0.04 <= wacc <= 0.40
    checks.append(("WACC in range (4%–40%)", wacc_ok, f"{wacc:.2%}", "4%–40%"))

    # 2. Terminal growth < WACC
    tg_ok = terminal_growth < wacc
    checks.append(("Terminal Growth < WACC", tg_ok, f"{terminal_growth:.2%} < {wacc:.2%}", "TG < WACC"))

    # 3. Terminal value % of EV < 85% (estimate)
    # Compute rough TV%
    last_hist = historical_income[-1] if historical_income else {}
    base_rev = last_hist.get("revenue", 1)
    rev = base_rev
    fcff_list = []
    for yr in projection_years:
        rev = rev * (1 + yr.get("rev_growth", 0))
        gm = yr.get("gross_margin", 0.4)
        rd = yr.get("rd_pct", 0)
        sga = yr.get("sga_pct", 0)
        da_pct = yr.get("da_pct", 0)
        capex_pct = yr.get("capex_pct", 0)
        nwc_pct = yr.get("nwc_pct", 0)
        t = yr.get("tax_rate", 0.25)
        gp = rev * gm
        ebit = gp - rev * (rd + sga)
        nopat = ebit * (1 - t)
        fcff = nopat + rev * da_pct - rev * capex_pct - rev * nwc_pct
        fcff_list.append(fcff)

    last_fcff = fcff_list[-1] if fcff_list else 0
    n = len(fcff_list)
    pv_fcff = sum(f / (1 + wacc) ** (i + 1) for i, f in enumerate(fcff_list))
    tv_gordon = last_fcff * (1 + terminal_growth) / (wacc - terminal_growth) if wacc > terminal_growth else 0
    pv_tv = tv_gordon / (1 + wacc) ** n if n > 0 else 0
    ev = pv_fcff + pv_tv
    tv_pct = pv_tv / ev if ev > 0 else 0
    tv_ok = tv_pct < 0.85
    checks.append(("Terminal Value < 85% of EV", tv_ok, f"{tv_pct:.1%}", "< 85%"))

    # 4. Revenue growth reasonable (all years < 50%)
    rev_growths = [yr.get("rev_growth", 0) for yr in projection_years]
    rev_gr_ok = all(abs(g) < 0.50 for g in rev_growths)
    max_gr = max(abs(g) for g in rev_growths) if rev_growths else 0
    checks.append(("Revenue growth < 50%", rev_gr_ok, f"Max: {max_gr:.1%}", "< 50%"))

    # 5. Gross margins stable (all positive)
    gm_vals = [yr.get("gross_margin", 0) for yr in projection_years]
    gm_ok = all(0 < g <= 1 for g in gm_vals)
    checks.append(("Gross margins positive", gm_ok,
                    f"{min(gm_vals):.1%}–{max(gm_vals):.1%}" if gm_vals else "N/A",
                    "0%–100%"))

    # 6. All projected FCFFs positive
    fcf_ok = all(f > 0 for f in fcff_list) if fcff_list else False
    checks.append(("All projected FCFs positive", fcf_ok,
                    f"Min: {min(fcff_list):,.0f}" if fcff_list else "N/A", "> 0"))

    # 7. Weights sum to ~100%
    weights_ok = abs(debt_weight + (1 - debt_weight) - 1.0) < 0.001
    checks.append(("Capital weights sum to 100%", weights_ok,
                    f"Wd={debt_weight:.1%}, We={(1 - debt_weight):.1%}", "Sum = 100%"))

    # 8. Exit multiple reasonable (if provided)
    if exit_ev_ebitda is not None:
        exit_ok = 3 <= exit_ev_ebitda <= 50
        checks.append(("Exit EV/EBITDA in range", exit_ok,
                        f"{exit_ev_ebitda:.1f}x", "3x–50x"))

    passed = 0
    total = len(checks)
    for i, (check, ok, value, threshold) in enumerate(checks):
        row = i + 4
        ws.cell(row=row, column=1, value=check)
        status_str = "\u2705 PASS" if ok else "\u274c FAIL"
        c = ws.cell(row=row, column=2, value=status_str)
        if ok:
            _style_check_pass(c)
            passed += 1
        else:
            _style_check_fail(c)
        ws.cell(row=row, column=3, value=value)
        ws.cell(row=row, column=4, value=threshold)

    # Summary row
    summary_row = len(checks) + 5
    _style_section_header(ws.cell(row=summary_row, column=1, value="Summary"))
    c = ws.cell(row=summary_row, column=2, value=f"{passed}/{total} checks passed")
    if passed == total:
        _style_check_pass(c)
    else:
        _style_check_fail(c)

    _auto_width(ws)
    return ws, passed, total


# ── Main builder function ───────────────────────────────────────────────────


def build_equity_valuation_model(
    db,
    company_name: str,
    ticker: str,
    currency: str = "USD",
    # Market data
    current_price: float = 0,
    shares_outstanding: float = 1,
    diluted_shares: float = None,
    high_52w: float = None,
    low_52w: float = None,
    annual_dividend: float = 0.0,
    # Beta & WACC inputs
    beta_raw: float = 1.0,
    risk_free_rate: float = 0.04,
    equity_risk_premium: float = 0.06,
    cost_of_debt: float = 0.05,
    tax_rate: float = 0.25,
    debt_weight: float = 0.02,
    blume_beta: float = None,
    bottom_up_beta: float = None,
    fama_french_beta: float = None,
    market_implied_beta: float = None,
    # Historical IS (3-5yr list of dicts)
    historical_income: list = None,
    # Historical BS (3-5yr list of dicts)
    historical_balance_sheet: list = None,
    # Historical CF (3-5yr list of dicts)
    historical_cash_flow: list = None,
    # Projection assumptions (5yr list of dicts)
    projection_years: list = None,
    # Terminal value
    terminal_growth: float = 0.03,
    exit_ev_ebitda: float = None,
    # DCF specifics
    net_debt: float = None,
    # Comps (list of peer dicts)
    peers: list = None,
    # Scenario analysis
    bull_price: float = None,
    bull_description: str = "",
    bear_price: float = None,
    bear_description: str = "",
    risks: str = "",
    catalysts: str = "",
) -> str:
    """Build an IB-grade equity valuation workbook with 10 interconnected sheets."""
    if not EXCEL_AVAILABLE:
        return json.dumps({"error": "openpyxl not installed"})

    # ── Defaults & validation ──
    if historical_income is None or len(historical_income) == 0:
        return json.dumps({"error": "historical_income is required (3-5 years of data)"})

    if projection_years is None or len(projection_years) == 0:
        return json.dumps({"error": "projection_years is required (list of projection assumptions)"})

    if diluted_shares is None:
        diluted_shares = shares_outstanding

    if blume_beta is None:
        blume_beta = round(2 / 3 * beta_raw + 1 / 3, 4)

    if exit_ev_ebitda is None:
        exit_ev_ebitda = 12.0  # sensible default

    equity_weight = 1 - debt_weight
    ke = risk_free_rate + beta_raw * equity_risk_premium
    wacc = equity_weight * ke + debt_weight * cost_of_debt * (1 - tax_rate)

    if wacc <= terminal_growth:
        return json.dumps({"error": f"WACC ({wacc:.2%}) must exceed terminal growth ({terminal_growth:.2%})"})

    # Net debt: derive from last BS if not provided
    if net_debt is None:
        if historical_balance_sheet:
            last_bs = historical_balance_sheet[-1]
            total_debt = last_bs.get("st_debt", 0) + last_bs.get("lt_debt", 0)
            cash = last_bs.get("cash", 0) + last_bs.get("st_investments", 0)
            net_debt = total_debt - cash
        else:
            net_debt = 0

    n_hist = len(historical_income)

    # ── Build workbook ──
    wb = openpyxl.Workbook()

    # Sheet 1: Assumptions
    _build_assumptions_sheet(
        wb, company_name, ticker, currency, current_price,
        shares_outstanding, diluted_shares, high_52w, low_52w,
        annual_dividend, beta_raw, risk_free_rate, equity_risk_premium,
        cost_of_debt, tax_rate, debt_weight, projection_years,
        terminal_growth, exit_ev_ebitda,
    )

    # Sheet 2: Income Statement
    _build_income_statement_sheet(wb, historical_income, projection_years, n_hist)

    # Sheet 3: Balance Sheet
    _build_balance_sheet_sheet(wb, historical_balance_sheet)

    # Sheet 4: Cash Flow
    _build_cash_flow_sheet(wb, historical_cash_flow, n_hist)

    # Sheet 5: Beta & WACC
    _build_beta_wacc_sheet(
        wb, beta_raw, blume_beta, bottom_up_beta, fama_french_beta,
        market_implied_beta, risk_free_rate, equity_risk_premium,
        cost_of_debt, tax_rate, debt_weight,
    )

    # Sheet 6: DCF
    _build_dcf_sheet(
        wb, n_hist, projection_years, net_debt, shares_outstanding,
        tax_rate, wacc, terminal_growth, exit_ev_ebitda,
    )

    # Sheet 7: Comps
    comps_result = _build_comps_sheet(wb, peers, n_hist)
    if isinstance(comps_result, tuple):
        _, comps_med_row, comps_impl_row = comps_result
    else:
        comps_med_row, comps_impl_row = 5, 7  # fallback if no peers

    # Sheet 8: Valuation Summary
    _build_valuation_summary_sheet(
        wb, current_price, high_52w, low_52w, bull_price, bear_price,
        bull_description, bear_description, risks, catalysts,
        comps_med_row, comps_impl_row,
    )

    # Sheet 9: Sensitivity
    _build_sensitivity_sheet(
        wb, projection_years, historical_income, wacc, terminal_growth,
        net_debt, diluted_shares, tax_rate,
    )

    # Sheet 10: Model Checks
    _, checks_passed, checks_total = _build_model_checks_sheet(
        wb, wacc, terminal_growth, projection_years, historical_income,
        exit_ev_ebitda, debt_weight,
    )

    # ── Compute summary values (Python-side for JSON) ──
    last_hist = historical_income[-1]
    base_rev = last_hist.get("revenue", 0)
    rev = base_rev
    fcff_list = []
    for yr in projection_years:
        rev = rev * (1 + yr.get("rev_growth", 0))
        gm = yr.get("gross_margin", 0.4)
        rd = yr.get("rd_pct", 0)
        sga = yr.get("sga_pct", 0)
        da_pct = yr.get("da_pct", 0)
        capex_pct = yr.get("capex_pct", 0)
        nwc_pct = yr.get("nwc_pct", 0)
        t = yr.get("tax_rate", tax_rate)
        gp = rev * gm
        ebit = gp - rev * (rd + sga)
        nopat = ebit * (1 - t)
        fcff = nopat + rev * da_pct - rev * capex_pct - rev * nwc_pct
        fcff_list.append(fcff)

    last_fcff = fcff_list[-1] if fcff_list else 0
    n = len(fcff_list)
    pv_fcff_sum = sum(f / (1 + wacc) ** (i + 1) for i, f in enumerate(fcff_list))
    tv_gordon = last_fcff * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_tv_gordon = tv_gordon / (1 + wacc) ** n
    ev_gordon = pv_fcff_sum + pv_tv_gordon
    price_gordon = round((ev_gordon - net_debt) / diluted_shares, 2) if diluted_shares else 0

    # Exit multiple
    last_ebitda = fcff_list[-1] if fcff_list else 0  # rough proxy; ideally from IS
    # Better: compute EBITDA from last projected year
    rev_final = base_rev
    for yr in projection_years:
        rev_final *= (1 + yr.get("rev_growth", 0))
    last_yr = projection_years[-1]
    gm_final = last_yr.get("gross_margin", 0.4)
    rd_final = last_yr.get("rd_pct", 0)
    sga_final = last_yr.get("sga_pct", 0)
    da_final = last_yr.get("da_pct", 0)
    ebit_final = rev_final * gm_final - rev_final * (rd_final + sga_final)
    ebitda_final = ebit_final + rev_final * da_final
    tv_exit = ebitda_final * exit_ev_ebitda
    pv_tv_exit = tv_exit / (1 + wacc) ** n
    ev_exit = pv_fcff_sum + pv_tv_exit
    price_exit = round((ev_exit - net_debt) / diluted_shares, 2) if diluted_shares else 0

    blended_price = round((price_gordon + price_exit) / 2, 2)
    upside_pct = round((blended_price - current_price) / current_price * 100, 1) if current_price else 0
    recommendation = "BUY" if upside_pct > 15 else ("SELL" if upside_pct < -15 else "HOLD")

    # ── Save ──
    file_id = _save_excel(wb, f"IB_Valuation_{ticker}")
    download_url = f"/api/financial-modeling/download/{file_id}" if file_id else None

    return json.dumps({
        "model_type": "equity_valuation",
        "company_name": company_name,
        "ticker": ticker,
        "implied_price_dcf_gordon": price_gordon,
        "implied_price_dcf_exit": price_exit,
        "implied_price_blended": blended_price,
        "wacc_pct": round(wacc * 100, 2),
        "upside_pct": upside_pct,
        "recommendation": recommendation,
        "model_checks_passed": checks_passed,
        "model_checks_total": checks_total,
        "download_url": download_url,
    })


# ── Tool registration ───────────────────────────────────────────────────────

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "build_equity_valuation_model",
            "description": (
                "Build an IB-grade equity valuation workbook with 10 interconnected sheets: "
                "Assumptions, Income Statement, Balance Sheet, Cash Flow, Beta & WACC, DCF, "
                "Comps, Valuation Summary, Sensitivity, Model Checks. Generates a single "
                "downloadable Excel file with cross-sheet formulas, dual terminal value "
                "(Gordon Growth + Exit Multiple), 5 beta methods, peer comps, football field, "
                "and integrity checks."
            ),
            "parameters": {
                "type": "object",
                "required": [
                    "company_name",
                    "ticker",
                    "current_price",
                    "shares_outstanding",
                    "beta_raw",
                    "risk_free_rate",
                    "equity_risk_premium",
                    "cost_of_debt",
                    "tax_rate",
                    "debt_weight",
                    "historical_income",
                    "projection_years",
                    "terminal_growth",
                ],
                "properties": {
                    "company_name": {
                        "type": "string",
                        "description": "Company name (e.g. 'Apple Inc.')",
                    },
                    "ticker": {
                        "type": "string",
                        "description": "Stock ticker symbol (e.g. 'AAPL')",
                    },
                    "currency": {
                        "type": "string",
                        "description": "Currency code (default: USD)",
                    },
                    "current_price": {
                        "type": "number",
                        "description": "Current share price",
                    },
                    "shares_outstanding": {
                        "type": "number",
                        "description": "Basic shares outstanding in millions",
                    },
                    "diluted_shares": {
                        "type": "number",
                        "description": "Diluted shares outstanding in millions (defaults to basic)",
                    },
                    "high_52w": {
                        "type": "number",
                        "description": "52-week high price",
                    },
                    "low_52w": {
                        "type": "number",
                        "description": "52-week low price",
                    },
                    "annual_dividend": {
                        "type": "number",
                        "description": "Annual dividend per share (default: 0)",
                    },
                    "beta_raw": {
                        "type": "number",
                        "description": "Raw regression beta",
                    },
                    "risk_free_rate": {
                        "type": "number",
                        "description": "Risk-free rate as decimal (e.g. 0.04 for 4%)",
                    },
                    "equity_risk_premium": {
                        "type": "number",
                        "description": "Equity risk premium as decimal (e.g. 0.06 for 6%)",
                    },
                    "cost_of_debt": {
                        "type": "number",
                        "description": "Pre-tax cost of debt as decimal",
                    },
                    "tax_rate": {
                        "type": "number",
                        "description": "Corporate tax rate as decimal (e.g. 0.25)",
                    },
                    "debt_weight": {
                        "type": "number",
                        "description": "Debt weight in capital structure as decimal (e.g. 0.02)",
                    },
                    "blume_beta": {
                        "type": "number",
                        "description": "Blume-adjusted beta (auto-calculated if omitted: 2/3×raw + 1/3)",
                    },
                    "bottom_up_beta": {
                        "type": "number",
                        "description": "Bottom-up (industry) beta",
                    },
                    "fama_french_beta": {
                        "type": "number",
                        "description": "Fama-French 3-factor beta",
                    },
                    "market_implied_beta": {
                        "type": "number",
                        "description": "Market-implied (reverse-engineered) beta",
                    },
                    "historical_income": {
                        "type": "array",
                        "description": "3-5 years of historical income statement data",
                        "items": {
                            "type": "object",
                            "required": ["year", "revenue"],
                            "properties": {
                                "year": {"type": ["string", "number"], "description": "Fiscal year label"},
                                "revenue": {"type": "number"},
                                "cogs": {"type": "number"},
                                "gross_profit": {"type": "number"},
                                "rd": {"type": "number", "description": "R&D expense"},
                                "sga": {"type": "number", "description": "SG&A expense"},
                                "total_opex": {"type": "number"},
                                "ebit": {"type": "number"},
                                "other_income": {"type": "number"},
                                "pbt": {"type": "number", "description": "Pre-tax income"},
                                "tax": {"type": "number"},
                                "net_income": {"type": "number"},
                                "da": {"type": "number", "description": "Depreciation & Amortization"},
                                "ebitda": {"type": "number"},
                            },
                        },
                    },
                    "historical_balance_sheet": {
                        "type": "array",
                        "description": "3-5 years of balance sheet data",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": ["string", "number"]},
                                "cash": {"type": "number"},
                                "st_investments": {"type": "number"},
                                "accounts_receivable": {"type": "number"},
                                "inventory": {"type": "number"},
                                "other_current_assets": {"type": "number"},
                                "lt_investments": {"type": "number"},
                                "ppe_net": {"type": "number"},
                                "other_noncurrent_assets": {"type": "number"},
                                "accounts_payable": {"type": "number"},
                                "st_debt": {"type": "number"},
                                "other_current_liabilities": {"type": "number"},
                                "lt_debt": {"type": "number"},
                                "other_noncurrent_liabilities": {"type": "number"},
                                "total_equity": {"type": "number"},
                            },
                        },
                    },
                    "historical_cash_flow": {
                        "type": "array",
                        "description": "3-5 years of cash flow data",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": ["string", "number"]},
                                "net_income": {"type": "number"},
                                "da": {"type": "number"},
                                "sbc": {"type": "number", "description": "Stock-based compensation"},
                                "wc_changes": {"type": "number"},
                                "other_operating": {"type": "number"},
                                "capex": {"type": "number", "description": "Capital expenditures (negative)"},
                                "other_investing": {"type": "number"},
                                "dividends": {"type": "number", "description": "Dividends paid (negative)"},
                                "buybacks": {"type": "number", "description": "Share buybacks (negative)"},
                                "net_debt_issuance": {"type": "number"},
                                "other_financing": {"type": "number"},
                            },
                        },
                    },
                    "projection_years": {
                        "type": "array",
                        "description": "5 years of projection assumptions",
                        "items": {
                            "type": "object",
                            "required": ["year", "rev_growth", "gross_margin"],
                            "properties": {
                                "year": {"type": ["string", "number"], "description": "Projection year label"},
                                "rev_growth": {"type": "number", "description": "Revenue growth rate (decimal)"},
                                "gross_margin": {"type": "number", "description": "Gross margin (decimal)"},
                                "rd_pct": {"type": "number", "description": "R&D as % of revenue"},
                                "sga_pct": {"type": "number", "description": "SG&A as % of revenue"},
                                "tax_rate": {"type": "number", "description": "Effective tax rate"},
                                "da_pct": {"type": "number", "description": "D&A as % of revenue"},
                                "capex_pct": {"type": "number", "description": "CapEx as % of revenue"},
                                "nwc_pct": {"type": "number", "description": "ΔNWC as % of revenue"},
                            },
                        },
                    },
                    "terminal_growth": {
                        "type": "number",
                        "description": "Terminal growth rate as decimal (e.g. 0.03 for 3%)",
                    },
                    "exit_ev_ebitda": {
                        "type": "number",
                        "description": "Exit EV/EBITDA multiple for terminal value (default: 12x)",
                    },
                    "net_debt": {
                        "type": "number",
                        "description": "Net debt in millions (auto-derived from BS if omitted)",
                    },
                    "peers": {
                        "type": "array",
                        "description": "Comparable companies for Comps analysis",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "ticker": {"type": "string"},
                                "ev": {"type": "number", "description": "Enterprise value (M)"},
                                "ev_revenue": {"type": "number"},
                                "ev_ebitda": {"type": "number"},
                                "pe": {"type": "number"},
                                "rev_growth": {"type": "number", "description": "Revenue growth (decimal)"},
                                "ebitda_margin": {"type": "number", "description": "EBITDA margin (decimal)"},
                                "ni_margin": {"type": "number", "description": "Net income margin (decimal)"},
                            },
                        },
                    },
                    "bull_price": {"type": "number", "description": "Bull-case target price"},
                    "bull_description": {"type": "string", "description": "Bull-case narrative"},
                    "bear_price": {"type": "number", "description": "Bear-case target price"},
                    "bear_description": {"type": "string", "description": "Bear-case narrative"},
                    "risks": {
                        "type": "string",
                        "description": "Key risks (newline-separated)",
                    },
                    "catalysts": {
                        "type": "string",
                        "description": "Key catalysts (newline-separated)",
                    },
                },
            },
        },
    },
]

TOOL_DISPATCH = {
    "build_equity_valuation_model": build_equity_valuation_model,
}
