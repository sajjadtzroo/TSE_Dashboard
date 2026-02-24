# rag/tools/financial_modeling.py
"""Financial modeling tools: DCF, P&L, Loan Amortization, Bond Pricing."""
from __future__ import annotations

import json
import logging
import math
import random
import statistics
import uuid
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)

# Lazy import to avoid circular imports at module load time
def _get_excel_models_dir() -> Path:
    from config.settings import EXCEL_MODELS_DIR
    return EXCEL_MODELS_DIR


# ── Excel styling helpers ─────────────────────────────────────────────────────

_FILL_HEADER  = PatternFill("solid", fgColor="1C2030") if EXCEL_AVAILABLE else None
_FILL_INPUT   = PatternFill("solid", fgColor="FFFDE7") if EXCEL_AVAILABLE else None
_FILL_FORMULA = PatternFill("solid", fgColor="E3F2FD") if EXCEL_AVAILABLE else None
_FILL_RESULT  = PatternFill("solid", fgColor="E8F5E9") if EXCEL_AVAILABLE else None

_FONT_HEADER  = Font(bold=True, color="FFFFFF") if EXCEL_AVAILABLE else None
_FONT_RESULT  = Font(bold=True) if EXCEL_AVAILABLE else None
_ALIGN_CENTER = Alignment(horizontal="center") if EXCEL_AVAILABLE else None


def _style_header(cell):
    cell.fill = _FILL_HEADER
    cell.font = _FONT_HEADER
    cell.alignment = _ALIGN_CENTER


def _style_input(cell):
    cell.fill = _FILL_INPUT


def _style_formula(cell):
    cell.fill = _FILL_FORMULA


def _style_result(cell):
    cell.fill = _FILL_RESULT
    cell.font = _FONT_RESULT


def _auto_width(ws, min_width: int = 12):
    """Set column widths to max content length."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _save_excel(wb, title: str) -> str | None:
    """Save workbook to data/models/{uuid}.xlsx, return file_id or None."""
    if not EXCEL_AVAILABLE:
        return None
    try:
        models_dir = _get_excel_models_dir()
        models_dir.mkdir(parents=True, exist_ok=True)
        file_id = str(uuid.uuid4())
        path = models_dir / f"{file_id}.xlsx"
        wb.save(str(path))
        return file_id
    except Exception as exc:
        logger.warning("Excel save failed: %s", exc)
        return None


# ── DCF Tool ──────────────────────────────────────────────────────────────────

def _compute_fcff(ebit: float, tax_rate: float, da: float, capex: float, delta_wc: float) -> float:
    """FCFF = EBIT(1-T) + D&A - CapEx - ΔWC"""
    return ebit * (1 - tax_rate) + da - capex - delta_wc


def _build_dcf_workbook(
    company_name: str,
    projections: list[dict],
    wacc: float,
    terminal_growth: float,
    net_debt: float,
    shares_outstanding: float,
    pv_sum: float,
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Assumptions (user-editable inputs) ───────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.sheet_view.rightToLeft = True

    ws_a["A1"] = f"DCF Model — {company_name}"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a.merge_cells("A1:B1")

    for col, h in enumerate(["Parameter", "Value"], 1):
        _style_header(ws_a.cell(row=3, column=col, value=h))

    inputs = [
        ("WACC", wacc),
        ("Terminal Growth Rate", terminal_growth),
        ("Net Debt (B IRR)", net_debt),
        ("Shares Outstanding (M)", shares_outstanding),
    ]
    for r, (label, val) in enumerate(inputs, 4):
        ws_a.cell(row=r, column=1, value=label)
        _style_input(ws_a.cell(row=r, column=2, value=val))

    _auto_width(ws_a)

    # ── Sheet 2: DCF Valuation ────────────────────────────────────────────────
    ws_d = wb.create_sheet("DCF")
    ws_d.sheet_view.rightToLeft = True

    n = len(projections)
    ws_d.cell(row=1, column=1, value="Item")
    _style_header(ws_d.cell(row=1, column=1))
    for i in range(n):
        c = ws_d.cell(row=1, column=i + 2, value=f"Year {i+1}")
        _style_header(c)

    row_labels = {
        2: "Year #",
        3: "EBIT (B IRR)",
        4: "Tax Rate",
        5: "D&A (B IRR)",
        6: "CapEx (B IRR)",
        7: "Delta WC (B IRR)",
    }
    for row_num, label in row_labels.items():
        ws_d.cell(row=row_num, column=1, value=label)

    for i, proj in enumerate(projections):
        col = i + 2
        ws_d.cell(row=2, column=col, value=i + 1)
        for row_num, key in [(3, "ebit"), (4, "tax_rate"), (5, "da"), (6, "capex"), (7, "delta_wc")]:
            _style_input(ws_d.cell(row=row_num, column=col, value=proj[key]))

    # Row 8: FCFF formula =B3*(1-B4)+B5-B6-B7
    ws_d.cell(row=8, column=1, value="FCFF (B IRR)")
    for i in range(n):
        cl = get_column_letter(i + 2)
        c = ws_d.cell(row=8, column=i + 2)
        c.value = f"={cl}3*(1-{cl}4)+{cl}5-{cl}6-{cl}7"
        _style_formula(c)

    # Row 9: Discount factor =(1+Assumptions!$B$4)^B2
    ws_d.cell(row=9, column=1, value="Discount Factor")
    for i in range(n):
        cl = get_column_letter(i + 2)
        c = ws_d.cell(row=9, column=i + 2)
        c.value = f"=(1+Assumptions!$B$4)^{cl}2"
        _style_formula(c)

    # Row 10: PV of FCFF =B8/B9
    ws_d.cell(row=10, column=1, value="PV of FCFF")
    for i in range(n):
        cl = get_column_letter(i + 2)
        c = ws_d.cell(row=10, column=i + 2)
        c.value = f"={cl}8/{cl}9"
        _style_formula(c)

    # Terminal value block (rows 13–21)
    last_col = get_column_letter(n + 1)
    first_pv_col = get_column_letter(2)

    tv_rows = [
        ("Sum PV FCFFs",        f"=SUM({first_pv_col}10:{last_col}10)"),
        ("Terminal FCFF",        f"={last_col}8*(1+Assumptions!$B$5)"),
        ("Terminal Value",       f"=B14/(Assumptions!$B$4-Assumptions!$B$5)"),
        ("PV Terminal Value",    f"=B15/(1+Assumptions!$B$4)^{n}"),
        ("Enterprise Value",     "=B13+B16"),
        ("(-) Net Debt",         "=Assumptions!$B$6"),
        ("Equity Value",         "=B17-B18"),
        ("Shares (M)",           "=Assumptions!$B$7"),
        ("Price / Share (IRR)",  "=B19/B20"),
    ]
    for r_offset, (label, formula) in enumerate(tv_rows):
        row = 13 + r_offset
        ws_d.cell(row=row, column=1, value=label)
        _style_result(ws_d.cell(row=row, column=2, value=formula))

    _auto_width(ws_d)

    # ── Sheet 3: Sensitivity (Python-computed) ────────────────────────────────
    ws_s = wb.create_sheet("Sensitivity")
    ws_s.sheet_view.rightToLeft = True

    _style_header(ws_s.cell(row=1, column=1, value="WACC \\ Terminal Growth"))

    # Cache all FCFFs for reuse across sensitivity rows
    _fcff_list = [
        _compute_fcff(p["ebit"], p["tax_rate"], p["da"], p["capex"], p["delta_wc"])
        for p in projections
    ]
    last_fcff_n = _fcff_list[-1]

    wacc_range = [round(wacc - 0.04 + i * 0.01, 4) for i in range(9)]
    tg_range = [round(terminal_growth - 0.02 + i * 0.01, 4) for i in range(5)]

    for col_i, tg in enumerate(tg_range, 2):
        _style_header(ws_s.cell(row=1, column=col_i, value=f"{round(tg * 100, 1)}%"))

    for row_i, w in enumerate(wacc_range, 2):
        ws_s.cell(row=row_i, column=1, value=f"{round(w * 100, 1)}%")
        for col_i, tg in enumerate(tg_range, 2):
            if w <= tg:
                ws_s.cell(row=row_i, column=col_i, value="N/A")
            else:
                # Recompute PV of explicit FCFFs at this row's WACC
                pv_fcff_at_w = sum(
                    fcff / (1 + w) ** (t + 1)
                    for t, fcff in enumerate(_fcff_list)
                )
                # Use sensitivity TG for TV numerator
                tv = last_fcff_n * (1 + tg) / (w - tg)
                pv_tv = tv / (1 + w) ** n
                ev = pv_fcff_at_w + pv_tv
                eq = ev - net_debt
                ps = round(eq / shares_outstanding, 2) if shares_outstanding else 0
                ws_s.cell(row=row_i, column=col_i, value=ps)

    _auto_width(ws_s)
    return wb


def build_dcf_model(
    db: Session,
    company_name: str,
    projections: list[dict],
    wacc: float,
    terminal_growth: float,
    net_debt: float = 0.0,
    shares_outstanding: float = 1000.0,
) -> str:
    """
    Build a DCF valuation model.

    Args:
        company_name: Company name for the spreadsheet title.
        projections: List of annual projections, each with keys:
            ebit, tax_rate, da, capex, delta_wc (all in billion IRR).
        wacc: Weighted average cost of capital (decimal, e.g. 0.22 for 22%).
        terminal_growth: Terminal growth rate (decimal, e.g. 0.03 for 3%).
        net_debt: Net debt = total debt - cash (billion IRR).
        shares_outstanding: Total shares in millions.

    Returns:
        JSON string with keys: enterprise_value, equity_value, price_per_share,
        pv_fcff, pv_terminal, download_url (null if Excel unavailable), model_type.
    """
    if wacc <= terminal_growth:
        return json.dumps({"error": "WACC must be greater than terminal growth rate (WACC > TG)"})

    if not projections:
        return json.dumps({"error": "At least one projection year is required"})

    # Compute values
    pv_sum = 0.0
    for i, proj in enumerate(projections):
        fcff = _compute_fcff(
            proj["ebit"], proj["tax_rate"], proj["da"], proj["capex"], proj["delta_wc"]
        )
        pv_sum += fcff / (1 + wacc) ** (i + 1)

    last_fcff = _compute_fcff(
        projections[-1]["ebit"],
        projections[-1]["tax_rate"],
        projections[-1]["da"],
        projections[-1]["capex"],
        projections[-1]["delta_wc"],
    )
    last_fcff_grown = last_fcff * (1 + terminal_growth)
    terminal_value = last_fcff_grown / (wacc - terminal_growth)
    pv_terminal = terminal_value / (1 + wacc) ** len(projections)

    enterprise_value = pv_sum + pv_terminal
    equity_value = enterprise_value - net_debt
    price_per_share = equity_value / shares_outstanding if shares_outstanding else 0

    download_url = None
    if EXCEL_AVAILABLE:
        try:
            wb = _build_dcf_workbook(
                company_name, projections, wacc, terminal_growth,
                net_debt, shares_outstanding, pv_sum,
            )
            file_id = _save_excel(wb, f"DCF-{company_name}")
            if file_id:
                download_url = f"/api/financial-modeling/download/{file_id}"
        except Exception as e:
            logger.warning("DCF Excel creation failed: %s", e)

    return json.dumps({
        "model_type": "dcf",
        "company_name": company_name,
        "enterprise_value": round(enterprise_value, 2),
        "equity_value": round(equity_value, 2),
        "price_per_share": round(price_per_share, 2),
        "pv_fcff": round(pv_sum, 2),
        "pv_terminal": round(pv_terminal, 2),
        "wacc_pct": round(wacc * 100, 2),
        "terminal_growth_pct": round(terminal_growth * 100, 2),
        "download_url": download_url,
    })


# ── P&L Tool ──────────────────────────────────────────────────────────────────

def _build_pl_workbook(
    company_name: str,
    years: list[dict],
    gross_margin: float,
    ebitda_margin: float,
    da_pct: float,
    interest_expense: float,
    tax_rate: float,
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Assumptions ─────────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "PL_Assumptions"
    ws_a.sheet_view.rightToLeft = True

    ws_a["A1"] = f"P&L Model — {company_name}"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a.merge_cells("A1:B1")

    for col, h in enumerate(["Parameter", "Value"], 1):
        _style_header(ws_a.cell(row=3, column=col, value=h))

    pl_inputs = [
        ("Gross Margin", gross_margin),
        ("EBITDA Margin", ebitda_margin),
        ("D&A % of Revenue", da_pct),
        ("Tax Rate", tax_rate),
        ("Interest Expense (B IRR)", interest_expense),
    ]
    for r, (label, val) in enumerate(pl_inputs, 4):
        ws_a.cell(row=r, column=1, value=label)
        _style_input(ws_a.cell(row=r, column=2, value=val))

    _auto_width(ws_a)

    # ── Sheet 2: P&L ─────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("P&L")
    ws_p.sheet_view.rightToLeft = True

    n = len(years)
    _style_header(ws_p.cell(row=1, column=1, value="Item"))
    for i, y in enumerate(years):
        _style_header(ws_p.cell(row=1, column=i + 2, value=f"Year {y['year']}"))

    # Row 2: Revenue (static computed values)
    ws_p.cell(row=2, column=1, value="Revenue (B IRR)")
    for i, y in enumerate(years):
        _style_input(ws_p.cell(row=2, column=i + 2, value=y["revenue"]))

    # Rows 3-5: margin formulas referencing PL_Assumptions
    # PL_Assumptions: B4=gross_margin, B5=ebitda_margin, B6=da_pct, B7=tax_rate, B8=interest_expense
    margin_rows = [
        (3, "Gross Profit",  "PL_Assumptions!$B$4"),
        (4, "EBITDA",        "PL_Assumptions!$B$5"),
        (5, "D&A",           "PL_Assumptions!$B$6"),
    ]
    for row_num, label, margin_ref in margin_rows:
        ws_p.cell(row=row_num, column=1, value=label)
        for i in range(n):
            cl = get_column_letter(i + 2)
            _style_formula(ws_p.cell(row=row_num, column=i + 2, value=f"={cl}2*{margin_ref}"))

    # Row 6: EBIT = EBITDA - D&A
    ws_p.cell(row=6, column=1, value="EBIT")
    for i in range(n):
        cl = get_column_letter(i + 2)
        _style_formula(ws_p.cell(row=6, column=i + 2, value=f"={cl}4-{cl}5"))

    # Row 7: Interest expense (constant from assumptions B8)
    ws_p.cell(row=7, column=1, value="Interest Expense (B IRR)")
    for i in range(n):
        _style_input(ws_p.cell(row=7, column=i + 2, value="=PL_Assumptions!$B$8"))

    # Row 8: EBT
    ws_p.cell(row=8, column=1, value="EBT")
    for i in range(n):
        cl = get_column_letter(i + 2)
        _style_formula(ws_p.cell(row=8, column=i + 2, value=f"={cl}6-{cl}7"))

    # Row 9: Tax =IF(B8>0,B8*PL_Assumptions!$B$7,0)
    ws_p.cell(row=9, column=1, value="Tax")
    for i in range(n):
        cl = get_column_letter(i + 2)
        _style_formula(ws_p.cell(row=9, column=i + 2, value=f"=IF({cl}8>0,{cl}8*PL_Assumptions!$B$7,0)"))

    # Row 10: Net Income =EBT - Tax
    ws_p.cell(row=10, column=1, value="Net Income")
    for i in range(n):
        cl = get_column_letter(i + 2)
        _style_result(ws_p.cell(row=10, column=i + 2, value=f"={cl}8-{cl}9"))

    _auto_width(ws_p)
    return wb


def build_pl_model(
    db: Session,
    company_name: str,
    base_revenue: float,
    revenue_growth_rates: list[float],
    gross_margin: float,
    ebitda_margin: float,
    da_pct: float,
    interest_expense: float,
    tax_rate: float,
) -> str:
    """
    Build a multi-year P&L projection model.

    Args:
        company_name: Company name.
        base_revenue: Base year revenue (billion IRR).
        revenue_growth_rates: List of annual growth rates (decimals), one per year.
        gross_margin: Gross profit / Revenue (decimal).
        ebitda_margin: EBITDA / Revenue (decimal).
        da_pct: D&A as % of Revenue (decimal).
        interest_expense: Annual interest expense (billion IRR, constant).
        tax_rate: Corporate tax rate (decimal).

    Returns:
        JSON string with projected P&L for each year + download_url.
    """
    years = []
    revenue = base_revenue

    for i, growth in enumerate(revenue_growth_rates):
        revenue = revenue * (1 + growth)
        gross_profit = revenue * gross_margin
        ebitda = revenue * ebitda_margin
        da = revenue * da_pct
        ebit = ebitda - da
        ebt = ebit - interest_expense
        net_income = ebt * (1 - tax_rate) if ebt > 0 else ebt
        net_margin = net_income / revenue if revenue else 0

        years.append({
            "year": i + 1,
            "revenue": round(revenue, 2),
            "gross_profit": round(gross_profit, 2),
            "gross_margin_pct": round(gross_margin * 100, 2),
            "ebitda": round(ebitda, 2),
            "ebitda_margin_pct": round(ebitda_margin * 100, 2),
            "da": round(da, 2),
            "ebit": round(ebit, 2),
            "ebit_margin_pct": round(ebit / revenue * 100, 2) if revenue else 0,
            "interest_expense": round(interest_expense, 2),
            "ebt": round(ebt, 2),
            "tax": round(ebt * tax_rate, 2) if ebt > 0 else 0,
            "net_income": round(net_income, 2),
            "net_margin_pct": round(net_margin * 100, 2),
        })

    download_url = None
    if EXCEL_AVAILABLE:
        try:
            wb = _build_pl_workbook(
                company_name, years,
                gross_margin, ebitda_margin, da_pct, interest_expense, tax_rate,
            )
            file_id = _save_excel(wb, f"PL-{company_name}")
            if file_id:
                download_url = f"/api/financial-modeling/download/{file_id}"
        except Exception as e:
            logger.warning("P&L Excel creation failed: %s", e)

    return json.dumps({
        "model_type": "pl",
        "company_name": company_name,
        "projections": years,
        "download_url": download_url,
    })


# ── Math helpers (numpy_financial fallbacks) ─────────────────────────────────

def _pmt(rate: float, nper: int, pv: float) -> float:
    """PMT: periodic payment for a fully-amortizing loan.

    Formula: PMT = PV * r * (1+r)^n / ((1+r)^n - 1)
    """
    if rate == 0:
        return pv / nper
    factor = (1 + rate) ** nper
    return pv * rate * factor / (factor - 1)


def _irr(cash_flows: list[float], max_iter: int = 100, tol: float = 1e-6) -> float | None:
    """Compute IRR via Newton-Raphson. Returns per-period rate or None on failure."""
    rate = 0.1
    for _ in range(max_iter):
        npv = sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))
        dnpv = sum(-t * cf / (1 + rate) ** (t + 1) for t, cf in enumerate(cash_flows))
        if abs(dnpv) < 1e-12:
            break
        new_rate = rate - npv / dnpv
        if abs(new_rate - rate) < tol:
            return new_rate
        rate = new_rate
    return rate if abs(sum(cf / (1 + rate) ** t for t, cf in enumerate(cash_flows))) < 0.01 else None


# ── Loan Amortization Tool ────────────────────────────────────────────────────

def _build_loan_workbook(
    principal: float,
    annual_rate: float,
    term_months: int,
    loan_type: str,
    schedule: list[dict],
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Inputs ───────────────────────────────────────────────────────
    ws_i = wb.active
    ws_i.title = "Inputs"
    ws_i.sheet_view.rightToLeft = True

    ws_i["A1"] = "Loan Amortization Schedule"
    ws_i["A1"].font = Font(bold=True, size=14)
    ws_i.merge_cells("A1:B1")

    for col, h in enumerate(["Parameter", "Value"], 1):
        _style_header(ws_i.cell(row=3, column=col, value=h))

    loan_inputs = [
        ("Principal (M IRR)", principal),
        ("Annual Rate",       annual_rate),
        ("Term (months)",     term_months),
        ("Loan Type",         loan_type),
    ]
    for r, (label, val) in enumerate(loan_inputs, 4):
        ws_i.cell(row=r, column=1, value=label)
        _style_input(ws_i.cell(row=r, column=2, value=val))

    # PMT formula for fully_amortizing (informational)
    if loan_type == "fully_amortizing":
        ws_i.cell(row=9, column=1, value="Monthly Payment (PMT)")
        _style_result(ws_i.cell(row=9, column=2, value="=PMT(B5/12,B6,-B4)"))

    _auto_width(ws_i)

    # ── Sheet 2: Schedule ─────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Schedule")
    ws_s.sheet_view.rightToLeft = True

    for col, h in enumerate(["Month", "Payment", "Principal", "Interest", "Balance"], 1):
        _style_header(ws_s.cell(row=1, column=col, value=h))

    if loan_type == "fully_amortizing":
        # Month 1 — reference Inputs!B9 for payment
        ws_s.cell(row=2, column=1, value=1)
        _style_formula(ws_s.cell(row=2, column=2, value="=Inputs!$B$9"))
        # Interest month 1: principal * monthly_rate
        _style_formula(ws_s.cell(row=2, column=4, value="=Inputs!$B$4*Inputs!$B$5/12"))
        # Principal = payment - interest
        _style_formula(ws_s.cell(row=2, column=3, value="=B2-D2"))
        # Balance = principal - principal_payment
        _style_formula(ws_s.cell(row=2, column=5, value="=Inputs!$B$4-C2"))

        for row in range(3, term_months + 2):
            prev = row - 1
            ws_s.cell(row=row, column=1, value=row - 1)
            _style_formula(ws_s.cell(row=row, column=2, value="=Inputs!$B$9"))
            _style_formula(ws_s.cell(row=row, column=4, value=f"=E{prev}*Inputs!$B$5/12"))
            _style_formula(ws_s.cell(row=row, column=3, value=f"=B{row}-D{row}"))
            _style_formula(ws_s.cell(row=row, column=5, value=f"=E{prev}-C{row}"))
    elif loan_type == "bullet":
        # Bullet: interest-only payments, principal returned at maturity
        # Inputs: B4=principal, B5=annual_rate, B6=term_months
        # Month 1
        ws_s.cell(row=2, column=1, value=1)
        # Interest = principal × monthly rate
        _style_formula(ws_s.cell(row=2, column=4, value="=Inputs!$B$4*Inputs!$B$5/12"))
        # Principal = 0 for all except last month
        _style_formula(ws_s.cell(row=2, column=3,
                                  value=f"=IF(A2=Inputs!$B$6,Inputs!$B$4,0)"))
        # Payment = Interest + Principal
        _style_formula(ws_s.cell(row=2, column=2, value="=C2+D2"))
        # Balance = principal - cumulative principal paid
        _style_formula(ws_s.cell(row=2, column=5, value="=Inputs!$B$4-C2"))

        for row in range(3, term_months + 2):
            prev = row - 1
            ws_s.cell(row=row, column=1, value=row - 1)
            _style_formula(ws_s.cell(row=row, column=4, value="=Inputs!$B$4*Inputs!$B$5/12"))
            _style_formula(ws_s.cell(row=row, column=3,
                                      value=f"=IF(A{row}=Inputs!$B$6,Inputs!$B$4,0)"))
            _style_formula(ws_s.cell(row=row, column=2, value=f"=C{row}+D{row}"))
            _style_formula(ws_s.cell(row=row, column=5, value=f"=E{prev}-C{row}"))

    else:
        # Balloon: amortizing payments over extended term, balloon at balloon_month
        # Write computed values (balloon logic is complex with variable cutoff)
        for row_i, s in enumerate(schedule, 2):
            ws_s.cell(row=row_i, column=1, value=s["month"])
            ws_s.cell(row=row_i, column=2, value=s["payment"])
            ws_s.cell(row=row_i, column=3, value=s["principal"])
            ws_s.cell(row=row_i, column=4, value=s["interest"])
            ws_s.cell(row=row_i, column=5, value=s["balance"])

    _auto_width(ws_s)
    return wb


def build_loan_amortization(
    db: Session,
    principal: float,
    annual_rate: float,
    term_months: int,
    loan_type: str = "fully_amortizing",
    balloon_month: Optional[int] = None,
) -> str:
    """
    Build a loan amortization schedule.

    Args:
        principal: Loan principal amount (million IRR).
        annual_rate: Annual interest rate (decimal, e.g. 0.18 for 18%).
        term_months: Total loan term in months.
        loan_type: One of 'fully_amortizing', 'bullet', 'balloon'.
        balloon_month: Month at which balloon payment occurs (required for 'balloon' type).

    Returns:
        JSON string with amortization schedule + summary + download_url.
    """
    try:
        import numpy_financial as npf
        _pmt_fn = lambda r, n, pv: -npf.pmt(r, n, pv)
        _irr_fn = npf.irr
    except ImportError:
        _pmt_fn = _pmt
        _irr_fn = _irr

    monthly_rate = annual_rate / 12

    if loan_type == "fully_amortizing":
        monthly_payment = _pmt_fn(monthly_rate, term_months, principal)

        schedule = []
        balance = principal
        total_interest = 0.0
        for month in range(1, term_months + 1):
            interest = balance * monthly_rate
            principal_payment = monthly_payment - interest
            balance = max(0.0, balance - principal_payment)
            total_interest += interest
            schedule.append({
                "month": month,
                "payment": round(monthly_payment, 2),
                "principal": round(principal_payment, 2),
                "interest": round(interest, 2),
                "balance": round(balance, 2),
            })

        summary = {
            "loan_type": "fully_amortizing",
            "principal": principal,
            "monthly_payment": round(monthly_payment, 2),
            "total_paid": round(monthly_payment * term_months, 2),
            "total_interest": round(total_interest, 2),
            "final_balance": schedule[-1]["balance"],
        }

    elif loan_type == "bullet":
        monthly_payment = principal * monthly_rate
        schedule = []
        total_interest = 0.0
        for month in range(1, term_months + 1):
            interest = principal * monthly_rate
            balloon = principal if month == term_months else 0.0
            payment = interest + balloon
            total_interest += interest
            schedule.append({
                "month": month,
                "payment": round(payment, 2),
                "principal": round(balloon, 2),
                "interest": round(interest, 2),
                "balance": round(principal if month < term_months else 0.0, 2),
            })

        summary = {
            "loan_type": "bullet",
            "principal": principal,
            "monthly_interest_payment": round(monthly_payment, 2),
            "balloon_payment_month": term_months,
            "balloon_amount": round(principal + monthly_payment, 2),
            "total_paid": round(monthly_payment * term_months + principal, 2),
            "total_interest": round(total_interest, 2),
            "final_balance": 0.0,
        }

    elif loan_type == "balloon":
        # balloon_month: the month at which the remaining balance is paid off.
        # Payments sized as if fully amortizing over extended_term (double the loan term).
        if balloon_month is None or balloon_month >= term_months:
            balloon_month = term_months  # default: balloon at loan end

        extended_term = term_months * 2
        monthly_payment = _pmt_fn(monthly_rate, extended_term, principal)

        schedule = []
        balance = principal
        total_interest = 0.0
        for month in range(1, balloon_month + 1):
            interest = balance * monthly_rate
            if month == balloon_month:
                # Balloon: pay entire remaining balance
                payment = balance + interest
                principal_payment = balance
                balance = 0.0
            else:
                principal_payment = monthly_payment - interest
                balance = max(0.0, balance - principal_payment)
                payment = monthly_payment
            total_interest += interest
            schedule.append({
                "month": month,
                "payment": round(payment, 2),
                "principal": round(principal_payment, 2),
                "interest": round(interest, 2),
                "balance": round(balance, 2),
            })

        summary = {
            "loan_type": "balloon",
            "principal": principal,
            "monthly_payment": round(monthly_payment, 2),
            "balloon_month": balloon_month,
            "balloon_amount": round(schedule[-1]["payment"], 2),
            "total_paid": round(sum(s["payment"] for s in schedule), 2),
            "total_interest": round(total_interest, 2),
            "final_balance": 0.0,
        }

    else:
        return json.dumps({"error": f"Unknown loan_type: {loan_type}. Use 'fully_amortizing', 'bullet', or 'balloon'."})

    download_url = None
    if EXCEL_AVAILABLE:
        try:
            wb = _build_loan_workbook(principal, annual_rate, term_months, loan_type, schedule)
            file_id = _save_excel(wb, f"Loan-{loan_type}")
            if file_id:
                download_url = f"/api/financial-modeling/download/{file_id}"
        except Exception as e:
            logger.warning("Loan Excel creation failed: %s", e)

    return json.dumps({
        "model_type": "loan_amortization",
        "summary": summary,
        "schedule": schedule,
        "download_url": download_url,
    })


# ── Bond Pricing Tool ─────────────────────────────────────────────────────────

def _build_bond_workbook(
    face_value: float,
    coupon_rate: float,
    ytm: float,
    periods: int,
    frequency: int,
    price: float,
    ytm_from_irr: float,
    macaulay_duration_years: float,
    modified_duration: float,
    schedule: list[dict],
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Inputs ───────────────────────────────────────────────────────
    ws_i = wb.active
    ws_i.title = "Inputs"
    ws_i.sheet_view.rightToLeft = True

    ws_i["A1"] = "Bond Pricing Model"
    ws_i["A1"].font = Font(bold=True, size=14)
    ws_i.merge_cells("A1:B1")

    for col, h in enumerate(["Parameter", "Value"], 1):
        _style_header(ws_i.cell(row=3, column=col, value=h))

    bond_inputs = [
        ("Face Value (IRR)", face_value),
        ("Coupon Rate",      coupon_rate),
        ("YTM",              ytm),
        ("Periods",          periods),
        ("Frequency",        frequency),
    ]
    for r, (label, val) in enumerate(bond_inputs, 4):
        ws_i.cell(row=r, column=1, value=label)
        _style_input(ws_i.cell(row=r, column=2, value=val))

    # Excel PV formula
    ws_i.cell(row=10, column=1, value="Bond Price (PV formula)")
    _style_result(ws_i.cell(row=10, column=2, value="=-PV(B6/B8,B7,-B4*B5/B8,-B4)"))

    # Computed metrics — formulas referencing Cash Flows sheet
    irr_row = periods + 4
    formula_metrics = [
        ("Computed Price",           "=-'Cash Flows'!D2"),
        ("Periodic YTM (IRR)",       f"='Cash Flows'!B{irr_row}"),
        ("Annual YTM",               f"='Cash Flows'!B{irr_row + 1}"),
        ("Macaulay Duration (yrs)",  f"='Cash Flows'!B{irr_row + 2}"),
        ("Modified Duration",        f"='Cash Flows'!B{irr_row + 3}"),
    ]
    for r_offset, (label, formula) in enumerate(formula_metrics):
        row = 12 + r_offset
        ws_i.cell(row=row, column=1, value=label)
        _style_formula(ws_i.cell(row=row, column=2, value=formula))

    _auto_width(ws_i)

    # ── Sheet 2: Cash Flows ───────────────────────────────────────────────────
    ws_c = wb.create_sheet("Cash Flows")
    ws_c.sheet_view.rightToLeft = True

    for col, h in enumerate(["Period", "Coupon", "Principal", "Total CF", "PV"], 1):
        _style_header(ws_c.cell(row=1, column=col, value=h))

    # Inputs refs: B4=face, B5=coupon_rate, B6=YTM, B7=periods, B8=frequency
    # Row 2: purchase price (negative) for IRR — formula = -PV(...)
    ws_c.cell(row=2, column=1, value=0)
    _style_formula(ws_c.cell(row=2, column=4, value="=PV(Inputs!$B$6/Inputs!$B$8,Inputs!$B$7,-Inputs!$B$4*Inputs!$B$5/Inputs!$B$8,-Inputs!$B$4)"))

    for row_i in range(3, periods + 3):
        period = row_i - 2
        ws_c.cell(row=row_i, column=1, value=period)
        # Coupon = face × coupon_rate / frequency
        _style_formula(ws_c.cell(row=row_i, column=2,
                                  value="=Inputs!$B$4*Inputs!$B$5/Inputs!$B$8"))
        # Principal = face if last period, else 0
        _style_formula(ws_c.cell(row=row_i, column=3,
                                  value=f"=IF(A{row_i}=Inputs!$B$7,Inputs!$B$4,0)"))
        # Total CF = Coupon + Principal
        _style_formula(ws_c.cell(row=row_i, column=4,
                                  value=f"=B{row_i}+C{row_i}"))
        # PV = Total CF / (1 + periodic YTM)^period
        _style_formula(ws_c.cell(row=row_i, column=5,
                                  value=f"=D{row_i}/(1+Inputs!$B$6/Inputs!$B$8)^A{row_i}"))

    # IRR formula over D range (D2 = -price, D3..D(n+2) = cash flows)
    irr_row = periods + 4
    ws_c.cell(row=irr_row, column=1, value="Periodic YTM (IRR)")
    _style_result(ws_c.cell(row=irr_row, column=2, value=f"=IRR(D2:D{periods + 2})"))

    # Duration formulas
    ws_c.cell(row=irr_row + 1, column=1, value="Annual YTM")
    _style_result(ws_c.cell(row=irr_row + 1, column=2,
                             value=f"={get_column_letter(2)}{irr_row}*Inputs!$B$8"))

    # Macaulay Duration = SUMPRODUCT(period, PV) / SUM(PV) / frequency
    ws_c.cell(row=irr_row + 2, column=1, value="Macaulay Duration (years)")
    _style_result(ws_c.cell(row=irr_row + 2, column=2,
                             value=f"=SUMPRODUCT(A3:A{periods+2},E3:E{periods+2})/SUM(E3:E{periods+2})/Inputs!$B$8"))

    # Modified Duration = Macaulay / (1 + YTM/frequency)
    ws_c.cell(row=irr_row + 3, column=1, value="Modified Duration")
    _style_result(ws_c.cell(row=irr_row + 3, column=2,
                             value=f"=B{irr_row+2}/(1+Inputs!$B$6/Inputs!$B$8)"))

    _auto_width(ws_c)
    return wb


def build_bond_model(
    db: Session,
    face_value: float,
    coupon_rate: float,
    periods: int,
    ytm: float,
    frequency: int = 1,
) -> str:
    """
    Build a bond pricing and risk metrics model.

    Args:
        face_value: Face/par value of the bond (IRR).
        coupon_rate: Annual coupon rate (decimal, e.g. 0.18 for 18%).
        periods: Number of coupon periods until maturity.
        ytm: Yield to maturity used for pricing (decimal).
        frequency: Coupon payments per year (1=annual, 2=semi-annual).

    Returns:
        JSON string with price, ytm, Macaulay duration, Modified duration, download_url.
    """
    try:
        import numpy_financial as npf
        _irr_fn = npf.irr
    except ImportError:
        _irr_fn = _irr

    periodic_ytm = ytm / frequency
    periodic_coupon = face_value * coupon_rate / frequency

    # Bond price = PV of coupons + PV of face value
    price = 0.0
    for t in range(1, periods + 1):
        price += periodic_coupon / (1 + periodic_ytm) ** t
    price += face_value / (1 + periodic_ytm) ** periods

    # Macaulay Duration
    weighted_time = 0.0
    for t in range(1, periods + 1):
        pv_cf = periodic_coupon / (1 + periodic_ytm) ** t
        weighted_time += t * pv_cf
    weighted_time += periods * face_value / (1 + periodic_ytm) ** periods
    macaulay_duration = weighted_time / price
    macaulay_duration_years = macaulay_duration / frequency

    # Modified Duration
    modified_duration = macaulay_duration_years / (1 + ytm / frequency)

    # Convexity: (1/P) × Σ [CFt × t × (t+1) / (1+y)^(t+2)]
    convexity_sum = 0.0
    for t in range(1, periods + 1):
        cf = periodic_coupon if t < periods else periodic_coupon + face_value
        convexity_sum += cf * t * (t + 1) / (1 + periodic_ytm) ** (t + 2)
    convexity_periodic = convexity_sum / price
    # Convert from periods² to years²
    convexity = convexity_periodic / (frequency ** 2)

    # DV01: dollar value of 1 basis point change in yield
    dv01 = modified_duration * price * 0.0001

    # YTM verification via IRR
    cash_flows = [-price] + [periodic_coupon] * (periods - 1) + [periodic_coupon + face_value]
    try:
        irr_periodic = _irr_fn(cash_flows)
        ytm_from_irr = irr_periodic * frequency if irr_periodic is not None else ytm
    except Exception:
        ytm_from_irr = ytm

    # Build schedule
    schedule = []
    for t in range(1, periods + 1):
        coupon_payment = periodic_coupon
        principal = face_value if t == periods else 0.0
        schedule.append({
            "period": t,
            "coupon": round(coupon_payment, 2),
            "principal": round(principal, 2),
            "total_cash_flow": round(coupon_payment + principal, 2),
            "pv": round((coupon_payment + principal) / (1 + periodic_ytm) ** t, 2),
        })

    download_url = None
    if EXCEL_AVAILABLE:
        try:
            wb = _build_bond_workbook(
                face_value, coupon_rate, ytm, periods, frequency,
                price, ytm_from_irr, macaulay_duration_years, modified_duration,
                schedule,
            )
            file_id = _save_excel(wb, "Bond")
            if file_id:
                download_url = f"/api/financial-modeling/download/{file_id}"
        except Exception as e:
            logger.warning("Bond Excel creation failed: %s", e)

    return json.dumps({
        "model_type": "bond",
        "face_value": face_value,
        "coupon_rate_pct": round(coupon_rate * 100, 2),
        "ytm_pct": round(ytm * 100, 2),
        "periods": periods,
        "frequency": frequency,
        "price": round(price, 2),
        "ytm_from_irr_pct": round(ytm_from_irr * 100, 4),
        "macaulay_duration_years": round(macaulay_duration_years, 4),
        "modified_duration": round(modified_duration, 4),
        "convexity": round(convexity, 6),
        "dv01": round(dv01, 4),
        "schedule": schedule,
        "download_url": download_url,
    })


# ── WACC Tool ─────────────────────────────────────────────────────────────────

def compute_wacc(
    db: Session,
    equity_value: float,
    debt_value: float,
    cost_of_equity: float,
    cost_of_debt: float,
    tax_rate: float,
) -> str:
    """
    Compute Weighted Average Cost of Capital (WACC).

    Formula: WACC = (E/V) × Ke + (D/V) × Kd × (1 − T)
    """
    if equity_value < 0 or debt_value < 0:
        return json.dumps({"error": "equity_value and debt_value must be non-negative"})
    total = equity_value + debt_value
    if total == 0:
        return json.dumps({"error": "equity_value + debt_value must be positive"})
    if not (0 <= tax_rate <= 1):
        return json.dumps({"error": "tax_rate must be between 0 and 1"})

    equity_weight = equity_value / total
    debt_weight = debt_value / total
    after_tax_kd = cost_of_debt * (1 - tax_rate)
    wacc = equity_weight * cost_of_equity + debt_weight * after_tax_kd

    return json.dumps({
        "model_type": "wacc",
        "wacc": round(wacc, 6),
        "wacc_pct": round(wacc * 100, 4),
        "equity_weight": round(equity_weight, 4),
        "debt_weight": round(debt_weight, 4),
        "cost_of_equity_pct": round(cost_of_equity * 100, 4),
        "cost_of_debt_pct": round(cost_of_debt * 100, 4),
        "after_tax_cost_of_debt_pct": round(after_tax_kd * 100, 4),
        "tax_rate_pct": round(tax_rate * 100, 2),
        "formula": (
            f"WACC = {round(equity_weight*100,1)}% × {round(cost_of_equity*100,2)}% + "
            f"{round(debt_weight*100,1)}% × {round(cost_of_debt*100,2)}% × "
            f"(1 − {round(tax_rate*100,0):.0f}%)"
        ),
    })


# ── CAPM Tool ─────────────────────────────────────────────────────────────────

def compute_capm(
    db: Session,
    risk_free_rate: float,
    beta: float,
    equity_risk_premium: float,
    size_premium: float = 0.0,
    specific_premium: float = 0.0,
) -> str:
    """
    Compute cost of equity using CAPM.

    Formula: Ke = Rf + β × ERP + size_premium + specific_premium
    """
    cost_of_equity = risk_free_rate + beta * equity_risk_premium + size_premium + specific_premium

    return json.dumps({
        "model_type": "capm",
        "cost_of_equity": round(cost_of_equity, 6),
        "cost_of_equity_pct": round(cost_of_equity * 100, 4),
        "risk_free_rate_pct": round(risk_free_rate * 100, 4),
        "beta": round(beta, 4),
        "equity_risk_premium_pct": round(equity_risk_premium * 100, 4),
        "beta_contribution_pct": round(beta * equity_risk_premium * 100, 4),
        "size_premium_pct": round(size_premium * 100, 4),
        "specific_premium_pct": round(specific_premium * 100, 4),
        "formula": (
            f"Ke = {round(risk_free_rate*100,2)}% + "
            f"{round(beta,2)} × {round(equity_risk_premium*100,2)}% + "
            f"{round(size_premium*100,2)}% + {round(specific_premium*100,2)}%"
            f" = {round(cost_of_equity*100,2)}%"
        ),
    })


# ── DDM Tool ──────────────────────────────────────────────────────────────────

def build_ddm_model(
    db: Session,
    current_dividend: float,
    discount_rate: float,
    model_type: str = "gordon",
    growth_rate: Optional[float] = None,
    short_term_growth: Optional[float] = None,
    long_term_growth: Optional[float] = None,
    half_life: Optional[float] = None,
    stage_growth_rates: Optional[list] = None,
    terminal_growth: Optional[float] = None,
) -> str:
    """
    Dividend Discount Model (DDM) valuation.

    Supports: 'gordon' (single-stage), 'h_model' (linearly declining), 'multistage'.
    """
    if model_type == "gordon":
        if growth_rate is None:
            return json.dumps({"error": "growth_rate required for Gordon Growth model"})
        if discount_rate <= growth_rate:
            return json.dumps({"error": "discount_rate must be greater than growth_rate (ke > g)"})
        d1 = current_dividend * (1 + growth_rate)
        intrinsic_value = d1 / (discount_rate - growth_rate)
        return json.dumps({
            "model_type": "ddm_gordon",
            "intrinsic_value": round(intrinsic_value, 4),
            "d0": current_dividend,
            "d1": round(d1, 4),
            "growth_rate_pct": round(growth_rate * 100, 2),
            "discount_rate_pct": round(discount_rate * 100, 2),
            "formula": (
                f"P₀ = D₁/(ke−g) = {round(d1, 2)}/"
                f"({round(discount_rate*100,2)}%−{round(growth_rate*100,2)}%)"
            ),
        })

    elif model_type == "h_model":
        if short_term_growth is None or long_term_growth is None or half_life is None:
            return json.dumps({"error": "short_term_growth, long_term_growth, and half_life required for H-model"})
        if discount_rate <= long_term_growth:
            return json.dumps({"error": "discount_rate must be greater than long_term_growth (ke > gL)"})
        long_term_component = current_dividend * (1 + long_term_growth) / (discount_rate - long_term_growth)
        h_component = current_dividend * half_life * (short_term_growth - long_term_growth) / (discount_rate - long_term_growth)
        intrinsic_value = long_term_component + h_component
        return json.dumps({
            "model_type": "ddm_h_model",
            "intrinsic_value": round(intrinsic_value, 4),
            "long_term_component": round(long_term_component, 4),
            "h_component": round(h_component, 4),
            "d0": current_dividend,
            "short_term_growth_pct": round(short_term_growth * 100, 2),
            "long_term_growth_pct": round(long_term_growth * 100, 2),
            "half_life_years": half_life,
            "discount_rate_pct": round(discount_rate * 100, 2),
        })

    elif model_type == "multistage":
        if stage_growth_rates is None or terminal_growth is None:
            return json.dumps({"error": "stage_growth_rates and terminal_growth required for multistage DDM"})
        if discount_rate <= terminal_growth:
            return json.dumps({"error": "discount_rate must be greater than terminal_growth (ke > g_terminal)"})
        dividend = current_dividend
        schedule = []
        pv_dividends = 0.0
        for t, g in enumerate(stage_growth_rates, start=1):
            dividend = dividend * (1 + g)
            pv = dividend / (1 + discount_rate) ** t
            pv_dividends += pv
            schedule.append({
                "period": t,
                "growth_pct": round(g * 100, 2),
                "dividend": round(dividend, 4),
                "pv": round(pv, 4),
            })
        n = len(stage_growth_rates)
        terminal_dividend = dividend * (1 + terminal_growth)
        terminal_price = terminal_dividend / (discount_rate - terminal_growth)
        pv_terminal = terminal_price / (1 + discount_rate) ** n
        intrinsic_value = pv_dividends + pv_terminal
        return json.dumps({
            "model_type": "ddm_multistage",
            "intrinsic_value": round(intrinsic_value, 4),
            "pv_dividends": round(pv_dividends, 4),
            "pv_terminal": round(pv_terminal, 4),
            "terminal_price": round(terminal_price, 4),
            "terminal_growth_pct": round(terminal_growth * 100, 2),
            "discount_rate_pct": round(discount_rate * 100, 2),
            "schedule": schedule,
        })

    else:
        return json.dumps({"error": f"Unknown model_type '{model_type}'. Use 'gordon', 'h_model', or 'multistage'"})


# ── Residual Income Tool ───────────────────────────────────────────────────────

def build_residual_income_model(
    db: Session,
    book_value_per_share: float,
    earnings_per_share_list: list,
    cost_of_equity: float,
    persistence_factor: float = 1.0,
    continuing_ri_growth: float = 0.0,
) -> str:
    """
    Residual Income (RI) valuation model.

    Formula: V₀ = B₀ + Σ PV(RIₜ) + PV(continuing RI)
    where RIₜ = EPSₜ − ke × BVPSₜ₋₁
    """
    if not earnings_per_share_list:
        return json.dumps({"error": "earnings_per_share_list must not be empty"})
    if cost_of_equity <= 0:
        return json.dumps({"error": "cost_of_equity must be positive"})
    if not (0 <= persistence_factor <= 1):
        return json.dumps({"error": "persistence_factor must be between 0 and 1"})

    schedule = []
    pv_ri_sum = 0.0
    bvps = book_value_per_share

    for t, eps in enumerate(earnings_per_share_list, start=1):
        equity_charge = cost_of_equity * bvps
        ri = eps - equity_charge
        pv_ri = ri / (1 + cost_of_equity) ** t
        pv_ri_sum += pv_ri
        schedule.append({
            "period": t,
            "bvps_start": round(bvps, 4),
            "eps": round(eps, 4),
            "equity_charge": round(equity_charge, 4),
            "ri": round(ri, 4),
            "pv_ri": round(pv_ri, 4),
        })
        bvps = bvps + eps  # book value grows by reinvested earnings

    last_ri = schedule[-1]["ri"]
    n = len(earnings_per_share_list)

    if persistence_factor == 0 or abs(last_ri) < 1e-12:
        pv_continuing_ri = 0.0
    elif continuing_ri_growth > 0 and cost_of_equity > continuing_ri_growth:
        continuing_ri_value = last_ri * (1 + continuing_ri_growth) / (cost_of_equity - continuing_ri_growth)
        pv_continuing_ri = continuing_ri_value / (1 + cost_of_equity) ** n
    else:
        pv_continuing_ri = (last_ri * persistence_factor / cost_of_equity) / (1 + cost_of_equity) ** n

    intrinsic_value = book_value_per_share + pv_ri_sum + pv_continuing_ri

    return json.dumps({
        "model_type": "residual_income",
        "intrinsic_value": round(intrinsic_value, 4),
        "book_value_per_share": book_value_per_share,
        "pv_explicit_ri": round(pv_ri_sum, 4),
        "pv_continuing_ri": round(pv_continuing_ri, 4),
        "premium_to_book": round(intrinsic_value - book_value_per_share, 4),
        "cost_of_equity_pct": round(cost_of_equity * 100, 2),
        "persistence_factor": persistence_factor,
        "schedule": schedule,
    })


# ── Multiples Valuation Tool ───────────────────────────────────────────────────

def build_multiples_model(
    db: Session,
    ebitda: float,
    net_income: float,
    book_value: float,
    revenue: float,
    shares_outstanding: float,
    net_debt: float,
    peer_ev_ebitda: float,
    peer_pe: float,
    peer_pb: float,
    peer_ps: float,
) -> str:
    """
    Peer comparables (multiples) valuation.

    Applies four CFA multiples: EV/EBITDA, P/E, P/B, P/S.
    """
    if shares_outstanding <= 0:
        return json.dumps({"error": "shares_outstanding must be positive"})

    multiples = {}
    prices = []

    if peer_ev_ebitda > 0 and ebitda > 0:
        implied_ev = ebitda * peer_ev_ebitda
        implied_equity = implied_ev - net_debt
        implied_price = implied_equity / shares_outstanding
        multiples["ev_ebitda"] = {
            "multiple": peer_ev_ebitda,
            "implied_ev": round(implied_ev, 2),
            "implied_equity": round(implied_equity, 2),
            "implied_price": round(implied_price, 2),
        }
        if implied_price > 0:
            prices.append(implied_price)

    if peer_pe > 0:
        eps = net_income / shares_outstanding
        implied_price = eps * peer_pe
        multiples["pe"] = {
            "multiple": peer_pe,
            "eps": round(eps, 4),
            "implied_price": round(implied_price, 2),
        }
        if implied_price > 0:
            prices.append(implied_price)

    if peer_pb > 0:
        bvps = book_value / shares_outstanding
        implied_price = bvps * peer_pb
        multiples["pb"] = {
            "multiple": peer_pb,
            "bvps": round(bvps, 4),
            "implied_price": round(implied_price, 2),
        }
        if implied_price > 0:
            prices.append(implied_price)

    if peer_ps > 0:
        rps = revenue / shares_outstanding
        implied_price = rps * peer_ps
        multiples["ps"] = {
            "multiple": peer_ps,
            "revenue_per_share": round(rps, 4),
            "implied_price": round(implied_price, 2),
        }
        if implied_price > 0:
            prices.append(implied_price)

    sorted_prices = sorted(prices)
    median = sorted_prices[len(sorted_prices) // 2] if sorted_prices else None

    return json.dumps({
        "model_type": "multiples",
        "multiples": multiples,
        "implied_price_min": round(min(prices), 2) if prices else None,
        "implied_price_max": round(max(prices), 2) if prices else None,
        "implied_price_median": round(median, 2) if median is not None else None,
    })


# ── FCFE Tool ─────────────────────────────────────────────────────────────────

def compute_fcfe(
    db: Session,
    net_income: Optional[float] = None,
    da: Optional[float] = None,
    capex: Optional[float] = None,
    delta_wc: Optional[float] = None,
    net_borrowing: Optional[float] = None,
    fcff: Optional[float] = None,
    interest_expense: Optional[float] = None,
    tax_rate: Optional[float] = None,
) -> str:
    """
    Compute Free Cash Flow to Equity (FCFE).

    Two paths:
    - Direct: FCFE = NI + D&A − CapEx − ΔNWC + Net Borrowing
    - From FCFF: FCFE = FCFF − Interest × (1 − T) + Net Borrowing
    """
    direct_params = [net_income, da, capex, delta_wc, net_borrowing]
    fcff_params = [fcff, interest_expense, tax_rate, net_borrowing]

    if all(p is not None for p in direct_params):
        fcfe_val = net_income + da - capex - delta_wc + net_borrowing
        return json.dumps({
            "model_type": "fcfe",
            "fcfe": round(fcfe_val, 4),
            "calculation_path": "direct",
            "net_income": net_income,
            "da": da,
            "capex": capex,
            "delta_wc": delta_wc,
            "net_borrowing": net_borrowing,
            "formula": "FCFE = NI + D&A − CapEx − ΔWC + Net Borrowing",
        })
    elif all(p is not None for p in fcff_params):
        after_tax_interest = interest_expense * (1 - tax_rate)
        fcfe_val = fcff - after_tax_interest + net_borrowing
        return json.dumps({
            "model_type": "fcfe",
            "fcfe": round(fcfe_val, 4),
            "calculation_path": "from_fcff",
            "fcff": fcff,
            "interest_expense": interest_expense,
            "tax_rate": tax_rate,
            "after_tax_interest": round(after_tax_interest, 4),
            "net_borrowing": net_borrowing,
            "formula": "FCFE = FCFF − Interest × (1 − T) + Net Borrowing",
        })
    else:
        return json.dumps({
            "error": (
                "Provide either (net_income, da, capex, delta_wc, net_borrowing) "
                "for direct calculation OR (fcff, interest_expense, tax_rate, net_borrowing) "
                "for FCFF-based calculation."
            )
        })


# ── Revenue Modeling Tool ─────────────────────────────────────────────────────

def build_revenue_model(
    db: Session,
    base_revenue: float,
    years: int,
    approach: str = "growth_rates",
    growth_rates: Optional[list] = None,
    market_size: Optional[float] = None,
    market_share_pct: Optional[float] = None,
    market_growth_rate: float = 0.0,
    units_sold: Optional[float] = None,
    price_per_unit: Optional[float] = None,
    volume_growth_rate: float = 0.0,
    price_growth_rate: float = 0.0,
) -> str:
    """
    Build a multi-year revenue projection.

    Three approaches:
    - 'growth_rates': compound base_revenue by a list of annual rates.
    - 'top_down': revenue = market_size × (1+market_growth_rate)^t × market_share_pct.
    - 'bottom_up': revenue = units_sold×(1+vol_g)^t × price_per_unit×(1+price_g)^t.
    """
    if years <= 0:
        return json.dumps({"error": "years must be positive"})

    projections = []

    if approach == "growth_rates":
        if not growth_rates:
            return json.dumps({"error": "growth_rates required for growth_rates approach"})
        if len(growth_rates) < years:
            return json.dumps({"error": f"growth_rates has {len(growth_rates)} items but years={years}"})
        revenue = base_revenue
        for t in range(years):
            g = growth_rates[t]
            revenue = revenue * (1 + g)
            projections.append({"year": t + 1, "revenue": round(revenue, 4), "growth_pct": round(g * 100, 4), "method": "growth_rates"})

    elif approach == "top_down":
        if market_size is None or market_share_pct is None:
            return json.dumps({"error": "market_size and market_share_pct required for top_down approach"})
        for t in range(1, years + 1):
            market = market_size * (1 + market_growth_rate) ** t
            revenue = market * market_share_pct
            prev_revenue = market_size * (1 + market_growth_rate) ** (t - 1) * market_share_pct
            growth_pct = (revenue / prev_revenue - 1) * 100 if prev_revenue else 0
            projections.append({"year": t, "revenue": round(revenue, 4), "market_size": round(market, 4), "growth_pct": round(growth_pct, 4), "method": "top_down"})

    elif approach == "bottom_up":
        if units_sold is None or price_per_unit is None:
            return json.dumps({"error": "units_sold and price_per_unit required for bottom_up approach"})
        for t in range(1, years + 1):
            units = units_sold * (1 + volume_growth_rate) ** t
            price = price_per_unit * (1 + price_growth_rate) ** t
            revenue = units * price
            prev_revenue = (
                units_sold * (1 + volume_growth_rate) ** (t - 1) *
                price_per_unit * (1 + price_growth_rate) ** (t - 1)
            )
            growth_pct = (revenue / prev_revenue - 1) * 100 if prev_revenue else 0
            projections.append({"year": t, "revenue": round(revenue, 4), "units": round(units, 4), "price": round(price, 4), "growth_pct": round(growth_pct, 4), "method": "bottom_up"})

    else:
        return json.dumps({"error": f"Unknown approach '{approach}'. Use 'growth_rates', 'top_down', or 'bottom_up'"})

    return json.dumps({
        "model_type": "revenue_model",
        "approach": approach,
        "projections": projections,
        "total_revenue": round(sum(p["revenue"] for p in projections), 4),
    })


# ── Working Capital Tool ──────────────────────────────────────────────────────

def build_wc_model(
    db: Session,
    revenue_list: list,
    cogs_pct: float,
    dso: float,
    dio: float,
    dpo: float,
    opening_nwc: float = 0.0,
) -> str:
    """
    Build a Working Capital model from revenue and day-metrics.

    Formulas:
        AR(t)        = (DSO / 365) × Revenue(t)
        Inventory(t) = (DIO / 365) × COGS(t)    where COGS = Revenue × cogs_pct
        AP(t)        = (DPO / 365) × COGS(t)
        NWC(t)       = AR(t) + Inventory(t) − AP(t)
        ΔWC(t)       = NWC(t) − NWC(t−1)        (positive = cash outflow in FCFF)
        CCC          = DSO + DIO − DPO
    """
    if not revenue_list:
        return json.dumps({"error": "revenue_list must not be empty"})
    if not (0 < cogs_pct <= 1):
        return json.dumps({"error": "cogs_pct must be between 0 and 1"})
    if dso < 0 or dio < 0 or dpo < 0:
        return json.dumps({"error": "dso, dio, dpo must be non-negative"})

    ccc = dso + dio - dpo
    projections = []
    prev_nwc = opening_nwc

    for t, revenue in enumerate(revenue_list, start=1):
        cogs = revenue * cogs_pct
        ar = (dso / 365) * revenue
        inventory = (dio / 365) * cogs
        ap = (dpo / 365) * cogs
        nwc = ar + inventory - ap
        delta_wc = nwc - prev_nwc
        projections.append({
            "year": t,
            "revenue": round(revenue, 4),
            "cogs": round(cogs, 4),
            "ar": round(ar, 4),
            "inventory": round(inventory, 4),
            "ap": round(ap, 4),
            "nwc": round(nwc, 4),
            "delta_wc": round(delta_wc, 4),
            "ccc": round(ccc, 4),
        })
        prev_nwc = nwc

    return json.dumps({"model_type": "wc_model", "dso": dso, "dio": dio, "dpo": dpo, "ccc": round(ccc, 4), "projections": projections})


# ── CapEx & PP&E Schedule Tool ────────────────────────────────────────────────

def build_capex_schedule(
    db: Session,
    opening_gross_ppe: float,
    opening_accum_dep: float,
    useful_life: float,
    years: int,
    capex_list: Optional[list] = None,
    capex_pct_revenue: Optional[float] = None,
    revenue_list: Optional[list] = None,
    disposals_list: Optional[list] = None,
) -> str:
    """
    Build a PP&E roll-forward and depreciation schedule.

    Roll-Forward:
        DA(t)          = Gross PP&E(t-1) / useful_life   (straight-line)
        Gross PP&E(t)  = Gross PP&E(t-1) + CapEx(t) − Disposals(t)
        Acc. Dep.(t)   = min(Acc. Dep.(t-1) + DA(t), Gross PP&E(t))
        Net PP&E(t)    = Gross PP&E(t) − Acc. Dep.(t)
    """
    if years <= 0:
        return json.dumps({"error": "years must be positive"})
    if useful_life <= 0:
        return json.dumps({"error": "useful_life must be positive"})

    if capex_list is not None:
        if len(capex_list) < years:
            return json.dumps({"error": f"capex_list has {len(capex_list)} items but years={years}"})
        capex_values = list(capex_list[:years])
    elif capex_pct_revenue is not None:
        if not revenue_list or len(revenue_list) < years:
            return json.dumps({"error": "revenue_list required (length >= years) when using capex_pct_revenue"})
        capex_values = [revenue_list[t] * capex_pct_revenue for t in range(years)]
    else:
        return json.dumps({"error": "Provide either capex_list or (capex_pct_revenue + revenue_list)"})

    disposals = list(disposals_list[:years]) if disposals_list else [0.0] * years

    projections = []
    gross_ppe = opening_gross_ppe
    accum_dep = opening_accum_dep

    for t in range(years):
        capex = capex_values[t]
        disposal = disposals[t]
        da = gross_ppe / useful_life
        gross_ppe_new = gross_ppe + capex - disposal
        accum_dep_new = min(accum_dep + da, gross_ppe_new)
        net_ppe = max(0.0, gross_ppe_new - accum_dep_new)

        projections.append({
            "year": t + 1,
            "capex": round(capex, 4),
            "disposals": round(disposal, 4),
            "da": round(da, 4),
            "gross_ppe": round(gross_ppe_new, 4),
            "accum_dep": round(accum_dep_new, 4),
            "net_ppe": round(net_ppe, 4),
        })
        gross_ppe = gross_ppe_new
        accum_dep = accum_dep_new

    return json.dumps({"model_type": "capex_schedule", "useful_life": useful_life, "projections": projections})


# ── Debt Schedule Tool ────────────────────────────────────────────────────────

def build_debt_schedule(
    db: Session,
    tranches: list,
    years: int,
    cash_list: Optional[list] = None,
) -> str:
    """
    Build a multi-tranche debt schedule with interest expense.

    Roll-Forward per tranche per year:
        Amortization(t) = Opening(t) × amortization_pct
        Ending(t)       = max(0, Opening(t) − Amortization(t))
        Interest(t)     = (Opening(t) + Ending(t)) / 2 × annual_rate
    """
    if not tranches:
        return json.dumps({"error": "tranches must not be empty"})
    if years <= 0:
        return json.dumps({"error": "years must be positive"})

    cash = list(cash_list) if cash_list else [0.0] * years
    if len(cash) < years:
        cash = cash + [0.0] * (years - len(cash))

    balances = [t["opening_balance"] for t in tranches]
    projections = []

    for yr in range(years):
        year_total_debt = 0.0
        year_interest = 0.0
        tranche_detail = []
        new_balances = []

        for i, tranche in enumerate(tranches):
            opening = balances[i]
            amort_pct = tranche.get("amortization_pct", 0.0)
            rate = tranche["annual_rate"]
            amortization = opening * amort_pct
            ending = max(0.0, opening - amortization)
            interest = (opening + ending) / 2 * rate
            year_total_debt += ending
            year_interest += interest
            tranche_detail.append({"name": tranche["name"], "opening": round(opening, 4), "amortization": round(amortization, 4), "ending": round(ending, 4), "interest": round(interest, 4)})
            new_balances.append(ending)

        balances = new_balances
        net_debt = year_total_debt - cash[yr]
        projections.append({
            "year": yr + 1,
            "total_debt": round(year_total_debt, 4),
            "interest_expense": round(year_interest, 4),
            "cash": round(cash[yr], 4),
            "net_debt": round(net_debt, 4),
            "tranches": tranche_detail,
        })

    return json.dumps({"model_type": "debt_schedule", "projections": projections})


# ── Three Financial Statements Tool ───────────────────────────────────────────

def build_three_statement_model(
    db: Session,
    revenue_list: list,
    ebit_list: list,
    interest_expense_list: list,
    tax_rate: float,
    da_list: list,
    capex_list: list,
    total_debt_list: list,
    net_borrowing_list: list,
    opening_bs: dict,
    ar_list: Optional[list] = None,
    inventory_list: Optional[list] = None,
    ap_list: Optional[list] = None,
    dividend_payout_ratio: float = 0.0,
) -> str:
    """
    Build linked Income Statement, Cash Flow Statement, and Balance Sheet.

    Linkages:
        IS → CFS: Net Income starts Operating CF.
        IS → BS:  NI adds to Equity (retained earnings roll-forward).
        CFS → BS: Cash(t) = Cash(t-1) + net_change_in_cash.
        CapEx/DA → BS: PP&E(t) = PP&E(t-1) + CapEx − DA.
        Debt → BS: total_debt_list drives the debt balance per year.
        WC → CFS+BS: delta_wc computed internally from ar/inv/ap for consistency.

    The balance sheet always balances (Assets = L+E) because cash is the CFS residual.

    Args:
        revenue_list: Revenue per year (from build_pl_model or build_revenue_model).
        ebit_list: EBIT per year (from build_pl_model).
        interest_expense_list: Interest expense per year (from build_debt_schedule).
        tax_rate: Corporate tax rate decimal (constant).
        da_list: D&A per year (from build_capex_schedule).
        capex_list: CapEx per year (from build_capex_schedule).
        total_debt_list: Total debt balance per year (from build_debt_schedule).
        net_borrowing_list: Net new debt per year; positive=drawdown, negative=repayment.
        opening_bs: Opening balance sheet. Required keys: cash, ppe_net, other_assets,
            other_liabilities, equity. Optional: opening_ar, opening_inventory, opening_ap.
        ar_list: AR per year (from build_wc_model). Optional.
        inventory_list: Inventory per year (from build_wc_model). Optional.
        ap_list: AP per year (from build_wc_model). Optional.
        dividend_payout_ratio: Fraction of positive NI paid as dividends. Default 0.

    Returns:
        JSON with model_type and years list. Each year: income_statement,
        cash_flow_statement, balance_sheet (with balance_check_passed).
    """
    n = len(revenue_list)

    required = {
        "ebit_list": ebit_list, "interest_expense_list": interest_expense_list,
        "da_list": da_list, "capex_list": capex_list,
        "total_debt_list": total_debt_list, "net_borrowing_list": net_borrowing_list,
    }
    for name, lst in required.items():
        if len(lst) != n:
            return json.dumps({"error": f"{name} has {len(lst)} items but revenue_list has {n}"})

    for name, lst in [("ar_list", ar_list), ("inventory_list", inventory_list), ("ap_list", ap_list)]:
        if lst is not None and len(lst) != n:
            return json.dumps({"error": f"{name} has {len(lst)} items but revenue_list has {n}"})

    # Opening balance sheet state
    cash = float(opening_bs.get("cash", 0.0))
    ppe_net = float(opening_bs.get("ppe_net", 0.0))
    other_assets = float(opening_bs.get("other_assets", 0.0))
    other_liabilities = float(opening_bs.get("other_liabilities", 0.0))
    equity = float(opening_bs.get("equity", 0.0))

    # Opening WC for delta_wc in year 1
    prev_nwc = (
        float(opening_bs.get("opening_ar", 0.0)) +
        float(opening_bs.get("opening_inventory", 0.0)) -
        float(opening_bs.get("opening_ap", 0.0))
    )

    years = []

    for t in range(n):
        # Income Statement
        ebit = ebit_list[t]
        interest = interest_expense_list[t]
        ebt = ebit - interest
        tax = max(0.0, ebt * tax_rate)
        net_income = ebt - tax
        dividends = max(0.0, net_income * dividend_payout_ratio) if net_income > 0 else 0.0

        # Working Capital (compute delta_wc internally for BS consistency)
        ar = float(ar_list[t]) if ar_list else 0.0
        inv = float(inventory_list[t]) if inventory_list else 0.0
        ap = float(ap_list[t]) if ap_list else 0.0
        nwc = ar + inv - ap
        delta_wc = nwc - prev_nwc
        prev_nwc = nwc

        # Cash Flow Statement
        da = da_list[t]
        capex = capex_list[t]
        net_borrowing = net_borrowing_list[t]
        operating_cf = net_income + da - delta_wc
        investing_cf = -capex
        financing_cf = net_borrowing - dividends
        net_change = operating_cf + investing_cf + financing_cf

        # Balance Sheet (roll forward)
        cash_new = cash + net_change
        ppe_net_new = ppe_net + capex - da
        total_assets = cash_new + ar + inv + ppe_net_new + other_assets
        total_debt = total_debt_list[t]
        total_liabilities = ap + total_debt + other_liabilities
        equity_new = equity + net_income - dividends
        total_le = total_liabilities + equity_new
        balance_error = abs(total_assets - total_le)

        years.append({
            "year": t + 1,
            "income_statement": {
                "revenue": round(revenue_list[t], 4),
                "ebit": round(ebit, 4),
                "interest_expense": round(interest, 4),
                "ebt": round(ebt, 4),
                "tax": round(tax, 4),
                "net_income": round(net_income, 4),
                "dividends": round(dividends, 4),
                "retained_earnings_addition": round(net_income - dividends, 4),
            },
            "cash_flow_statement": {
                "da": round(da, 4),
                "delta_wc": round(delta_wc, 4),
                "operating_cf": round(operating_cf, 4),
                "capex": round(capex, 4),
                "investing_cf": round(investing_cf, 4),
                "net_borrowing": round(net_borrowing, 4),
                "dividends": round(dividends, 4),
                "financing_cf": round(financing_cf, 4),
                "net_change_in_cash": round(net_change, 4),
            },
            "balance_sheet": {
                "cash": round(cash_new, 4),
                "ar": round(ar, 4),
                "inventory": round(inv, 4),
                "ppe_net": round(ppe_net_new, 4),
                "other_assets": round(other_assets, 4),
                "total_assets": round(total_assets, 4),
                "ap": round(ap, 4),
                "total_debt": round(total_debt, 4),
                "other_liabilities": round(other_liabilities, 4),
                "total_liabilities": round(total_liabilities, 4),
                "equity": round(equity_new, 4),
                "total_liabilities_and_equity": round(total_le, 4),
                "balance_check_passed": balance_error < 0.01,
                "balance_error": round(balance_error, 6),
            },
        })

        cash = cash_new
        ppe_net = ppe_net_new
        equity = equity_new

    # ── Excel workbook ──────────────────────────────────────────────────────
    download_url = None
    if EXCEL_AVAILABLE:
        try:
            wb = _build_three_statement_workbook(
                years=years, opening_bs=opening_bs, tax_rate=tax_rate,
                dividend_payout_ratio=dividend_payout_ratio,
                revenue_list=revenue_list, ebit_list=ebit_list,
                interest_expense_list=interest_expense_list,
                da_list=da_list, capex_list=capex_list,
                total_debt_list=total_debt_list,
                net_borrowing_list=net_borrowing_list,
                ar_list=ar_list, inventory_list=inventory_list, ap_list=ap_list,
            )
            file_id = _save_excel(wb, "ThreeStatement")
            if file_id:
                download_url = f"/api/financial-modeling/download/{file_id}"
        except Exception as e:
            logger.warning("Three-statement Excel creation failed: %s", e)

    return json.dumps({
        "model_type": "three_statement_model",
        "years": years,
        "download_url": download_url,
    })


def _build_three_statement_workbook(
    years: list[dict],
    opening_bs: dict,
    tax_rate: float,
    dividend_payout_ratio: float = 0.0,
    revenue_list: list | None = None,
    ebit_list: list | None = None,
    interest_expense_list: list | None = None,
    da_list: list | None = None,
    capex_list: list | None = None,
    total_debt_list: list | None = None,
    net_borrowing_list: list | None = None,
    ar_list: list | None = None,
    inventory_list: list | None = None,
    ap_list: list | None = None,
):
    """Build a 5-sheet Excel workbook with formulas: Assumptions, Summary, IS, BS, CFS."""
    wb = openpyxl.Workbook()
    n = len(years)

    # ── Sheet 1: Assumptions (editable inputs) ───────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.sheet_view.rightToLeft = True

    ws_a["A1"] = "Three-Statement Model — Assumptions"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a.merge_cells("A1:C1")

    # Section A: Key rates (rows 3-4)
    _style_header(ws_a.cell(row=3, column=1, value="Parameter"))
    _style_header(ws_a.cell(row=3, column=2, value="Value"))

    assumptions = [
        ("Tax Rate", tax_rate),                          # B4
        ("Dividend Payout Ratio", dividend_payout_ratio), # B5
    ]
    for r, (label, val) in enumerate(assumptions, 4):
        ws_a.cell(row=r, column=1, value=label)
        _style_input(ws_a.cell(row=r, column=2, value=val))

    # Section B: Opening Balance Sheet (rows 7+)
    ws_a.cell(row=7, column=1, value="Opening Balance Sheet")
    ws_a.cell(row=7, column=1).font = Font(bold=True, size=12)

    _style_header(ws_a.cell(row=8, column=1, value="Item"))
    _style_header(ws_a.cell(row=8, column=2, value="Value"))

    opening_items = [
        ("Cash", float(opening_bs.get("cash", 0))),                      # B9
        ("Accounts Receivable", float(opening_bs.get("opening_ar", 0))),  # B10
        ("Inventory", float(opening_bs.get("opening_inventory", 0))),     # B11
        ("PP&E (Net)", float(opening_bs.get("ppe_net", 0))),              # B12
        ("Other Assets", float(opening_bs.get("other_assets", 0))),       # B13
        ("Accounts Payable", float(opening_bs.get("opening_ap", 0))),     # B14
        ("Other Liabilities", float(opening_bs.get("other_liabilities", 0))),  # B15
        ("Equity", float(opening_bs.get("equity", 0))),                   # B16
    ]
    for r, (label, val) in enumerate(opening_items, 9):
        ws_a.cell(row=r, column=1, value=label)
        _style_input(ws_a.cell(row=r, column=2, value=val))

    # Computed opening totals with formulas
    ws_a.cell(row=17, column=1, value="Total Assets")
    _style_formula(ws_a.cell(row=17, column=2, value="=SUM(B9:B13)"))  # B17
    ws_a.cell(row=18, column=1, value="Total Liabilities")
    _style_formula(ws_a.cell(row=18, column=2, value="=B14+B15"))      # B18
    ws_a.cell(row=19, column=1, value="Total L + E")
    _style_result(ws_a.cell(row=19, column=2, value="=B18+B16"))       # B19

    # Section C: Annual inputs (rows 22+)
    ws_a.cell(row=21, column=1, value="Annual Inputs")
    ws_a.cell(row=21, column=1).font = Font(bold=True, size=12)

    _style_header(ws_a.cell(row=22, column=1, value="Item"))
    for i in range(n):
        _style_header(ws_a.cell(row=22, column=i + 2, value=f"Year {i+1}"))

    # Annual input rows — row numbers: Revenue=23, EBIT=24, Interest=25,
    # D&A=26, CapEx=27, AR=28, Inventory=29, AP=30, Total Debt=31, Net Borrowing=32
    annual_inputs = [
        ("Revenue",           revenue_list or [0] * n),
        ("EBIT",              ebit_list or [0] * n),
        ("Interest Expense",  interest_expense_list or [0] * n),
        ("D&A",               da_list or [0] * n),
        ("CapEx",             capex_list or [0] * n),
        ("Accounts Receivable", ar_list or [0] * n),
        ("Inventory",         inventory_list or [0] * n),
        ("Accounts Payable",  ap_list or [0] * n),
        ("Total Debt",        total_debt_list or [0] * n),
        ("Net Borrowing",     net_borrowing_list or [0] * n),
    ]
    for r, (label, vals) in enumerate(annual_inputs, 23):
        ws_a.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, 2):
            _style_input(ws_a.cell(row=r, column=c, value=float(v)))

    _auto_width(ws_a)

    # ── Row references for Assumptions sheet ─────────────────────────────
    # Rates: B4=tax_rate, B5=dividend_payout
    # Opening BS: B9=cash, B10=AR, B11=inv, B12=PPE, B13=other_assets,
    #             B14=AP, B15=other_liab, B16=equity
    # Annual: row 23=Revenue, 24=EBIT, 25=Interest, 26=D&A, 27=CapEx,
    #         28=AR, 29=Inv, 30=AP, 31=TotalDebt, 32=NetBorrowing

    def _acol(year_idx):
        """Column letter for year_idx (0-based) in Assumptions annual section."""
        return get_column_letter(year_idx + 2)

    # ── Sheet 2: Income Statement ────────────────────────────────────────
    ws_is = wb.create_sheet("Income Statement")
    ws_is.sheet_view.rightToLeft = True

    ws_is["A1"] = "Income Statement"
    ws_is["A1"].font = Font(bold=True, size=14)
    ws_is.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)

    _style_header(ws_is.cell(row=3, column=1, value="Item"))
    for i in range(n):
        _style_header(ws_is.cell(row=3, column=i + 2, value=f"Year {i+1}"))

    # IS rows: 4=Revenue, 5=EBIT, 6=Interest, 7=EBT, 8=Tax, 9=NI, 10=Dividends, 11=RetEarnings
    is_labels = [
        (4, "Revenue"),
        (5, "EBIT"),
        (6, "Interest Expense"),
        (7, "EBT (Earnings Before Tax)"),
        (8, "Tax"),
        (9, "Net Income"),
        (10, "Dividends"),
        (11, "Retained Earnings Addition"),
    ]
    for row, label in is_labels:
        ws_is.cell(row=row, column=1, value=label)

    for i in range(n):
        cl = get_column_letter(i + 2)
        acl = _acol(i)

        # Row 4: Revenue = Assumptions!annual row 23
        c = ws_is.cell(row=4, column=i + 2, value=f"=Assumptions!{acl}23")
        _style_input(c)

        # Row 5: EBIT = Assumptions!annual row 24
        c = ws_is.cell(row=5, column=i + 2, value=f"=Assumptions!{acl}24")
        _style_input(c)

        # Row 6: Interest = Assumptions!annual row 25
        c = ws_is.cell(row=6, column=i + 2, value=f"=Assumptions!{acl}25")
        _style_input(c)

        # Row 7: EBT = EBIT - Interest
        c = ws_is.cell(row=7, column=i + 2, value=f"={cl}5-{cl}6")
        _style_formula(c)

        # Row 8: Tax = MAX(0, EBT * TaxRate)
        c = ws_is.cell(row=8, column=i + 2, value=f"=MAX(0,{cl}7*Assumptions!$B$4)")
        _style_formula(c)

        # Row 9: Net Income = EBT - Tax
        c = ws_is.cell(row=9, column=i + 2, value=f"={cl}7-{cl}8")
        _style_result(c)

        # Row 10: Dividends = MAX(0, NI * DividendPayout) if NI>0
        c = ws_is.cell(row=10, column=i + 2,
                        value=f"=IF({cl}9>0,MAX(0,{cl}9*Assumptions!$B$5),0)")
        _style_formula(c)

        # Row 11: Retained Earnings Addition = NI - Dividends
        c = ws_is.cell(row=11, column=i + 2, value=f"={cl}9-{cl}10")
        _style_result(c)

    _auto_width(ws_is)

    # ── Sheet 3: Cash Flow Statement ─────────────────────────────────────
    ws_cf = wb.create_sheet("Cash Flow Statement")
    ws_cf.sheet_view.rightToLeft = True

    ws_cf["A1"] = "Cash Flow Statement"
    ws_cf["A1"].font = Font(bold=True, size=14)
    ws_cf.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)

    _style_header(ws_cf.cell(row=3, column=1, value="Item"))
    for i in range(n):
        _style_header(ws_cf.cell(row=3, column=i + 2, value=f"Year {i+1}"))

    # CFS rows:
    # 4=header OPERATING, 5=NI, 6=D&A, 7=DeltaWC, 8=OperatingCF
    # 9=spacer, 10=header INVESTING, 11=CapEx, 12=InvestingCF
    # 13=spacer, 14=header FINANCING, 15=NetBorrowing, 16=DividendsPaid, 17=FinancingCF
    # 18=spacer, 19=NetChange, 20=EndingCash
    cf_labels = {
        4: ("OPERATING", True),
        5: ("Net Income", False),
        6: ("+ D&A", False),
        7: ("- Change in Working Capital", False),
        8: ("Operating Cash Flow", False),
        10: ("INVESTING", True),
        11: ("Capital Expenditure", False),
        12: ("Investing Cash Flow", False),
        14: ("FINANCING", True),
        15: ("Net Borrowing", False),
        16: ("Dividends Paid", False),
        17: ("Financing Cash Flow", False),
        19: ("Net Change in Cash", False),
        20: ("Ending Cash Balance", False),
    }
    for row, (label, is_header) in cf_labels.items():
        c = ws_cf.cell(row=row, column=1, value=label)
        if is_header:
            c.font = Font(bold=True)

    for i in range(n):
        cl = get_column_letter(i + 2)
        acl = _acol(i)
        is_cl = cl  # same column letter in IS sheet

        # Row 5: Net Income = 'Income Statement'!row9
        c = ws_cf.cell(row=5, column=i + 2,
                        value=f"='Income Statement'!{is_cl}9")
        _style_formula(c)

        # Row 6: D&A = Assumptions!annual row 26
        c = ws_cf.cell(row=6, column=i + 2, value=f"=Assumptions!{acl}26")
        _style_input(c)

        # Row 7: Delta WC — computed as change in (AR+Inv-AP)
        # Year 1: NWC_1 - NWC_opening where opening NWC = B10+B11-B14
        # Year t>1: NWC_t - NWC_(t-1)  — use Assumptions annual rows 28,29,30
        if i == 0:
            # NWC_1 = Assumptions!(B28+B29-B30), NWC_0 = Assumptions!(B10+B11-B14)
            formula = (
                f"=(Assumptions!{acl}28+Assumptions!{acl}29-Assumptions!{acl}30)"
                f"-(Assumptions!$B$10+Assumptions!$B$11-Assumptions!$B$14)"
            )
        else:
            prev_cl = _acol(i - 1)
            formula = (
                f"=(Assumptions!{acl}28+Assumptions!{acl}29-Assumptions!{acl}30)"
                f"-(Assumptions!{prev_cl}28+Assumptions!{prev_cl}29-Assumptions!{prev_cl}30)"
            )
        c = ws_cf.cell(row=7, column=i + 2, value=formula)
        _style_formula(c)

        # Row 8: Operating CF = NI + D&A - DeltaWC
        c = ws_cf.cell(row=8, column=i + 2, value=f"={cl}5+{cl}6-{cl}7")
        _style_result(c)

        # Row 11: CapEx (negative) = -Assumptions!annual row 27
        c = ws_cf.cell(row=11, column=i + 2, value=f"=-Assumptions!{acl}27")
        _style_formula(c)

        # Row 12: Investing CF = CapEx (already negative)
        c = ws_cf.cell(row=12, column=i + 2, value=f"={cl}11")
        _style_result(c)

        # Row 15: Net Borrowing = Assumptions!annual row 32
        c = ws_cf.cell(row=15, column=i + 2, value=f"=Assumptions!{acl}32")
        _style_input(c)

        # Row 16: Dividends Paid = -'Income Statement'!row10
        c = ws_cf.cell(row=16, column=i + 2,
                        value=f"=-'Income Statement'!{is_cl}10")
        _style_formula(c)

        # Row 17: Financing CF = NetBorrowing + DividendsPaid
        c = ws_cf.cell(row=17, column=i + 2, value=f"={cl}15+{cl}16")
        _style_result(c)

        # Row 19: Net Change = Operating + Investing + Financing
        c = ws_cf.cell(row=19, column=i + 2, value=f"={cl}8+{cl}12+{cl}17")
        _style_result(c)

        # Row 20: Ending Cash = previous cash + net change
        if i == 0:
            c = ws_cf.cell(row=20, column=i + 2,
                            value=f"=Assumptions!$B$9+{cl}19")
        else:
            prev_cl = get_column_letter(i + 1)
            c = ws_cf.cell(row=20, column=i + 2,
                            value=f"={prev_cl}20+{cl}19")
        _style_result(c)

    _auto_width(ws_cf)

    # ── Sheet 4: Balance Sheet ───────────────────────────────────────────
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.sheet_view.rightToLeft = True

    ws_bs["A1"] = "Balance Sheet"
    ws_bs["A1"].font = Font(bold=True, size=14)
    ws_bs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 2)

    _style_header(ws_bs.cell(row=3, column=1, value="Item"))
    _style_header(ws_bs.cell(row=3, column=2, value="Opening"))
    for i in range(n):
        _style_header(ws_bs.cell(row=3, column=i + 3, value=f"Year {i+1}"))

    # BS rows layout:
    # 4=header ASSETS, 5=Cash, 6=AR, 7=Inv, 8=PPE, 9=OtherAssets, 10=TotalAssets
    # 11=spacer, 12=header LIAB&EQ, 13=AP, 14=TotalDebt, 15=OtherLiab,
    # 16=TotalLiab, 17=Equity, 18=TotalL+E, 19=BalanceCheck
    bs_labels = {
        4: ("ASSETS", True),
        5: ("Cash", False),
        6: ("Accounts Receivable", False),
        7: ("Inventory", False),
        8: ("PP&E (Net)", False),
        9: ("Other Assets", False),
        10: ("Total Assets", False),
        12: ("LIABILITIES & EQUITY", True),
        13: ("Accounts Payable", False),
        14: ("Total Debt", False),
        15: ("Other Liabilities", False),
        16: ("Total Liabilities", False),
        17: ("Equity", False),
        18: ("Total L + E", False),
        19: ("Balance Check (A - L&E)", False),
    }
    for row, (label, is_header) in bs_labels.items():
        c = ws_bs.cell(row=row, column=1, value=label)
        if is_header:
            c.font = Font(bold=True)

    # Opening column (col 2) — formulas referencing Assumptions sheet
    opening_formulas = {
        5: "=Assumptions!$B$9",    # Cash
        6: "=Assumptions!$B$10",   # AR
        7: "=Assumptions!$B$11",   # Inventory
        8: "=Assumptions!$B$12",   # PPE
        9: "=Assumptions!$B$13",   # Other Assets
        10: "=SUM(B5:B9)",         # Total Assets
        13: "=Assumptions!$B$14",  # AP
        14: 0,                     # Total Debt (opening = 0 or user can edit)
        15: "=Assumptions!$B$15",  # Other Liabilities
        16: "=B13+B14+B15",        # Total Liabilities
        17: "=Assumptions!$B$16",  # Equity
        18: "=B16+B17",            # Total L+E
        19: "=B10-B18",            # Balance check
    }
    for row, val in opening_formulas.items():
        c = ws_bs.cell(row=row, column=2, value=val)
        if row in (10, 18, 19):
            _style_result(c)
        elif isinstance(val, str) and val.startswith("="):
            _style_formula(c)
        else:
            _style_input(c)

    # Year columns (col 3..n+2) — formulas
    for i in range(n):
        col = i + 3
        cl = get_column_letter(col)
        prev_cl = get_column_letter(col - 1)
        acl = _acol(i)
        cf_cl = get_column_letter(i + 2)  # CFS uses col i+2

        # Row 5: Cash = 'Cash Flow Statement'!row20
        c = ws_bs.cell(row=5, column=col,
                        value=f"='Cash Flow Statement'!{cf_cl}20")
        _style_formula(c)

        # Row 6: AR = Assumptions!annual row 28
        c = ws_bs.cell(row=6, column=col, value=f"=Assumptions!{acl}28")
        _style_input(c)

        # Row 7: Inventory = Assumptions!annual row 29
        c = ws_bs.cell(row=7, column=col, value=f"=Assumptions!{acl}29")
        _style_input(c)

        # Row 8: PPE = previous PPE + CapEx - D&A
        # (CapEx adds to PPE, D&A reduces it)
        c = ws_bs.cell(row=8, column=col,
                        value=f"={prev_cl}8+Assumptions!{acl}27-Assumptions!{acl}26")
        _style_formula(c)

        # Row 9: Other Assets = previous (constant)
        c = ws_bs.cell(row=9, column=col, value=f"={prev_cl}9")
        _style_input(c)

        # Row 10: Total Assets = SUM(Cash..OtherAssets)
        c = ws_bs.cell(row=10, column=col, value=f"=SUM({cl}5:{cl}9)")
        _style_result(c)

        # Row 13: AP = Assumptions!annual row 30
        c = ws_bs.cell(row=13, column=col, value=f"=Assumptions!{acl}30")
        _style_input(c)

        # Row 14: Total Debt = Assumptions!annual row 31
        c = ws_bs.cell(row=14, column=col, value=f"=Assumptions!{acl}31")
        _style_input(c)

        # Row 15: Other Liabilities = previous (constant)
        c = ws_bs.cell(row=15, column=col, value=f"={prev_cl}15")
        _style_input(c)

        # Row 16: Total Liabilities = AP + Debt + OtherLiab
        c = ws_bs.cell(row=16, column=col, value=f"={cl}13+{cl}14+{cl}15")
        _style_formula(c)

        # Row 17: Equity = previous equity + retained earnings from IS
        c = ws_bs.cell(row=17, column=col,
                        value=f"={prev_cl}17+'Income Statement'!{cf_cl}11")
        _style_result(c)

        # Row 18: Total L + E = Total Liab + Equity
        c = ws_bs.cell(row=18, column=col, value=f"={cl}16+{cl}17")
        _style_result(c)

        # Row 19: Balance Check = Total Assets - Total L+E (should be 0)
        c = ws_bs.cell(row=19, column=col, value=f"={cl}10-{cl}18")
        _style_result(c)

    _auto_width(ws_bs)

    # ── Sheet 5: Summary (cross-sheet formula references) ────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.sheet_view.rightToLeft = True

    ws_s["A1"] = "Financial Model — Three-Statement Summary"
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)

    _style_header(ws_s.cell(row=3, column=1, value="Item"))
    for i in range(n):
        _style_header(ws_s.cell(row=3, column=i + 2, value=f"Year {i+1}"))

    # Summary rows referencing other sheets
    # IS refs: row4=Revenue, row5=EBIT, row9=NI
    # CFS refs: row8=OpCF, row12=InvCF, row17=FinCF, row19=NetChange
    # BS refs (col offset +1): row10=TotalAssets, row16=TotalLiab, row17=Equity, row19=BalCheck
    summary_defs = [
        (4,  "Revenue",           "'Income Statement'!{cl}4"),
        (5,  "EBIT",              "'Income Statement'!{cl}5"),
        (6,  "Net Income",        "'Income Statement'!{cl}9"),
        # spacer row 7
        (8,  "Operating CF",      "'Cash Flow Statement'!{cl}8"),
        (9,  "Investing CF",      "'Cash Flow Statement'!{cl}12"),
        (10, "Financing CF",      "'Cash Flow Statement'!{cl}17"),
        (11, "Net Change in Cash","'Cash Flow Statement'!{cl}19"),
        # spacer row 12
        (13, "Total Assets",      "'Balance Sheet'!{bcl}10"),
        (14, "Total Liabilities", "'Balance Sheet'!{bcl}16"),
        (15, "Equity",            "'Balance Sheet'!{bcl}17"),
        (16, "Balance Check",     "'Balance Sheet'!{bcl}19"),
    ]

    for row, label, _ in summary_defs:
        ws_s.cell(row=row, column=1, value=label)

    for i in range(n):
        is_cf_cl = get_column_letter(i + 2)   # IS/CFS column
        bs_cl = get_column_letter(i + 3)       # BS column (offset by opening col)

        for row, label, tmpl in summary_defs:
            formula = "=" + tmpl.format(cl=is_cf_cl, bcl=bs_cl)
            c = ws_s.cell(row=row, column=i + 2, value=formula)
            if label in ("Net Income", "Net Change in Cash", "Equity"):
                _style_result(c)
            elif label == "Balance Check":
                _style_result(c)
            else:
                _style_formula(c)

    _auto_width(ws_s)

    return wb


# ── Phase 5: Advanced Wall Street Tools ──────────────────────────────────────

def build_lbo_model(
    db: Session,
    entry_ebitda: float,
    entry_multiple: float,
    debt_pct: float,
    interest_rate: float,
    ebitda_growth_rates: list,
    exit_multiple: float,
    tax_rate: float = 0.25,
    da_pct_ebitda: float = 0.15,
    capex_pct_ebitda: float = 0.12,
    mandatory_amort_pct: float = 0.05,
) -> str:
    """
    Build a Leveraged Buyout (LBO) model.

    Mechanics:
        Entry EV = entry_ebitda × entry_multiple
        Entry Debt = Entry EV × debt_pct; Entry Equity = EV × (1 - debt_pct)
        Each year: EBITDA grows, interest accrues, FCF pays down debt (mandatory + sweep)
        Exit EV = Exit EBITDA × exit_multiple; Exit Equity = Exit EV − Remaining Debt
        MOIC = Exit Equity / Entry Equity
        IRR: solve 0 = −Entry Equity + Exit Equity / (1+IRR)^n  [no interim distributions]

    Returns attribution: EBITDA Growth effect, Multiple Expansion effect, Debt Paydown effect.
    """
    if not ebitda_growth_rates:
        return json.dumps({"error": "ebitda_growth_rates must not be empty"})
    if debt_pct <= 0 or debt_pct >= 1:
        return json.dumps({"error": "debt_pct must be between 0 and 1"})

    years = len(ebitda_growth_rates)
    entry_ev = entry_ebitda * entry_multiple
    entry_debt = entry_ev * debt_pct
    entry_equity = entry_ev * (1 - debt_pct)

    schedule = []
    debt = entry_debt
    ebitda = entry_ebitda
    mandatory_amort_annual = entry_debt * mandatory_amort_pct

    for t, g in enumerate(ebitda_growth_rates, start=1):
        ebitda = ebitda * (1 + g)
        da = ebitda * da_pct_ebitda
        ebit = ebitda - da
        interest = debt * interest_rate
        ebt = ebit - interest
        tax = max(0.0, ebt * tax_rate)
        net_income = ebt - tax
        capex = ebitda * capex_pct_ebitda
        fcf = net_income + da - capex

        # Debt paydown: mandatory + cash sweep (excess FCF)
        mandatory = min(mandatory_amort_annual, debt)
        sweep = max(0.0, fcf - mandatory)
        total_paydown = min(mandatory + sweep, debt)
        debt = max(0.0, debt - total_paydown)

        schedule.append({
            "year": t,
            "ebitda": round(ebitda, 4),
            "da": round(da, 4),
            "ebit": round(ebit, 4),
            "interest": round(interest, 4),
            "net_income": round(net_income, 4),
            "fcf": round(fcf, 4),
            "debt_paydown": round(total_paydown, 4),
            "remaining_debt": round(debt, 4),
        })

    exit_ebitda = ebitda
    exit_ev = exit_ebitda * exit_multiple
    exit_equity = max(0.0, exit_ev - debt)
    moic = exit_equity / entry_equity if entry_equity > 0 else 0

    # IRR: solve for r such that entry_equity = exit_equity / (1+r)^n
    irr = (exit_equity / entry_equity) ** (1.0 / years) - 1 if entry_equity > 0 and exit_equity > 0 else None

    # Returns attribution
    debt_paydown_effect = entry_debt - debt
    ebitda_growth_effect = (exit_ebitda - entry_ebitda) * exit_multiple
    multiple_expansion_effect = entry_ebitda * (exit_multiple - entry_multiple)

    return json.dumps({
        "model_type": "lbo",
        "entry": {"ev": round(entry_ev, 4), "debt": round(entry_debt, 4), "equity": round(entry_equity, 4), "ebitda": entry_ebitda, "multiple": entry_multiple},
        "exit": {"ev": round(exit_ev, 4), "debt": round(debt, 4), "equity": round(exit_equity, 4), "ebitda": round(exit_ebitda, 4), "multiple": exit_multiple},
        "returns": {"moic": round(moic, 4), "irr_pct": round(irr * 100, 2) if irr else None, "hold_years": years},
        "attribution": {
            "ebitda_growth_effect": round(ebitda_growth_effect, 4),
            "multiple_expansion_effect": round(multiple_expansion_effect, 4),
            "debt_paydown_effect": round(debt_paydown_effect, 4),
        },
        "schedule": schedule,
    })


def build_ma_model(
    db: Session,
    acquirer_net_income: float,
    acquirer_shares: float,
    acquirer_share_price: float,
    target_net_income: float,
    deal_value: float,
    cash_pct: float,
    tax_rate: float,
    cost_synergies: float = 0.0,
    revenue_synergies: float = 0.0,
    financing_rate: float = 0.08,
    target_book_equity: Optional[float] = None,
) -> str:
    """
    Build an M&A Accretion/Dilution model.

    Mechanics:
        Cash portion: financed with new debt → interest cost × (1-T) reduces NI
        Stock portion: new shares issued at acquirer share price → dilutes EPS
        Pro Forma NI = Acquirer NI + Target NI + Synergies - After-Tax Financing Cost
        Pro Forma Shares = Acquirer Shares + New Shares Issued (for stock portion)
        Accretion % = (Pro Forma EPS - Standalone EPS) / Standalone EPS × 100
        Goodwill = Deal Value - Target Book Equity (if provided)
    """
    if acquirer_shares <= 0:
        return json.dumps({"error": "acquirer_shares must be positive"})
    if not (0 <= cash_pct <= 1):
        return json.dumps({"error": "cash_pct must be between 0 and 1"})

    standalone_eps = acquirer_net_income / acquirer_shares

    # Financing structure
    cash_consideration = deal_value * cash_pct
    stock_consideration = deal_value * (1 - cash_pct)

    # New debt interest cost (after-tax)
    financing_cost_after_tax = cash_consideration * financing_rate * (1 - tax_rate)

    # New shares issued (stock portion at current acquirer price)
    new_shares_issued = stock_consideration / acquirer_share_price if acquirer_share_price > 0 else 0

    # Pro forma
    pro_forma_ni = acquirer_net_income + target_net_income + cost_synergies + revenue_synergies - financing_cost_after_tax
    pro_forma_shares = acquirer_shares + new_shares_issued
    pro_forma_eps = pro_forma_ni / pro_forma_shares if pro_forma_shares > 0 else 0

    accretion_pct = (pro_forma_eps - standalone_eps) / standalone_eps * 100 if standalone_eps else 0
    accretive = accretion_pct > 0

    # Goodwill
    if target_book_equity is None:
        target_book_equity = deal_value * 0.6
    goodwill = max(0.0, deal_value - target_book_equity)

    return json.dumps({
        "model_type": "ma_accretion_dilution",
        "standalone_eps": round(standalone_eps, 4),
        "pro_forma_eps": round(pro_forma_eps, 4),
        "accretion_pct": round(accretion_pct, 4),
        "accretive": accretive,
        "deal_structure": {
            "deal_value": deal_value,
            "cash_consideration": round(cash_consideration, 4),
            "stock_consideration": round(stock_consideration, 4),
            "new_shares_issued": round(new_shares_issued, 4),
            "pro_forma_shares": round(pro_forma_shares, 4),
        },
        "pro_forma_income": {
            "acquirer_ni": acquirer_net_income,
            "target_ni": target_net_income,
            "cost_synergies": cost_synergies,
            "revenue_synergies": revenue_synergies,
            "financing_cost_after_tax": round(financing_cost_after_tax, 4),
            "pro_forma_ni": round(pro_forma_ni, 4),
        },
        "goodwill": round(goodwill, 4),
    })


def compute_credit_metrics(
    db: Session,
    ebitda: float,
    interest_expense: float,
    total_debt: float,
    cash: float,
    capex: float = 0.0,
    mandatory_amortization: float = 0.0,
    revenue: Optional[float] = None,
    max_leverage_multiple: float = 4.0,
) -> str:
    """
    Compute key credit and leverage metrics.

    Metrics:
        Net Debt = Total Debt - Cash
        Net Leverage = Net Debt / EBITDA
        Interest Coverage = EBITDA / Interest
        DSCR = (EBITDA - CapEx) / (Interest + Mandatory Amort)
        Debt Capacity (at max_leverage × EBITDA)
    """
    if ebitda <= 0:
        return json.dumps({"error": "ebitda must be positive"})

    net_debt = total_debt - cash
    net_leverage = net_debt / ebitda
    gross_leverage = total_debt / ebitda
    interest_coverage = ebitda / interest_expense if interest_expense > 0 else None

    debt_service = interest_expense + mandatory_amortization
    dscr = (ebitda - capex) / debt_service if debt_service > 0 else None

    debt_capacity = ebitda * max_leverage_multiple
    debt_headroom = debt_capacity - total_debt

    result = {
        "model_type": "credit_metrics",
        "net_debt": round(net_debt, 4),
        "gross_leverage": round(gross_leverage, 4),
        "net_leverage": round(net_leverage, 4),
        "interest_coverage": round(interest_coverage, 4) if interest_coverage else None,
        "dscr": round(dscr, 4) if dscr else None,
        "debt_capacity": round(debt_capacity, 4),
        "debt_headroom": round(debt_headroom, 4),
        "credit_rating_proxy": (
            "Investment Grade" if net_leverage < 2.5 else
            "High Yield" if net_leverage < 5.0 else
            "Distressed"
        ),
    }
    if revenue:
        result["debt_to_revenue"] = round(total_debt / revenue, 4)
    return json.dumps(result)


def compute_liquidation_value(
    db: Session,
    assets: list,
    claims: list,
    admin_costs_pct: float = 0.05,
) -> str:
    """
    Compute liquidation value and creditor recovery in a distressed scenario.

    Apply recovery rates to each asset, deduct admin costs, then distribute
    through claims waterfall in priority order.

    Args:
        assets: [{name, book_value, recovery_rate}]
        claims: [{name, amount, priority}] (priority 1 = highest, e.g. secured)
        admin_costs_pct: Bankruptcy admin costs as fraction of liquidation proceeds.
    """
    if not assets:
        return json.dumps({"error": "assets list must not be empty"})
    if not claims:
        return json.dumps({"error": "claims list must not be empty"})

    # Compute liquidation proceeds
    asset_recoveries = []
    total_liquidation = 0.0
    for a in assets:
        recovery = a["book_value"] * a.get("recovery_rate", 0.5)
        total_liquidation += recovery
        asset_recoveries.append({
            "name": a["name"],
            "book_value": a["book_value"],
            "recovery_rate": a.get("recovery_rate", 0.5),
            "recovery_value": round(recovery, 4),
        })

    admin_costs = total_liquidation * admin_costs_pct
    net_proceeds = total_liquidation - admin_costs

    # Distribute through waterfall
    remaining = net_proceeds
    sorted_claims = sorted(claims, key=lambda c: c.get("priority", 999))
    waterfall = []
    for claim in sorted_claims:
        amount = claim["amount"]
        recovery = min(remaining, amount)
        recovery_pct = (recovery / amount * 100) if amount > 0 else 0
        remaining = max(0.0, remaining - recovery)
        waterfall.append({
            "name": claim["name"],
            "claim_amount": amount,
            "recovery_amount": round(recovery, 4),
            "recovery_pct": round(recovery_pct, 2),
            "priority": claim.get("priority", 999),
            "fulcrum": recovery < amount and remaining == 0,
        })

    # Identify fulcrum security (last class to get partial recovery)
    fulcrum_security = next((w["name"] for w in waterfall if w["fulcrum"]), None)

    return json.dumps({
        "model_type": "liquidation_value",
        "total_book_value": round(sum(a["book_value"] for a in assets), 4),
        "total_liquidation_value": round(total_liquidation, 4),
        "admin_costs": round(admin_costs, 4),
        "net_proceeds": round(net_proceeds, 4),
        "asset_recoveries": asset_recoveries,
        "waterfall": waterfall,
        "fulcrum_security": fulcrum_security,
    })


def compute_ipo_pricing(
    db: Session,
    shares_offered: float,
    offer_price: float,
    pre_ipo_shares: float,
    underwriting_discount_pct: float = 0.05,
    first_day_close: Optional[float] = None,
    greenshoe_pct: float = 0.15,
) -> str:
    """
    Analyze IPO economics: gross proceeds, net proceeds, underpricing, and greenshoe.

    Args:
        shares_offered: Shares in IPO (millions).
        offer_price: IPO offer price (IRR).
        pre_ipo_shares: Existing shares before IPO (millions).
        underwriting_discount_pct: Underwriting fee fraction. Default 5%.
        first_day_close: First-day closing price for underpricing calc. Optional.
        greenshoe_pct: Overallotment option fraction. Default 15%.
    """
    gross_proceeds = shares_offered * offer_price
    underwriting_fee = gross_proceeds * underwriting_discount_pct
    net_proceeds = gross_proceeds - underwriting_fee

    total_shares = pre_ipo_shares + shares_offered
    float_pct = shares_offered / total_shares * 100
    market_cap_at_offer = total_shares * offer_price

    greenshoe_shares = shares_offered * greenshoe_pct
    greenshoe_proceeds = greenshoe_shares * offer_price

    result = {
        "model_type": "ipo_pricing",
        "shares_offered": shares_offered,
        "offer_price": offer_price,
        "gross_proceeds": round(gross_proceeds, 4),
        "underwriting_fee": round(underwriting_fee, 4),
        "net_proceeds": round(net_proceeds, 4),
        "market_cap_at_offer": round(market_cap_at_offer, 4),
        "float_pct": round(float_pct, 2),
        "greenshoe": {
            "shares": round(greenshoe_shares, 4),
            "max_proceeds": round(greenshoe_proceeds, 4),
        },
    }

    if first_day_close is not None:
        underpricing_pct = (first_day_close - offer_price) / offer_price * 100
        money_left_on_table = (first_day_close - offer_price) * shares_offered
        market_cap_at_close = total_shares * first_day_close
        result.update({
            "first_day_close": first_day_close,
            "underpricing_pct": round(underpricing_pct, 2),
            "money_left_on_table": round(money_left_on_table, 4),
            "market_cap_at_close": round(market_cap_at_close, 4),
        })

    return json.dumps(result)


# ── Phase 6: Earnings Quality & FP&A Tools ───────────────────────────────────

def compute_altman_z(
    db: Session,
    working_capital: float,
    total_assets: float,
    retained_earnings: float,
    ebit: float,
    total_liabilities: float,
    revenue: float,
    market_cap: Optional[float] = None,
    book_equity: Optional[float] = None,
) -> str:
    """
    Compute Altman Z-Score for bankruptcy prediction.

    Public Z-Score: Z = 1.2×X1 + 1.4×X2 + 3.3×X3 + 0.6×X4 + 1.0×X5
        X1 = Working Capital / Total Assets
        X2 = Retained Earnings / Total Assets
        X3 = EBIT / Total Assets
        X4 = Market Cap / Total Liabilities (public)
        X5 = Revenue / Total Assets

    Private Z'-Score: uses Book Equity instead of Market Cap in X4.
    Zones: Public: Z>2.99=Safe, 1.81-2.99=Grey, <1.81=Distress
           Private: Z'>2.9=Safe, 1.23-2.9=Grey, <1.23=Distress
    """
    if total_assets <= 0:
        return json.dumps({"error": "total_assets must be positive"})
    if total_liabilities <= 0:
        return json.dumps({"error": "total_liabilities must be positive"})

    x1 = working_capital / total_assets
    x2 = retained_earnings / total_assets
    x3 = ebit / total_assets
    x5 = revenue / total_assets

    results = {}

    if market_cap is not None:
        x4_public = market_cap / total_liabilities
        z_public = 1.2 * x1 + 1.4 * x2 + 3.3 * x3 + 0.6 * x4_public + 1.0 * x5
        zone_public = "Safe" if z_public > 2.99 else "Grey Zone" if z_public > 1.81 else "Distress"
        results["public"] = {
            "z_score": round(z_public, 4),
            "zone": zone_public,
            "x4_market_cap_to_liabilities": round(x4_public, 4),
        }

    if book_equity is not None or market_cap is None:
        equity_for_private = book_equity if book_equity is not None else (total_assets - total_liabilities)
        x4_private = equity_for_private / total_liabilities
        z_prime = 0.717 * x1 + 0.847 * x2 + 3.107 * x3 + 0.420 * x4_private + 0.998 * x5
        zone_private = "Safe" if z_prime > 2.9 else "Grey Zone" if z_prime > 1.23 else "Distress"
        results["private"] = {
            "z_prime_score": round(z_prime, 4),
            "zone": zone_private,
            "x4_book_equity_to_liabilities": round(x4_private, 4),
        }

    return json.dumps({
        "model_type": "altman_z",
        "components": {
            "x1_working_capital_to_assets": round(x1, 4),
            "x2_retained_earnings_to_assets": round(x2, 4),
            "x3_ebit_to_assets": round(x3, 4),
            "x5_revenue_to_assets": round(x5, 4),
        },
        **results,
    })


def compute_beneish_score(
    db: Session,
    ar_current: float, sales_current: float,
    ar_prior: float, sales_prior: float,
    gross_profit_current: float, gross_profit_prior: float,
    current_assets_current: float, ppe_current: float, total_assets_current: float,
    current_assets_prior: float, ppe_prior: float, total_assets_prior: float,
    dep_current: float, dep_prior: float,
    sga_current: float, sga_prior: float,
    net_income: float, cfo: float,
    ltd_current: float, current_liab_current: float,
    ltd_prior: float, current_liab_prior: float,
) -> str:
    """
    Compute Beneish M-Score for earnings manipulation detection.

    M = -4.84 + 0.920(DSRI) + 0.528(GMI) + 0.404(AQI) + 0.892(SGI)
        + 0.115(DEPI) - 0.172(SGAI) + 4.679(TATA) - 0.327(LVGI)

    M > -1.78 suggests likely manipulation.
    """
    if sales_prior == 0 or total_assets_current == 0:
        return json.dumps({"error": "sales_prior and total_assets_current must be non-zero"})

    # DSRI: Days Sales Receivables Index
    dsri = (ar_current / sales_current) / (ar_prior / sales_prior) if (ar_prior / sales_prior) != 0 else 1

    # GMI: Gross Margin Index (prior / current — declining margin → GMI > 1)
    gm_current = gross_profit_current / sales_current if sales_current else 0
    gm_prior = gross_profit_prior / sales_prior if sales_prior else 0
    gmi = gm_prior / gm_current if gm_current != 0 else 1

    # AQI: Asset Quality Index
    non_current_current = 1 - (current_assets_current + ppe_current) / total_assets_current
    non_current_prior = 1 - (current_assets_prior + ppe_prior) / total_assets_prior if total_assets_prior else 0
    aqi = non_current_current / non_current_prior if non_current_prior != 0 else 1

    # SGI: Sales Growth Index
    sgi = sales_current / sales_prior

    # DEPI: Depreciation Index (lower dep rate → DEPI > 1)
    dep_rate_prior = dep_prior / (dep_prior + ppe_prior) if (dep_prior + ppe_prior) != 0 else 0
    dep_rate_current = dep_current / (dep_current + ppe_current) if (dep_current + ppe_current) != 0 else 0
    depi = dep_rate_prior / dep_rate_current if dep_rate_current != 0 else 1

    # SGAI: SG&A Index
    sga_ratio_current = sga_current / sales_current if sales_current else 0
    sga_ratio_prior = sga_prior / sales_prior if sales_prior else 0
    sgai = sga_ratio_current / sga_ratio_prior if sga_ratio_prior != 0 else 1

    # TATA: Total Accruals to Total Assets
    tata = (net_income - cfo) / total_assets_current

    # LVGI: Leverage Index
    lev_current = (ltd_current + current_liab_current) / total_assets_current
    lev_prior = (ltd_prior + current_liab_prior) / total_assets_prior if total_assets_prior else 0
    lvgi = lev_current / lev_prior if lev_prior != 0 else 1

    m_score = (-4.84 + 0.920 * dsri + 0.528 * gmi + 0.404 * aqi + 0.892 * sgi
               + 0.115 * depi - 0.172 * sgai + 4.679 * tata - 0.327 * lvgi)

    return json.dumps({
        "model_type": "beneish_m_score",
        "m_score": round(m_score, 4),
        "manipulation_likely": m_score > -1.78,
        "interpretation": "Likely manipulation" if m_score > -1.78 else "No manipulation signal",
        "components": {
            "dsri": round(dsri, 4),
            "gmi": round(gmi, 4),
            "aqi": round(aqi, 4),
            "sgi": round(sgi, 4),
            "depi": round(depi, 4),
            "sgai": round(sgai, 4),
            "tata": round(tata, 4),
            "lvgi": round(lvgi, 4),
        },
    })


def compute_accrual_ratios(
    db: Session,
    net_income: float,
    cfo: float,
    cfi: float,
    total_assets_current: float,
    total_assets_prior: float,
    cash_current: float = 0.0,
    cash_prior: float = 0.0,
    total_debt_current: float = 0.0,
    total_debt_prior: float = 0.0,
) -> str:
    """
    Compute accrual-based earnings quality metrics.

    CF-based: Accrual Ratio = (NI - CFO - CFI) / Average Total Assets
    BS-based: Accrual Ratio = (ΔNon-Cash Assets - ΔNon-Debt Liabilities) / Average Total Assets
    High accruals (positive) indicate lower earnings quality.
    """
    avg_assets = (total_assets_current + total_assets_prior) / 2
    if avg_assets == 0:
        return json.dumps({"error": "average total assets must be non-zero"})

    # CF-based accrual ratio
    cf_accrual_ratio = (net_income - cfo - cfi) / avg_assets

    # BS-based accrual ratio (simplified)
    delta_non_cash_assets = (total_assets_current - cash_current) - (total_assets_prior - cash_prior)
    delta_non_debt_liabilities = (
        (total_assets_current - total_debt_current - (total_assets_current - total_debt_current - 0))
    )
    # Simplified version: ΔNet Assets excluding cash and debt changes
    bs_accrual_ratio = (
        ((total_assets_current - cash_current) - (total_assets_prior - cash_prior))
        - ((total_debt_current) - (total_debt_prior))
    ) / avg_assets

    total_accruals = net_income - cfo

    return json.dumps({
        "model_type": "accrual_ratios",
        "total_accruals": round(total_accruals, 4),
        "cf_accrual_ratio": round(cf_accrual_ratio, 4),
        "bs_accrual_ratio": round(bs_accrual_ratio, 4),
        "interpretation": (
            "High accruals — lower earnings quality" if cf_accrual_ratio > 0.05
            else "Low accruals — higher earnings quality" if cf_accrual_ratio < -0.05
            else "Moderate accruals"
        ),
        "avg_total_assets": round(avg_assets, 4),
    })


def compute_variance_analysis(
    db: Session,
    actual_price: float,
    budget_price: float,
    actual_volume: float,
    budget_volume: float,
    actual_hours: Optional[float] = None,
    budget_hours: Optional[float] = None,
    actual_rate: Optional[float] = None,
    budget_rate: Optional[float] = None,
) -> str:
    """
    Compute budget vs. actual variance analysis for FP&A.

    Revenue variances:
        Price Variance = (Actual Price - Budget Price) × Actual Volume
        Volume Variance = (Actual Volume - Budget Volume) × Budget Price
        Total Variance = Price Variance + Volume Variance

    Labor variances (if hours and rates provided):
        Spending (Rate) Variance = (Actual Rate - Budget Rate) × Actual Hours
        Efficiency Variance = (Actual Hours - Budget Hours) × Budget Rate

    Favorable = positive (more revenue or less cost than budget).
    """
    actual_revenue = actual_price * actual_volume
    budget_revenue = budget_price * budget_volume

    price_variance = (actual_price - budget_price) * actual_volume
    volume_variance = (actual_volume - budget_volume) * budget_price
    total_revenue_variance = actual_revenue - budget_revenue

    result = {
        "model_type": "variance_analysis",
        "revenue": {
            "actual": round(actual_revenue, 4),
            "budget": round(budget_revenue, 4),
            "total_variance": round(total_revenue_variance, 4),
            "price_variance": round(price_variance, 4),
            "volume_variance": round(volume_variance, 4),
            "price_variance_favorable": price_variance >= 0,
            "volume_variance_favorable": volume_variance >= 0,
        },
    }

    if all(p is not None for p in [actual_hours, budget_hours, actual_rate, budget_rate]):
        spending_variance = (actual_rate - budget_rate) * actual_hours
        efficiency_variance = (actual_hours - budget_hours) * budget_rate
        total_labor_variance = spending_variance + efficiency_variance
        result["labor"] = {
            "actual_cost": round(actual_hours * actual_rate, 4),
            "budget_cost": round(budget_hours * budget_rate, 4),
            "total_variance": round(total_labor_variance, 4),
            "spending_variance": round(spending_variance, 4),
            "efficiency_variance": round(efficiency_variance, 4),
            "spending_variance_favorable": spending_variance <= 0,  # less cost = favorable
            "efficiency_variance_favorable": efficiency_variance <= 0,
        }

    return json.dumps(result)


# ── Industry Benchmarks Tool ──────────────────────────────────────────────────

def lookup_industry_benchmarks(
    db: Session,
    business_type: str,
    country: str = "Iran",
) -> str:
    """
    Look up industry financial benchmarks for a given business type using web search.

    Performs 3 targeted web searches to find:
    - Revenue range (small/medium/large businesses)
    - Gross margin and EBITDA margin benchmarks
    - Key cost drivers and working capital norms

    The LLM should extract structured benchmark data from the search results
    and use them as default inputs when building financial models.

    Args:
        business_type: Business description in English or Persian
            (e.g. "coffee roastery", "online trading exchange", "pharmacy", "قهوه‌برشته‌کاری")
        country: Country context for localized benchmarks. Default "Iran".

    Returns:
        JSON with search results for the LLM to extract benchmark data from.
    """
    from rag.tools.web import web_search

    queries = [
        f"{business_type} annual revenue benchmark {country} 2024 2025",
        f"{business_type} gross margin EBITDA margin industry average",
        f"{business_type} درآمد سالانه بنچمارک ایران حاشیه سود",
    ]

    all_results = []
    for query in queries:
        result_json = web_search(db, query=query, max_results=3)
        try:
            data = json.loads(result_json)
            if "results" in data:
                all_results.extend(data["results"])
        except Exception:
            pass

    return json.dumps({
        "model_type": "industry_benchmarks",
        "business_type": business_type,
        "country": country,
        "search_results": all_results[:9],
        "instruction": (
            "Extract from these results: revenue range (min/max), "
            "gross margin %, EBITDA margin %, typical CapEx % of revenue, "
            "DSO/DIO/DPO norms. Use as DEFAULT inputs for modeling tools. "
            "Always cite the source when presenting benchmarks to the user."
        ),
    }, ensure_ascii=False)


# ── Beta Estimation Tool ──────────────────────────────────────────────────────

def compute_beta(
    db: Session,
    levered_beta: float,
    debt_to_equity: float,
    tax_rate: float,
    target_debt_to_equity: Optional[float] = None,
) -> str:
    """
    Compute unlevered and re-levered beta using the Hamada equation.

    Formulas:
        β_U   = β_L / [1 + (1−T) × (D/E)]
        β_L   = β_U × [1 + (1−T) × (D/E)_target]
        β_adj = (2/3) × β_L + (1/3) × 1.0    (Bloomberg adjusted)
    """
    if target_debt_to_equity is None:
        target_debt_to_equity = debt_to_equity

    hamada_factor = 1 + (1 - tax_rate) * debt_to_equity
    unlevered_beta = levered_beta / hamada_factor
    target_hamada_factor = 1 + (1 - tax_rate) * target_debt_to_equity
    re_levered_beta = unlevered_beta * target_hamada_factor
    adjusted_beta = (2 / 3) * levered_beta + (1 / 3) * 1.0

    return json.dumps({
        "model_type": "beta",
        "levered_beta": round(levered_beta, 4),
        "unlevered_beta": round(unlevered_beta, 6),
        "re_levered_beta": round(re_levered_beta, 6),
        "adjusted_beta": round(adjusted_beta, 6),
        "debt_to_equity": debt_to_equity,
        "target_debt_to_equity": target_debt_to_equity,
        "tax_rate": tax_rate,
        "hamada_formula": f"β_U = {levered_beta} / [1 + (1−{tax_rate}) × {debt_to_equity}] = {round(unlevered_beta, 4)}",
        "bloomberg_formula": f"β_adj = (2/3) × {levered_beta} + 1/3 = {round(adjusted_beta, 4)}",
    })


# ── Scenario Analysis Tool ────────────────────────────────────────────────────

def build_scenario_model(
    db: Session,
    base_results: dict,
    bear_pct: float,
    bull_pct: float,
    bear_overrides: Optional[dict] = None,
    bull_overrides: Optional[dict] = None,
    scenario_labels: Optional[dict] = None,
) -> str:
    """
    Apply bear/base/bull scenarios to an already-computed result dict.

    For each numeric metric in base_results:
        bear_value = base × (1 + bear_overrides.get(key, bear_pct))
        bull_value = base × (1 + bull_overrides.get(key, bull_pct))
    """
    if not base_results:
        return json.dumps({"error": "base_results must not be empty"})

    bear_ov = bear_overrides or {}
    bull_ov = bull_overrides or {}
    labels = scenario_labels or {"bear": "Bear", "base": "Base", "bull": "Bull"}

    bear_scenario, bull_scenario = {}, {}
    downside_pct, upside_pct = {}, {}

    for key, base_val in base_results.items():
        if not isinstance(base_val, (int, float)):
            continue
        b_pct = bear_ov.get(key, bear_pct)
        u_pct = bull_ov.get(key, bull_pct)
        bear_val = base_val * (1 + b_pct)
        bull_val = base_val * (1 + u_pct)
        bear_scenario[key] = round(bear_val, 4)
        bull_scenario[key] = round(bull_val, 4)
        downside_pct[key] = round((bear_val / base_val - 1) * 100, 2) if base_val else 0
        upside_pct[key] = round((bull_val / base_val - 1) * 100, 2) if base_val else 0

    return json.dumps({
        "model_type": "scenario_model",
        "scenarios": {
            labels["bear"]: bear_scenario,
            labels["base"]: {k: round(v, 4) for k, v in base_results.items() if isinstance(v, (int, float))},
            labels["bull"]: bull_scenario,
        },
        "summary": {
            "downside_pct": downside_pct,
            "upside_pct": upside_pct,
            "bear_pct_global": round(bear_pct * 100, 2),
            "bull_pct_global": round(bull_pct * 100, 2),
        },
    })


# ── Operating Leverage Tool ───────────────────────────────────────────────────

def compute_operating_leverage(
    db: Session,
    revenue: float,
    variable_costs: float,
    fixed_costs: float,
    units_sold: Optional[float] = None,
) -> str:
    """
    Compute operating leverage: DOL, contribution margin, and breakeven.

    Formulas:
        CM            = Revenue − Variable Costs
        CM Ratio      = CM / Revenue
        EBIT          = CM − Fixed Costs
        DOL           = CM / EBIT
        Breakeven Rev = Fixed Costs / CM Ratio
    """
    if revenue == 0:
        return json.dumps({"error": "revenue must be positive"})
    cm = revenue - variable_costs
    cm_ratio = cm / revenue
    if cm_ratio == 0:
        return json.dumps({"error": "Contribution margin is zero — cannot compute DOL or breakeven"})

    ebit = cm - fixed_costs
    dol = (cm / ebit) if ebit != 0 else None
    breakeven_revenue = fixed_costs / cm_ratio

    breakeven_units = None
    if units_sold and units_sold > 0:
        price_per_unit = revenue / units_sold
        vc_per_unit = variable_costs / units_sold
        cm_per_unit = price_per_unit - vc_per_unit
        if cm_per_unit > 0:
            breakeven_units = round(fixed_costs / cm_per_unit, 4)

    return json.dumps({
        "model_type": "operating_leverage",
        "revenue": revenue,
        "variable_costs": variable_costs,
        "fixed_costs": fixed_costs,
        "contribution_margin": round(cm, 4),
        "cm_ratio": round(cm_ratio, 6),
        "ebit": round(ebit, 4),
        "dol": round(dol, 4) if dol is not None else None,
        "breakeven_revenue": round(breakeven_revenue, 4),
        "breakeven_units": breakeven_units,
    })


# ── PVGO Tool ─────────────────────────────────────────────────────────────────

def compute_pvgo(
    db: Session,
    intrinsic_value: float,
    earnings_per_share: float,
    cost_of_equity: float,
    growth_rate: float,
    payout_ratio: float,
) -> str:
    """
    Compute PVGO and justified P/E ratios.

    Formulas:
        No-Growth Value  = E₁ / ke
        PVGO             = Intrinsic Value − No-Growth Value
        Justified P/E (leading)  = (1 − b) / (ke − g)
        Justified P/E (trailing) = (1 − b) × (1 + g) / (ke − g)
    """
    if cost_of_equity <= growth_rate:
        return json.dumps({"error": "cost_of_equity must be greater than growth_rate (ke > g)"})
    if cost_of_equity <= 0:
        return json.dumps({"error": "cost_of_equity must be positive"})

    no_growth_value = earnings_per_share / cost_of_equity
    pvgo = intrinsic_value - no_growth_value
    pvgo_pct = (pvgo / intrinsic_value * 100) if intrinsic_value else 0
    ke_minus_g = cost_of_equity - growth_rate
    justified_pe_leading = (1 - payout_ratio) / ke_minus_g
    justified_pe_trailing = justified_pe_leading * (1 + growth_rate)

    return json.dumps({
        "model_type": "pvgo",
        "intrinsic_value": round(intrinsic_value, 4),
        "no_growth_value": round(no_growth_value, 4),
        "pvgo": round(pvgo, 4),
        "pvgo_pct_of_value": round(pvgo_pct, 2),
        "earnings_per_share": earnings_per_share,
        "cost_of_equity_pct": round(cost_of_equity * 100, 2),
        "growth_rate_pct": round(growth_rate * 100, 2),
        "payout_ratio": payout_ratio,
        "justified_pe_leading": round(justified_pe_leading, 4),
        "justified_pe_trailing": round(justified_pe_trailing, 4),
    })


# ── EVA Tool ──────────────────────────────────────────────────────────────────

def compute_eva(
    db: Session,
    ebit: float,
    tax_rate: float,
    wacc: float,
    invested_capital: float,
    market_value_of_firm: Optional[float] = None,
) -> str:
    """
    Compute Economic Value Added (EVA) and Market Value Added (MVA).

    Formulas:
        NOPAT      = EBIT × (1 − T)
        ROIC       = NOPAT / Invested Capital
        EVA        = (ROIC − WACC) × IC
        EVA Spread = ROIC − WACC
        MVA        = Market Value − Book Value of IC   [optional]
    """
    if invested_capital <= 0:
        return json.dumps({"error": "invested_capital must be positive"})

    nopat = ebit * (1 - tax_rate)
    roic = nopat / invested_capital
    capital_charge = wacc * invested_capital
    eva = nopat - capital_charge
    eva_spread = roic - wacc
    mva = (market_value_of_firm - invested_capital) if market_value_of_firm is not None else None

    return json.dumps({
        "model_type": "eva",
        "ebit": ebit,
        "tax_rate": tax_rate,
        "nopat": round(nopat, 4),
        "invested_capital": invested_capital,
        "wacc_pct": round(wacc * 100, 2),
        "capital_charge": round(capital_charge, 4),
        "roic": round(roic, 6),
        "roic_pct": round(roic * 100, 4),
        "eva": round(eva, 4),
        "eva_spread": round(eva_spread, 6),
        "eva_spread_pct": round(eva_spread * 100, 4),
        "mva": round(mva, 4) if mva is not None else None,
        "value_creation": "positive" if eva > 0 else "negative" if eva < 0 else "neutral",
    })


# ── Phase 7: Portfolio & Risk Analytics ───────────────────────────────────────

def compute_portfolio_stats(
    db: Session,
    assets: list,
    correlation_matrix: list,
) -> str:
    """
    Compute portfolio return, volatility, and diversification ratio.

    E(Rp) = Σ wᵢ×E(Rᵢ)
    σp = √(w'Σw)
    Diversification ratio = (Σ wᵢσᵢ) / σp
    """
    n = len(assets)
    if n == 0:
        return json.dumps({"error": "assets list must not be empty"})
    if len(correlation_matrix) != n:
        return json.dumps({"error": f"correlation_matrix must be {n}x{n}, got {len(correlation_matrix)} rows"})
    for row in correlation_matrix:
        if len(row) != n:
            return json.dumps({"error": f"correlation_matrix must be {n}x{n}"})

    weights = [a["weight"] for a in assets]
    expected_returns = [a["expected_return"] for a in assets]
    volatilities = [a["volatility"] for a in assets]

    weight_sum = sum(weights)
    if abs(weight_sum - 1.0) > 0.01:
        return json.dumps({"error": f"weights must sum to ~1.0, got {round(weight_sum, 6)}"})

    # Portfolio return
    port_return = sum(w * r for w, r in zip(weights, expected_returns))

    # Covariance matrix from correlation and volatilities
    # Σ_ij = ρ_ij × σ_i × σ_j
    # Portfolio variance = w' Σ w
    port_variance = 0.0
    for i in range(n):
        for j in range(n):
            port_variance += (
                weights[i] * weights[j]
                * volatilities[i] * volatilities[j]
                * correlation_matrix[i][j]
            )
    port_volatility = math.sqrt(max(port_variance, 0.0))

    # Diversification ratio
    weighted_vol_sum = sum(w * s for w, s in zip(weights, volatilities))
    div_ratio = weighted_vol_sum / port_volatility if port_volatility > 0 else 1.0

    return json.dumps({
        "model_type": "portfolio_stats",
        "portfolio_return": round(port_return, 4),
        "portfolio_volatility": round(port_volatility, 4),
        "diversification_ratio": round(div_ratio, 4),
        "assets": [
            {
                "name": a["name"],
                "weight": round(a["weight"], 4),
                "expected_return": round(a["expected_return"], 4),
                "volatility": round(a["volatility"], 4),
            }
            for a in assets
        ],
    })


def compute_risk_metrics(
    db: Session,
    returns: list,
    risk_free_rate: float = 0.0,
    benchmark_returns: Optional[list] = None,
    periods_per_year: int = 12,
) -> str:
    """
    Compute risk-adjusted performance metrics from a return series.

    Metrics: annualized return, annualized volatility, Sharpe ratio,
    Sortino ratio, max drawdown, Calmar ratio.
    If benchmark provided: beta, Treynor ratio, information ratio, tracking error.
    """
    if len(returns) < 2:
        return json.dumps({"error": "need at least 2 return observations"})

    mean_r = statistics.mean(returns)
    std_r = statistics.pstdev(returns)  # population stdev for consistency

    ann_return = mean_r * periods_per_year
    ann_vol = std_r * math.sqrt(periods_per_year)

    # Sharpe
    sharpe = (ann_return - risk_free_rate) / ann_vol if ann_vol > 0 else 0.0

    # Sortino — downside deviation
    downside = [r for r in returns if r < 0]
    if len(downside) > 0:
        downside_var = sum(r ** 2 for r in downside) / len(returns)
        downside_std = math.sqrt(downside_var) * math.sqrt(periods_per_year)
    else:
        downside_std = 0.0
    sortino = (ann_return - risk_free_rate) / downside_std if downside_std > 0 else 0.0

    # Max drawdown
    cumulative = 1.0
    peak = 1.0
    max_dd = 0.0
    for r in returns:
        cumulative *= (1 + r)
        if cumulative > peak:
            peak = cumulative
        dd = (peak - cumulative) / peak
        if dd > max_dd:
            max_dd = dd

    # Calmar
    calmar = ann_return / max_dd if max_dd > 0 else 0.0

    result = {
        "model_type": "risk_metrics",
        "annualized_return": round(ann_return, 4),
        "annualized_volatility": round(ann_vol, 4),
        "sharpe_ratio": round(sharpe, 4),
        "sortino_ratio": round(sortino, 4),
        "max_drawdown": round(max_dd, 4),
        "calmar_ratio": round(calmar, 4),
    }

    # Benchmark-relative metrics
    if benchmark_returns is not None:
        if len(benchmark_returns) != len(returns):
            return json.dumps({"error": "benchmark_returns must have same length as returns"})
        excess = [r - b for r, b in zip(returns, benchmark_returns)]
        mean_excess = statistics.mean(excess)
        tracking_error = statistics.pstdev(excess) * math.sqrt(periods_per_year)
        info_ratio = (mean_excess * periods_per_year) / tracking_error if tracking_error > 0 else 0.0

        # Beta = Cov(Rp, Rb) / Var(Rb)
        mean_b = statistics.mean(benchmark_returns)
        cov_pb = sum((r - mean_r) * (b - mean_b) for r, b in zip(returns, benchmark_returns)) / len(returns)
        var_b = sum((b - mean_b) ** 2 for b in benchmark_returns) / len(benchmark_returns)
        beta = cov_pb / var_b if var_b > 0 else 0.0

        # Treynor
        treynor = (ann_return - risk_free_rate) / beta if beta != 0 else 0.0

        result["beta"] = round(beta, 4)
        result["treynor_ratio"] = round(treynor, 4)
        result["information_ratio"] = round(info_ratio, 4)
        result["tracking_error"] = round(tracking_error, 4)

    return json.dumps(result)


def compute_var(
    db: Session,
    portfolio_value: float,
    confidence_level: float = 0.95,
    method: str = "parametric",
    horizon_days: int = 1,
    expected_return: Optional[float] = None,
    volatility: Optional[float] = None,
    returns: Optional[list] = None,
    num_simulations: int = 10000,
    seed: Optional[int] = None,
) -> str:
    """
    Compute Value at Risk (VaR).

    Methods:
      - parametric: VaR = |−μ_daily×t + z_α × σ_daily × √t| × portfolio_value
      - historical: percentile of sorted returns
      - monte_carlo: simulate N paths, take percentile
    """
    z_scores = {0.90: 1.2816, 0.95: 1.6449, 0.99: 2.3263, 0.975: 1.9600}

    if method == "parametric":
        if expected_return is None or volatility is None:
            return json.dumps({"error": "parametric method requires expected_return and volatility"})
        z = z_scores.get(confidence_level)
        if z is None:
            return json.dumps({"error": f"unsupported confidence_level {confidence_level}, use one of {list(z_scores.keys())}"})
        mu_daily = expected_return / 252
        sigma_daily = volatility / math.sqrt(252)
        t = horizon_days
        var_pct = -mu_daily * t + z * sigma_daily * math.sqrt(t)
        var_pct = abs(var_pct)
        var_amount = var_pct * portfolio_value
        return json.dumps({
            "model_type": "var",
            "method": "parametric",
            "confidence_level": confidence_level,
            "horizon_days": horizon_days,
            "var_pct": round(var_pct, 4),
            "var_amount": round(var_amount, 4),
            "portfolio_value": round(portfolio_value, 4),
        })

    elif method == "historical":
        if returns is None or len(returns) < 2:
            return json.dumps({"error": "historical method requires returns list with at least 2 observations"})
        sorted_returns = sorted(returns)
        idx = int((1 - confidence_level) * len(sorted_returns))
        idx = max(0, min(idx, len(sorted_returns) - 1))
        var_return = sorted_returns[idx]
        if horizon_days > 1:
            var_return = var_return * math.sqrt(horizon_days)
        var_pct = abs(var_return)
        var_amount = var_pct * portfolio_value
        return json.dumps({
            "model_type": "var",
            "method": "historical",
            "confidence_level": confidence_level,
            "horizon_days": horizon_days,
            "var_pct": round(var_pct, 4),
            "var_amount": round(var_amount, 4),
            "portfolio_value": round(portfolio_value, 4),
            "num_observations": len(returns),
        })

    elif method == "monte_carlo":
        if expected_return is None or volatility is None:
            return json.dumps({"error": "monte_carlo method requires expected_return and volatility"})
        rng = random.Random(seed)
        mu_daily = expected_return / 252
        sigma_daily = volatility / math.sqrt(252)
        terminal_returns = []
        for _ in range(num_simulations):
            cum_value = 1.0
            for _d in range(horizon_days):
                cum_value *= math.exp(
                    (mu_daily - 0.5 * sigma_daily ** 2)
                    + sigma_daily * rng.gauss(0, 1)
                )
            terminal_returns.append(cum_value - 1.0)
        terminal_returns.sort()
        idx = int((1 - confidence_level) * num_simulations)
        idx = max(0, min(idx, num_simulations - 1))
        var_return = terminal_returns[idx]
        var_pct = abs(var_return)
        var_amount = var_pct * portfolio_value
        return json.dumps({
            "model_type": "var",
            "method": "monte_carlo",
            "confidence_level": confidence_level,
            "horizon_days": horizon_days,
            "var_pct": round(var_pct, 4),
            "var_amount": round(var_amount, 4),
            "portfolio_value": round(portfolio_value, 4),
            "num_simulations": num_simulations,
        })

    else:
        return json.dumps({"error": f"unknown method '{method}', use parametric, historical, or monte_carlo"})


def compute_cvar(
    db: Session,
    portfolio_value: float,
    confidence_level: float = 0.95,
    horizon_days: int = 1,
    expected_return: Optional[float] = None,
    volatility: Optional[float] = None,
    returns: Optional[list] = None,
) -> str:
    """
    Compute Conditional Value at Risk (CVaR / Expected Shortfall).

    Auto-selects method: historical if returns provided, else parametric.
    CVaR = E[Loss | Loss > VaR]
    """
    z_scores = {0.90: 1.2816, 0.95: 1.6449, 0.99: 2.3263, 0.975: 1.9600}

    if returns is not None and len(returns) >= 2:
        # Historical CVaR
        sorted_returns = sorted(returns)
        cutoff_idx = int((1 - confidence_level) * len(sorted_returns))
        cutoff_idx = max(1, cutoff_idx)  # at least 1 observation
        tail = sorted_returns[:cutoff_idx]
        var_return = sorted_returns[max(0, cutoff_idx - 1)]
        cvar_return = statistics.mean(tail) if tail else var_return
        if horizon_days > 1:
            var_return = var_return * math.sqrt(horizon_days)
            cvar_return = cvar_return * math.sqrt(horizon_days)
        var_pct = abs(var_return)
        cvar_pct = abs(cvar_return)
        var_amount = var_pct * portfolio_value
        cvar_amount = cvar_pct * portfolio_value
        return json.dumps({
            "model_type": "cvar",
            "method": "historical",
            "confidence_level": confidence_level,
            "horizon_days": horizon_days,
            "var_pct": round(var_pct, 4),
            "var_amount": round(var_amount, 4),
            "cvar_pct": round(cvar_pct, 4),
            "cvar_amount": round(cvar_amount, 4),
            "portfolio_value": round(portfolio_value, 4),
            "num_observations": len(returns),
        })
    elif expected_return is not None and volatility is not None:
        # Parametric CVaR (normal distribution)
        z = z_scores.get(confidence_level)
        if z is None:
            return json.dumps({"error": f"unsupported confidence_level {confidence_level}, use one of {list(z_scores.keys())}"})
        mu_daily = expected_return / 252
        sigma_daily = volatility / math.sqrt(252)
        t = horizon_days
        mu_t = mu_daily * t
        sigma_t = sigma_daily * math.sqrt(t)

        var_pct = abs(-mu_t + z * sigma_t)

        # CVaR for normal: σ × φ(z)/(1-α) - μ  (scaled to horizon)
        # φ(z) = pdf of standard normal at z
        phi_z = (1.0 / math.sqrt(2 * math.pi)) * math.exp(-0.5 * z ** 2)
        alpha = 1 - confidence_level
        cvar_pct = abs(-mu_t + sigma_t * phi_z / alpha)

        var_amount = var_pct * portfolio_value
        cvar_amount = cvar_pct * portfolio_value
        return json.dumps({
            "model_type": "cvar",
            "method": "parametric",
            "confidence_level": confidence_level,
            "horizon_days": horizon_days,
            "var_pct": round(var_pct, 4),
            "var_amount": round(var_amount, 4),
            "cvar_pct": round(cvar_pct, 4),
            "cvar_amount": round(cvar_amount, 4),
            "portfolio_value": round(portfolio_value, 4),
        })
    else:
        return json.dumps({"error": "provide either returns (for historical) or expected_return + volatility (for parametric)"})


def run_monte_carlo(
    db: Session,
    initial_value: float,
    expected_return: float,
    volatility: float,
    horizon_years: float = 1.0,
    num_simulations: int = 10000,
    num_steps: int = 252,
    seed: Optional[int] = None,
) -> str:
    """
    Run Monte Carlo simulation using Geometric Brownian Motion (GBM).

    S(t+dt) = S(t) × exp((μ − σ²/2)dt + σ√dt × Z)
    """
    if initial_value <= 0:
        return json.dumps({"error": "initial_value must be positive"})
    if num_steps < 1:
        return json.dumps({"error": "num_steps must be >= 1"})

    rng = random.Random(seed)
    dt = horizon_years / num_steps
    drift = (expected_return - 0.5 * volatility ** 2) * dt
    diffusion = volatility * math.sqrt(dt)

    terminal_values = []
    for _ in range(num_simulations):
        s = initial_value
        for _step in range(num_steps):
            z = rng.gauss(0, 1)
            s *= math.exp(drift + diffusion * z)
        terminal_values.append(s)

    terminal_values.sort()
    n = len(terminal_values)
    prob_loss = sum(1 for v in terminal_values if v < initial_value) / n

    mean_val = statistics.mean(terminal_values)
    median_val = statistics.median(terminal_values)
    std_val = statistics.pstdev(terminal_values)

    def _percentile(data, pct):
        k = (len(data) - 1) * pct / 100.0
        f = int(k)
        c = f + 1
        if c >= len(data):
            return data[-1]
        return data[f] + (k - f) * (data[c] - data[f])

    return json.dumps({
        "model_type": "monte_carlo",
        "initial_value": round(initial_value, 4),
        "expected_return": round(expected_return, 4),
        "volatility": round(volatility, 4),
        "horizon_years": round(horizon_years, 4),
        "num_simulations": num_simulations,
        "num_steps": num_steps,
        "terminal_stats": {
            "mean": round(mean_val, 4),
            "median": round(median_val, 4),
            "std": round(std_val, 4),
            "min": round(terminal_values[0], 4),
            "max": round(terminal_values[-1], 4),
        },
        "percentile_paths": {
            "p5": round(_percentile(terminal_values, 5), 4),
            "p25": round(_percentile(terminal_values, 25), 4),
            "p50": round(_percentile(terminal_values, 50), 4),
            "p75": round(_percentile(terminal_values, 75), 4),
            "p95": round(_percentile(terminal_values, 95), 4),
        },
        "probability_of_loss": round(prob_loss, 4),
    })


def optimize_portfolio(
    db: Session,
    assets: list,
    correlation_matrix: list,
    risk_free_rate: float = 0.0,
    objective: str = "max_sharpe",
    target_return: Optional[float] = None,
    min_weight: float = 0.0,
    max_weight: float = 1.0,
    num_portfolios: int = 10000,
    seed: Optional[int] = None,
) -> str:
    """
    Portfolio optimization via Monte Carlo random weight generation.

    Objectives: min_variance, max_sharpe, target_return.
    """
    n = len(assets)
    if n == 0:
        return json.dumps({"error": "assets list must not be empty"})
    if len(correlation_matrix) != n:
        return json.dumps({"error": f"correlation_matrix must be {n}x{n}"})
    if objective not in ("min_variance", "max_sharpe", "target_return"):
        return json.dumps({"error": f"unknown objective '{objective}'"})
    if objective == "target_return" and target_return is None:
        return json.dumps({"error": "target_return required when objective is 'target_return'"})

    expected_returns = [a["expected_return"] for a in assets]
    volatilities = [a["volatility"] for a in assets]
    rng = random.Random(seed)

    def _port_metrics(w):
        ret = sum(wi * ri for wi, ri in zip(w, expected_returns))
        var = 0.0
        for i in range(n):
            for j in range(n):
                var += w[i] * w[j] * volatilities[i] * volatilities[j] * correlation_matrix[i][j]
        vol = math.sqrt(max(var, 0.0))
        sharpe = (ret - risk_free_rate) / vol if vol > 0 else 0.0
        return ret, vol, sharpe

    best_weights = None
    best_ret = 0.0
    best_vol = float("inf")
    best_sharpe = float("-inf")

    for _ in range(num_portfolios):
        # Generate random weights within bounds
        raw = [rng.uniform(min_weight, max_weight) for _ in range(n)]
        total = sum(raw)
        if total == 0:
            continue
        w = [x / total for x in raw]
        # Check bounds after normalization
        if any(wi < min_weight - 1e-9 or wi > max_weight + 1e-9 for wi in w):
            continue

        ret, vol, sharpe = _port_metrics(w)

        if objective == "min_variance":
            if vol < best_vol:
                best_weights, best_ret, best_vol, best_sharpe = w, ret, vol, sharpe
        elif objective == "max_sharpe":
            if sharpe > best_sharpe:
                best_weights, best_ret, best_vol, best_sharpe = w, ret, vol, sharpe
        elif objective == "target_return":
            # Find portfolio closest to target return with minimum variance
            if abs(ret - target_return) < 0.01:
                if vol < best_vol:
                    best_weights, best_ret, best_vol, best_sharpe = w, ret, vol, sharpe

    if best_weights is None:
        return json.dumps({"error": "could not find feasible portfolio"})

    return json.dumps({
        "model_type": "portfolio_optimization",
        "objective": objective,
        "optimal_weights": [
            {"name": assets[i]["name"], "weight": round(best_weights[i], 4)}
            for i in range(n)
        ],
        "portfolio_return": round(best_ret, 4),
        "portfolio_volatility": round(best_vol, 4),
        "sharpe_ratio": round(best_sharpe, 4),
    })


def compute_efficient_frontier(
    db: Session,
    assets: list,
    correlation_matrix: list,
    risk_free_rate: float = 0.0,
    num_points: int = 50,
) -> str:
    """
    Generate efficient frontier by finding minimum-variance portfolios
    at different target return levels.
    """
    n = len(assets)
    if n == 0:
        return json.dumps({"error": "assets list must not be empty"})
    if len(correlation_matrix) != n:
        return json.dumps({"error": f"correlation_matrix must be {n}x{n}"})

    expected_returns = [a["expected_return"] for a in assets]
    volatilities = [a["volatility"] for a in assets]

    min_ret = min(expected_returns)
    max_ret = max(expected_returns)

    def _port_metrics(w):
        ret = sum(wi * ri for wi, ri in zip(w, expected_returns))
        var = 0.0
        for i in range(n):
            for j in range(n):
                var += w[i] * w[j] * volatilities[i] * volatilities[j] * correlation_matrix[i][j]
        vol = math.sqrt(max(var, 0.0))
        sharpe = (ret - risk_free_rate) / vol if vol > 0 else 0.0
        return ret, vol, sharpe

    rng = random.Random(42)  # fixed seed for reproducibility
    num_mc = 5000  # Monte Carlo per target

    if max_ret - min_ret < 1e-8:
        return json.dumps({"error": "all assets have the same expected return; frontier is a single point"})

    frontier_points = []
    step = (max_ret - min_ret) / max(num_points - 1, 1)
    target_returns = [min_ret + i * step for i in range(num_points)]

    global_best_sharpe = float("-inf")
    tangent_portfolio = None
    min_var_vol = float("inf")
    min_var_portfolio = None

    for target in target_returns:
        best_vol = float("inf")
        best_weights = None
        for _ in range(num_mc):
            raw = [rng.random() for _ in range(n)]
            total = sum(raw)
            w = [x / total for x in raw]
            ret, vol, sharpe = _port_metrics(w)
            if abs(ret - target) < step * 0.5 and vol < best_vol:
                best_vol = vol
                best_weights = w

        if best_weights is not None:
            ret, vol, sharpe = _port_metrics(best_weights)
            frontier_points.append({
                "return": round(ret, 4),
                "volatility": round(vol, 4),
            })
            if vol < min_var_vol:
                min_var_vol = vol
                min_var_portfolio = {
                    "weights": [{"name": assets[i]["name"], "weight": round(best_weights[i], 4)} for i in range(n)],
                    "return": round(ret, 4),
                    "volatility": round(vol, 4),
                }
            if sharpe > global_best_sharpe:
                global_best_sharpe = sharpe
                tangent_portfolio = {
                    "weights": [{"name": assets[i]["name"], "weight": round(best_weights[i], 4)} for i in range(n)],
                    "return": round(ret, 4),
                    "volatility": round(vol, 4),
                    "sharpe_ratio": round(sharpe, 4),
                }

    return json.dumps({
        "model_type": "efficient_frontier",
        "num_points": len(frontier_points),
        "frontier_points": frontier_points,
        "min_variance_portfolio": min_var_portfolio,
        "tangent_portfolio": tangent_portfolio,
    })


def compute_risk_parity(
    db: Session,
    assets: list,
    correlation_matrix: list,
) -> str:
    """
    Compute risk parity portfolio weights.

    Method: iterative algorithm starting from inverse-volatility weights,
    adjusting to equalize risk contributions.
    Risk contribution_i = w_i × (Σw)_i / σp
    """
    n = len(assets)
    if n == 0:
        return json.dumps({"error": "assets list must not be empty"})
    if len(correlation_matrix) != n:
        return json.dumps({"error": f"correlation_matrix must be {n}x{n}"})

    volatilities = [a["volatility"] for a in assets]

    # Check for zero volatilities
    if any(v <= 0 for v in volatilities):
        return json.dumps({"error": "all asset volatilities must be positive"})

    # Start with inverse-volatility weights
    inv_vol = [1.0 / v for v in volatilities]
    total_inv = sum(inv_vol)
    weights = [iv / total_inv for iv in inv_vol]

    def _risk_contributions(w):
        """Compute marginal risk contributions."""
        # Portfolio variance components: (Σw)_i
        marginal = []
        for i in range(n):
            mc_i = 0.0
            for j in range(n):
                mc_i += w[j] * volatilities[i] * volatilities[j] * correlation_matrix[i][j]
            marginal.append(mc_i)
        port_var = sum(w[i] * marginal[i] for i in range(n))
        port_vol = math.sqrt(max(port_var, 0.0))
        if port_vol == 0:
            return [0.0] * n, 0.0
        rc = [w[i] * marginal[i] / port_vol for i in range(n)]
        return rc, port_vol

    # Iterative adjustment (simple gradient-like)
    for _iteration in range(200):
        rc, port_vol = _risk_contributions(weights)
        if port_vol == 0:
            break
        target_rc = port_vol / n  # equal risk per asset
        # Adjust weights proportionally to deviation from target
        new_weights = []
        for i in range(n):
            if rc[i] > 0:
                adj = weights[i] * (target_rc / rc[i]) ** 0.5
            else:
                adj = weights[i]
            new_weights.append(adj)
        total_w = sum(new_weights)
        weights = [w / total_w for w in new_weights]

    rc_final, port_vol_final = _risk_contributions(weights)

    return json.dumps({
        "model_type": "risk_parity",
        "weights": [
            {"name": assets[i]["name"], "weight": round(weights[i], 4)}
            for i in range(n)
        ],
        "risk_contributions": [
            {"name": assets[i]["name"], "risk_contribution": round(rc_final[i], 4)}
            for i in range(n)
        ],
        "total_volatility": round(port_vol_final, 4),
    })


def compute_factor_model(
    db: Session,
    asset_returns: list,
    market_returns: list,
    risk_free_rate: float = 0.0,
    smb_returns: Optional[list] = None,
    hml_returns: Optional[list] = None,
) -> str:
    """
    Compute single-factor (CAPM) or Fama-French three-factor regression.

    Single: Rᵢ−Rf = α + β(Rm−Rf) + ε
    Three: Rᵢ−Rf = α + β_m(Rm−Rf) + β_smb×SMB + β_hml×HML + ε
    Uses OLS via normal equations.
    """
    n = len(asset_returns)
    if n != len(market_returns):
        return json.dumps({"error": "asset_returns and market_returns must have same length"})
    if n < 3:
        return json.dumps({"error": "need at least 3 observations"})

    # Excess returns
    y = [asset_returns[i] - risk_free_rate for i in range(n)]
    x_market = [market_returns[i] - risk_free_rate for i in range(n)]

    three_factor = smb_returns is not None and hml_returns is not None
    if three_factor:
        if len(smb_returns) != n or len(hml_returns) != n:
            return json.dumps({"error": "smb_returns and hml_returns must have same length as asset_returns"})

    if not three_factor:
        # Single factor OLS: y = alpha + beta * x_market
        # Normal equations: [Σ1, Σx; Σx, Σx²] [α; β] = [Σy; Σxy]
        sum_x = sum(x_market)
        sum_y = sum(y)
        sum_xy = sum(xi * yi for xi, yi in zip(x_market, y))
        sum_xx = sum(xi * xi for xi in x_market)

        det = n * sum_xx - sum_x * sum_x
        if abs(det) < 1e-15:
            return json.dumps({"error": "singular matrix in regression"})

        alpha = (sum_y * sum_xx - sum_x * sum_xy) / det
        beta = (n * sum_xy - sum_x * sum_y) / det

        # R-squared
        y_pred = [alpha + beta * x_market[i] for i in range(n)]
        ss_res = sum((y[i] - y_pred[i]) ** 2 for i in range(n))
        mean_y = sum(y) / n
        ss_tot = sum((y[i] - mean_y) ** 2 for i in range(n))
        r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
        residual_std = math.sqrt(ss_res / max(n - 2, 1))

        return json.dumps({
            "model_type": "factor_model",
            "factors": "single",
            "alpha": round(alpha, 4),
            "beta_market": round(beta, 4),
            "r_squared": round(r_squared, 4),
            "residual_std": round(residual_std, 4),
            "num_observations": n,
        })
    else:
        # Three-factor OLS via normal equations (4x4 system)
        # y = b0 + b1*x1 + b2*x2 + b3*x3
        # where x1=market, x2=smb, x3=hml
        X = []
        for i in range(n):
            X.append([1.0, x_market[i], smb_returns[i], hml_returns[i]])

        # X'X matrix (4x4) and X'y vector (4x1)
        k = 4
        xtx = [[0.0] * k for _ in range(k)]
        xty = [0.0] * k
        for i in range(n):
            for r in range(k):
                xty[r] += X[i][r] * y[i]
                for c in range(k):
                    xtx[r][c] += X[i][r] * X[i][c]

        # Gaussian elimination with partial pivoting
        aug = [xtx[r][:] + [xty[r]] for r in range(k)]
        for col in range(k):
            # Pivot
            max_row = col
            for row in range(col + 1, k):
                if abs(aug[row][col]) > abs(aug[max_row][col]):
                    max_row = row
            aug[col], aug[max_row] = aug[max_row], aug[col]

            if abs(aug[col][col]) < 1e-15:
                return json.dumps({"error": "singular matrix in 3-factor regression"})

            for row in range(col + 1, k):
                factor = aug[row][col] / aug[col][col]
                for j in range(col, k + 1):
                    aug[row][j] -= factor * aug[col][j]

        # Back substitution
        coeffs = [0.0] * k
        for row in range(k - 1, -1, -1):
            coeffs[row] = aug[row][k]
            for col in range(row + 1, k):
                coeffs[row] -= aug[row][col] * coeffs[col]
            coeffs[row] /= aug[row][row]

        alpha, beta_m, beta_smb, beta_hml = coeffs

        # R-squared
        y_pred = [alpha + beta_m * x_market[i] + beta_smb * smb_returns[i] + beta_hml * hml_returns[i] for i in range(n)]
        ss_res = sum((y[i] - y_pred[i]) ** 2 for i in range(n))
        mean_y = sum(y) / n
        ss_tot = sum((y[i] - mean_y) ** 2 for i in range(n))
        r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else 0.0
        residual_std = math.sqrt(ss_res / max(n - k, 1))

        return json.dumps({
            "model_type": "factor_model",
            "factors": "three_factor",
            "alpha": round(alpha, 4),
            "beta_market": round(beta_m, 4),
            "beta_smb": round(beta_smb, 4),
            "beta_hml": round(beta_hml, 4),
            "r_squared": round(r_squared, 4),
            "residual_std": round(residual_std, 4),
            "num_observations": n,
        })


def run_stress_test(
    db: Session,
    portfolio: list,
    scenarios: list,
) -> str:
    """
    Run stress test scenarios on a portfolio.

    Each scenario defines shocks (pct changes) per asset.
    Assets not in shocks default to 0% change.
    """
    if not portfolio:
        return json.dumps({"error": "portfolio must not be empty"})
    if not scenarios:
        return json.dumps({"error": "scenarios must not be empty"})

    total_value = sum(p["current_value"] * p["weight"] for p in portfolio)
    if total_value <= 0:
        return json.dumps({"error": "total portfolio value must be positive"})

    scenario_results = []
    for scenario in scenarios:
        name = scenario["name"]
        shocks = scenario.get("shocks", {})
        details = []
        total_pnl = 0.0
        worst_pnl = float("inf")
        best_pnl = float("-inf")
        worst_asset = None
        best_asset = None

        for p in portfolio:
            asset_name = p["asset"]
            asset_value = p["current_value"] * p["weight"]
            shock_pct = shocks.get(asset_name, 0.0)
            pnl = asset_value * shock_pct
            total_pnl += pnl
            details.append({
                "asset": asset_name,
                "weight": round(p["weight"], 4),
                "shock_pct": round(shock_pct, 4),
                "pnl": round(pnl, 4),
            })
            if pnl < worst_pnl:
                worst_pnl = pnl
                worst_asset = asset_name
            if pnl > best_pnl:
                best_pnl = pnl
                best_asset = asset_name

        portfolio_pct_change = total_pnl / total_value if total_value > 0 else 0.0
        scenario_results.append({
            "name": name,
            "portfolio_pnl": round(total_pnl, 4),
            "portfolio_pct_change": round(portfolio_pct_change, 4),
            "worst_asset": worst_asset,
            "best_asset": best_asset,
            "details": details,
        })

    return json.dumps({
        "model_type": "stress_test",
        "portfolio_value": round(total_value, 4),
        "scenario_results": scenario_results,
    })


# ── Phase 8: Derivatives & Options ────────────────────────────────────────────

def _norm_cdf(x: float) -> float:
    """Standard normal CDF using math.erf."""
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def _norm_pdf(x: float) -> float:
    """Standard normal PDF."""
    return math.exp(-x * x / 2.0) / math.sqrt(2.0 * math.pi)


def _bsm_d1_d2(
    spot: float, strike: float, t: float, r: float, sigma: float, q: float = 0.0,
) -> tuple[float, float]:
    """Compute d1 and d2 for Black-Scholes-Merton."""
    d1 = (math.log(spot / strike) + (r - q + sigma * sigma / 2.0) * t) / (sigma * math.sqrt(t))
    d2 = d1 - sigma * math.sqrt(t)
    return d1, d2


def _bsm_price(
    spot: float, strike: float, t: float, r: float, sigma: float,
    option_type: str = "call", q: float = 0.0,
) -> float:
    """Return BSM option price (call or put)."""
    d1, d2 = _bsm_d1_d2(spot, strike, t, r, sigma, q)
    if option_type == "call":
        return spot * math.exp(-q * t) * _norm_cdf(d1) - strike * math.exp(-r * t) * _norm_cdf(d2)
    else:
        return strike * math.exp(-r * t) * _norm_cdf(-d2) - spot * math.exp(-q * t) * _norm_cdf(-d1)


def _bsm_vega_raw(
    spot: float, strike: float, t: float, r: float, sigma: float, q: float = 0.0,
) -> float:
    """Return BSM vega (raw, not per-1%)."""
    d1, _ = _bsm_d1_d2(spot, strike, t, r, sigma, q)
    return spot * math.exp(-q * t) * _norm_pdf(d1) * math.sqrt(t)


def price_option_bsm(
    db: Session,
    spot: float,
    strike: float,
    time_to_expiry: float,
    risk_free_rate: float,
    volatility: float,
    option_type: str = "call",
    dividend_yield: float = 0.0,
) -> str:
    """Price a European option using Black-Scholes-Merton."""
    option_type = option_type.lower()
    if option_type not in ("call", "put"):
        return json.dumps({"error": "option_type must be 'call' or 'put'"})
    if volatility <= 0:
        return json.dumps({"error": "volatility must be positive"})
    if spot <= 0 or strike <= 0:
        return json.dumps({"error": "spot and strike must be positive"})

    S, K, T, r, sigma, q = spot, strike, time_to_expiry, risk_free_rate, volatility, dividend_yield

    # Handle expired option — return intrinsic value
    if T <= 0:
        if option_type == "call":
            intrinsic = max(S - K, 0.0)
        else:
            intrinsic = max(K - S, 0.0)
        return json.dumps({
            "model_type": "bsm",
            "price": round(intrinsic, 4),
            "d1": None,
            "d2": None,
            "intrinsic_value": round(intrinsic, 4),
            "time_value": 0.0,
            "inputs": {
                "spot": S, "strike": K, "time_to_expiry": T,
                "risk_free_rate": r, "volatility": sigma,
                "option_type": option_type, "dividend_yield": q,
            },
        })

    d1, d2 = _bsm_d1_d2(S, K, T, r, sigma, q)
    price = _bsm_price(S, K, T, r, sigma, option_type, q)

    if option_type == "call":
        intrinsic = max(S - K, 0.0)
    else:
        intrinsic = max(K - S, 0.0)
    time_val = price - intrinsic

    return json.dumps({
        "model_type": "bsm",
        "price": round(price, 4),
        "d1": round(d1, 4),
        "d2": round(d2, 4),
        "intrinsic_value": round(intrinsic, 4),
        "time_value": round(time_val, 4),
        "inputs": {
            "spot": S, "strike": K, "time_to_expiry": T,
            "risk_free_rate": r, "volatility": sigma,
            "option_type": option_type, "dividend_yield": q,
        },
    })


def price_option_binomial(
    db: Session,
    spot: float,
    strike: float,
    time_to_expiry: float,
    risk_free_rate: float,
    volatility: float,
    option_type: str = "call",
    dividend_yield: float = 0.0,
    steps: int = 100,
    exercise: str = "european",
) -> str:
    """Price an option using the Cox-Ross-Rubinstein binomial tree."""
    option_type = option_type.lower()
    exercise = exercise.lower()
    if option_type not in ("call", "put"):
        return json.dumps({"error": "option_type must be 'call' or 'put'"})
    if exercise not in ("european", "american"):
        return json.dumps({"error": "exercise must be 'european' or 'american'"})
    if volatility <= 0:
        return json.dumps({"error": "volatility must be positive"})
    if spot <= 0 or strike <= 0:
        return json.dumps({"error": "spot and strike must be positive"})
    if time_to_expiry <= 0:
        # Expired — intrinsic value
        if option_type == "call":
            intrinsic = max(spot - strike, 0.0)
        else:
            intrinsic = max(strike - spot, 0.0)
        return json.dumps({
            "model_type": "binomial_tree",
            "price": round(intrinsic, 4),
            "exercise": exercise,
            "steps": steps,
            "u": None, "d": None, "p": None,
        })
    steps = max(steps, 1)

    S, K, T, r, sigma, q = spot, strike, time_to_expiry, risk_free_rate, volatility, dividend_yield
    dt = T / steps
    u = math.exp(sigma * math.sqrt(dt))
    d_val = 1.0 / u
    disc = math.exp(-r * dt)
    p = (math.exp((r - q) * dt) - d_val) / (u - d_val)

    # Terminal payoffs
    prices = [S * (u ** (steps - i)) * (d_val ** i) for i in range(steps + 1)]
    if option_type == "call":
        values = [max(px - K, 0.0) for px in prices]
    else:
        values = [max(K - px, 0.0) for px in prices]

    # Backward induction
    for step in range(steps - 1, -1, -1):
        for i in range(step + 1):
            hold = disc * (p * values[i] + (1.0 - p) * values[i + 1])
            if exercise == "american":
                node_price = S * (u ** (step - i)) * (d_val ** i)
                if option_type == "call":
                    ex_val = max(node_price - K, 0.0)
                else:
                    ex_val = max(K - node_price, 0.0)
                values[i] = max(hold, ex_val)
            else:
                values[i] = hold

    return json.dumps({
        "model_type": "binomial_tree",
        "price": round(values[0], 4),
        "exercise": exercise,
        "steps": steps,
        "u": round(u, 4),
        "d": round(d_val, 4),
        "p": round(p, 4),
    })


def compute_greeks(
    db: Session,
    spot: float,
    strike: float,
    time_to_expiry: float,
    risk_free_rate: float,
    volatility: float,
    option_type: str = "call",
    dividend_yield: float = 0.0,
) -> str:
    """Compute option Greeks (delta, gamma, vega, theta, rho) using BSM."""
    option_type = option_type.lower()
    if option_type not in ("call", "put"):
        return json.dumps({"error": "option_type must be 'call' or 'put'"})
    if volatility <= 0:
        return json.dumps({"error": "volatility must be positive"})
    if time_to_expiry <= 0:
        return json.dumps({"error": "time_to_expiry must be positive for Greeks computation"})
    if spot <= 0 or strike <= 0:
        return json.dumps({"error": "spot and strike must be positive"})

    S, K, T, r, sigma, q = spot, strike, time_to_expiry, risk_free_rate, volatility, dividend_yield
    d1, d2 = _bsm_d1_d2(S, K, T, r, sigma, q)
    sqrt_t = math.sqrt(T)
    exp_qt = math.exp(-q * T)
    exp_rt = math.exp(-r * T)
    nd1 = _norm_pdf(d1)

    # Delta
    if option_type == "call":
        delta = exp_qt * _norm_cdf(d1)
    else:
        delta = exp_qt * (_norm_cdf(d1) - 1.0)

    # Gamma (same for call and put)
    gamma = exp_qt * nd1 / (S * sigma * sqrt_t)

    # Vega (per 1% vol change)
    vega = S * exp_qt * nd1 * sqrt_t / 100.0

    # Theta (per day)
    if option_type == "call":
        theta = (
            -(S * nd1 * sigma * exp_qt) / (2.0 * sqrt_t)
            - r * K * exp_rt * _norm_cdf(d2)
            + q * S * exp_qt * _norm_cdf(d1)
        ) / 365.0
    else:
        theta = (
            -(S * nd1 * sigma * exp_qt) / (2.0 * sqrt_t)
            + r * K * exp_rt * _norm_cdf(-d2)
            - q * S * exp_qt * _norm_cdf(-d1)
        ) / 365.0

    # Rho (per 1% rate change)
    if option_type == "call":
        rho = K * T * exp_rt * _norm_cdf(d2) / 100.0
    else:
        rho = -K * T * exp_rt * _norm_cdf(-d2) / 100.0

    return json.dumps({
        "model_type": "greeks",
        "delta": round(delta, 4),
        "gamma": round(gamma, 4),
        "vega": round(vega, 4),
        "theta": round(theta, 4),
        "rho": round(rho, 4),
        "d1": round(d1, 4),
        "d2": round(d2, 4),
    })


def compute_implied_volatility(
    db: Session,
    market_price: float,
    spot: float,
    strike: float,
    time_to_expiry: float,
    risk_free_rate: float,
    option_type: str = "call",
    dividend_yield: float = 0.0,
    max_iterations: int = 100,
    tolerance: float = 1e-6,
) -> str:
    """Compute implied volatility using Newton-Raphson with bisection fallback."""
    option_type = option_type.lower()
    if option_type not in ("call", "put"):
        return json.dumps({"error": "option_type must be 'call' or 'put'"})
    if market_price <= 0:
        return json.dumps({"error": "market_price must be positive"})
    if time_to_expiry <= 0:
        return json.dumps({"error": "time_to_expiry must be positive"})
    if spot <= 0 or strike <= 0:
        return json.dumps({"error": "spot and strike must be positive"})

    S, K, T, r, q = spot, strike, time_to_expiry, risk_free_rate, dividend_yield

    # Newton-Raphson
    sigma = 0.20
    converged = False
    iterations = 0

    for i in range(max_iterations):
        iterations = i + 1
        price = _bsm_price(S, K, T, r, sigma, option_type, q)
        vega_raw = _bsm_vega_raw(S, K, T, r, sigma, q)

        diff = price - market_price
        if abs(diff) < tolerance:
            converged = True
            break

        if abs(vega_raw) < 1e-12:
            break  # vega too small, switch to bisection

        sigma = sigma - diff / vega_raw
        sigma = max(0.001, min(sigma, 5.0))

    # If Newton didn't converge, try bisection
    if not converged:
        lo, hi = 0.001, 5.0
        for i in range(max_iterations):
            iterations += 1
            mid = (lo + hi) / 2.0
            price = _bsm_price(S, K, T, r, mid, option_type, q)
            diff = price - market_price
            if abs(diff) < tolerance:
                sigma = mid
                converged = True
                break
            if diff > 0:
                hi = mid
            else:
                lo = mid
        if not converged:
            sigma = (lo + hi) / 2.0

    return json.dumps({
        "model_type": "implied_volatility",
        "implied_volatility": round(sigma, 4),
        "iterations": iterations,
        "convergence": "converged" if converged else "not_converged",
    })


def check_put_call_parity(
    db: Session,
    call_price: float,
    put_price: float,
    spot: float,
    strike: float,
    time_to_expiry: float,
    risk_free_rate: float,
    dividend_yield: float = 0.0,
) -> str:
    """Check European put-call parity and detect arbitrage opportunities."""
    if time_to_expiry < 0:
        return json.dumps({"error": "time_to_expiry must be non-negative"})
    if spot <= 0 or strike <= 0:
        return json.dumps({"error": "spot and strike must be positive"})

    S, K, T, r, q = spot, strike, time_to_expiry, risk_free_rate, dividend_yield

    lhs = call_price - put_price  # C - P
    rhs = S * math.exp(-q * T) - K * math.exp(-r * T)  # S·e^(-qT) - K·e^(-rT)
    deviation = lhs - rhs

    threshold = 0.50
    arbitrage = abs(deviation) > threshold

    # Describe the arbitrage strategy
    if not arbitrage:
        strategy = "No significant arbitrage opportunity detected."
    elif deviation > 0:
        # C - P > S·e^(-qT) - K·e^(-rT): call overpriced relative to put
        strategy = (
            "Call overpriced relative to put. Strategy: sell call, buy put, "
            "buy underlying, borrow K·e^(-rT)."
        )
    else:
        # C - P < S·e^(-qT) - K·e^(-rT): put overpriced relative to call
        strategy = (
            "Put overpriced relative to call. Strategy: buy call, sell put, "
            "sell underlying, invest K·e^(-rT)."
        )

    return json.dumps({
        "model_type": "put_call_parity",
        "lhs_c_minus_p": round(lhs, 4),
        "rhs_s_minus_pvk": round(rhs, 4),
        "deviation": round(deviation, 4),
        "arbitrage_opportunity": arbitrage,
        "strategy": strategy,
    })


def build_option_strategy(
    db: Session,
    legs: list,
    underlying_price: float,
    spot_min: float = 0.0,
    spot_max: float = 0.0,
    spot_steps: int = 50,
) -> str:
    """Build and analyze a multi-leg option strategy (payoff diagram)."""
    if not legs:
        return json.dumps({"error": "At least one leg is required"})
    if underlying_price <= 0:
        return json.dumps({"error": "underlying_price must be positive"})

    # Default spot range
    if spot_max <= spot_min:
        spot_min = underlying_price * 0.7
        spot_max = underlying_price * 1.3

    spot_steps = max(spot_steps, 2)
    step_size = (spot_max - spot_min) / (spot_steps - 1)
    spot_prices = [spot_min + i * step_size for i in range(spot_steps)]

    payoff_table = []
    for sp in spot_prices:
        total_payoff = 0.0
        total_cost = 0.0
        for leg in legs:
            leg_type = leg.get("type", "call").lower()
            position = leg.get("position", "long").lower()
            strike_l = leg.get("strike", 0.0)
            premium = leg.get("premium", 0.0)
            sign = 1.0 if position == "long" else -1.0

            if leg_type == "call":
                payoff = max(sp - strike_l, 0.0)
            elif leg_type == "put":
                payoff = max(strike_l - sp, 0.0)
            else:  # stock
                payoff = sp - strike_l  # strike acts as purchase price

            total_payoff += sign * payoff
            total_cost += sign * premium

        profit = total_payoff - total_cost
        payoff_table.append({
            "spot": round(sp, 4),
            "payoff": round(total_payoff, 4),
            "profit": round(profit, 4),
        })

    # Find max profit, max loss
    profits = [row["profit"] for row in payoff_table]
    max_profit = max(profits)
    max_loss = min(profits)

    # Find breakevens by linear interpolation
    breakevens = []
    for i in range(len(payoff_table) - 1):
        p1 = payoff_table[i]["profit"]
        p2 = payoff_table[i + 1]["profit"]
        if p1 * p2 < 0:  # sign change
            s1 = payoff_table[i]["spot"]
            s2 = payoff_table[i + 1]["spot"]
            # Linear interpolation
            be = s1 + (s2 - s1) * abs(p1) / (abs(p1) + abs(p2))
            breakevens.append(round(be, 4))
        elif p1 == 0.0:
            breakevens.append(payoff_table[i]["spot"])

    return json.dumps({
        "model_type": "option_strategy",
        "legs": legs,
        "underlying_price": round(underlying_price, 4),
        "max_profit": round(max_profit, 4),
        "max_loss": round(max_loss, 4),
        "breakevens": breakevens,
        "payoff_table": payoff_table,
    })


# ── Phase 9: Iranian Market & Real Estate ──────────────────────────────────────

def compute_real_estate_noi(
    db: Session,
    gross_rental_income: float,
    vacancy_rate: float,
    operating_expenses: float,
    property_value: Optional[float] = None,
    cap_rate: Optional[float] = None,
    debt_service: Optional[float] = None,
    equity_invested: Optional[float] = None,
) -> str:
    """Compute Net Operating Income and related real estate metrics."""
    if gross_rental_income <= 0:
        return json.dumps({"error": "gross_rental_income must be positive"})
    if not (0 <= vacancy_rate < 1):
        return json.dumps({"error": "vacancy_rate must be between 0 and 1"})

    egi = gross_rental_income * (1 - vacancy_rate)
    noi = egi - operating_expenses

    result: dict = {
        "model_type": "real_estate_noi",
        "gross_rental_income": round(gross_rental_income, 4),
        "vacancy_rate": round(vacancy_rate, 4),
        "effective_gross_income": round(egi, 4),
        "operating_expenses": round(operating_expenses, 4),
        "noi": round(noi, 4),
    }

    if property_value is not None and property_value > 0:
        result["property_value"] = round(property_value, 4)
        result["cap_rate"] = round(noi / property_value, 4)
        if gross_rental_income > 0:
            result["grm"] = round(property_value / gross_rental_income, 4)
    elif cap_rate is not None and cap_rate > 0:
        implied_value = noi / cap_rate
        result["cap_rate"] = round(cap_rate, 4)
        result["implied_value"] = round(implied_value, 4)

    if debt_service is not None and debt_service > 0:
        dscr = noi / debt_service
        cash_after_debt = noi - debt_service
        result["debt_service"] = round(debt_service, 4)
        result["dscr"] = round(dscr, 4)
        result["cash_after_debt"] = round(cash_after_debt, 4)
        if equity_invested is not None and equity_invested > 0:
            result["cash_on_cash_return"] = round(cash_after_debt / equity_invested, 4)

    return json.dumps(result)


def build_development_proforma(
    db: Session,
    land_cost: float,
    construction_cost_per_sqm: float,
    total_sqm: float,
    sellable_pct: float,
    sale_price_per_sqm: float,
    construction_months: int,
    absorption_months: int = 6,
    financing_rate: float = 0.20,
    equity_pct: float = 0.40,
) -> str:
    """Build a real estate development proforma."""
    if total_sqm <= 0 or sellable_pct <= 0:
        return json.dumps({"error": "total_sqm and sellable_pct must be positive"})
    if not (0 < equity_pct <= 1):
        return json.dumps({"error": "equity_pct must be between 0 and 1"})

    construction_cost = construction_cost_per_sqm * total_sqm
    hard_costs = land_cost + construction_cost
    equity_amount = hard_costs * equity_pct
    debt_amount = hard_costs * (1 - equity_pct)

    total_months = construction_months + absorption_months
    monthly_rate = financing_rate / 12
    # Average draw: 0.5 factor for gradual drawdown
    financing_cost = debt_amount * monthly_rate * total_months * 0.5
    total_cost = hard_costs + financing_cost

    sellable_area = total_sqm * sellable_pct
    revenue = sellable_area * sale_price_per_sqm
    gross_profit = revenue - total_cost
    margin = gross_profit / revenue if revenue > 0 else 0.0

    equity_multiple = (gross_profit + equity_amount) / equity_amount if equity_amount > 0 else 0.0
    # Approximate annualized IRR from equity multiple
    irr = (equity_multiple ** (12 / total_months) - 1) if total_months > 0 and equity_multiple > 0 else 0.0
    breakeven_price = total_cost / sellable_area if sellable_area > 0 else 0.0

    return json.dumps({
        "model_type": "development_proforma",
        "land_cost": round(land_cost, 4),
        "construction_cost": round(construction_cost, 4),
        "hard_costs": round(hard_costs, 4),
        "equity_amount": round(equity_amount, 4),
        "debt_amount": round(debt_amount, 4),
        "financing_cost": round(financing_cost, 4),
        "total_cost": round(total_cost, 4),
        "sellable_area": round(sellable_area, 4),
        "revenue": round(revenue, 4),
        "gross_profit": round(gross_profit, 4),
        "margin": round(margin, 4),
        "equity_multiple": round(equity_multiple, 4),
        "irr": round(irr, 4),
        "breakeven_price": round(breakeven_price, 4),
        "total_months": total_months,
    })


def build_sukuk_model(
    db: Session,
    face_value: float,
    profit_rate: float,
    periods: int,
    sukuk_type: str = "ijara",
    market_yield: Optional[float] = None,
) -> str:
    """Build an Islamic bond (sukuk) model for ijara, murabaha, or musharaka."""
    sukuk_type = sukuk_type.lower()
    if sukuk_type not in ("ijara", "murabaha", "musharaka"):
        return json.dumps({"error": "sukuk_type must be 'ijara', 'murabaha', or 'musharaka'"})
    if periods <= 0:
        return json.dumps({"error": "periods must be positive"})
    if face_value <= 0:
        return json.dumps({"error": "face_value must be positive"})

    cash_flows: list[float] = []

    if sukuk_type == "ijara":
        # Periodic rental payments + face returned at maturity (like a bond)
        periodic_payment = face_value * profit_rate
        for t in range(1, periods + 1):
            if t < periods:
                cash_flows.append(periodic_payment)
            else:
                cash_flows.append(periodic_payment + face_value)

    elif sukuk_type == "murabaha":
        # Total price = cost + markup, divided into equal installments
        total_price = face_value + face_value * profit_rate
        installment = total_price / periods
        cash_flows = [installment] * periods

    elif sukuk_type == "musharaka":
        # Declining balance: each period buyback face/periods + profit share on remaining
        buyback = face_value / periods
        for t in range(1, periods + 1):
            remaining = face_value - buyback * (t - 1)
            profit_share = remaining * profit_rate
            cash_flows.append(buyback + profit_share)

    # Compute price if market_yield provided
    price = None
    if market_yield is not None and market_yield > 0:
        price = sum(cf / (1 + market_yield) ** t for t, cf in enumerate(cash_flows, 1))
        price = round(price, 4)

    # Compute Macaulay duration (weight by PV of cash flows)
    discount_rate = market_yield if market_yield is not None and market_yield > 0 else profit_rate
    duration = None
    if discount_rate > 0:
        pv_total = sum(cf / (1 + discount_rate) ** t for t, cf in enumerate(cash_flows, 1))
        if pv_total > 0:
            weighted_sum = sum(t * cf / (1 + discount_rate) ** t for t, cf in enumerate(cash_flows, 1))
            duration = round(weighted_sum / pv_total, 4)

    total_payments = round(sum(cash_flows), 4)

    result: dict = {
        "model_type": "sukuk",
        "sukuk_type": sukuk_type,
        "face_value": round(face_value, 4),
        "profit_rate": round(profit_rate, 4),
        "periods": periods,
        "cash_flows": [round(cf, 4) for cf in cash_flows],
        "total_payments": total_payments,
    }

    if price is not None:
        result["market_yield"] = round(market_yield, 4)
        result["price"] = price
        result["premium_discount"] = round(price - face_value, 4)

    if duration is not None:
        result["macaulay_duration"] = duration

    return json.dumps(result)


def build_murabaha_schedule(
    db: Session,
    cost_price: float,
    markup_rate: float,
    installments: int,
    grace_months: int = 0,
) -> str:
    """Build a Murabaha (cost-plus financing) payment schedule."""
    if cost_price <= 0:
        return json.dumps({"error": "cost_price must be positive"})
    if installments <= 0:
        return json.dumps({"error": "installments must be positive"})

    profit_amount = cost_price * markup_rate
    total_price = cost_price + profit_amount
    installment_amount = total_price / installments

    schedule = []
    remaining = total_price

    # Grace months (no payment)
    for p in range(1, grace_months + 1):
        schedule.append({
            "period": p,
            "payment": 0.0,
            "remaining": round(remaining, 4),
        })

    # Installment months
    for p in range(1, installments + 1):
        remaining -= installment_amount
        schedule.append({
            "period": grace_months + p,
            "payment": round(installment_amount, 4),
            "remaining": round(max(remaining, 0.0), 4),
        })

    return json.dumps({
        "model_type": "murabaha_schedule",
        "cost_price": round(cost_price, 4),
        "markup_rate": round(markup_rate, 4),
        "profit_amount": round(profit_amount, 4),
        "total_price": round(total_price, 4),
        "installment_amount": round(installment_amount, 4),
        "installments": installments,
        "grace_months": grace_months,
        "schedule": schedule,
    })


def build_ijara_model(
    db: Session,
    asset_value: float,
    lease_term_months: int,
    monthly_rent: float,
    transfer_price: float,
    maintenance_pct: float = 0.01,
) -> str:
    """Build an Ijara (Islamic lease) model with optional transfer of ownership."""
    if asset_value <= 0:
        return json.dumps({"error": "asset_value must be positive"})
    if lease_term_months <= 0:
        return json.dumps({"error": "lease_term_months must be positive"})

    total_rental = monthly_rent * lease_term_months
    maintenance = asset_value * maintenance_pct * (lease_term_months / 12)
    net_rental = total_rental - maintenance
    net_yield = net_rental / asset_value if asset_value > 0 else 0.0

    # Approximate effective annual rate:
    # PV of all payments should equal asset_value
    # Payments: monthly_rent for each month + transfer_price at end
    # Use bisection to solve for monthly rate
    effective_annual_rate = 0.0
    lo, hi = 0.0, 1.0  # monthly rate range
    for _ in range(200):
        mid = (lo + hi) / 2.0
        if mid == 0:
            pv = monthly_rent * lease_term_months + transfer_price
        else:
            pv = sum(monthly_rent / (1 + mid) ** t for t in range(1, lease_term_months + 1))
            pv += transfer_price / (1 + mid) ** lease_term_months
        if pv > asset_value:
            lo = mid
        else:
            hi = mid
    effective_annual_rate = (1 + (lo + hi) / 2.0) ** 12 - 1

    # Build schedule
    schedule = []
    for t in range(1, lease_term_months + 1):
        entry: dict = {
            "month": t,
            "payment": round(monthly_rent, 4),
            "type": "rent",
        }
        if t == lease_term_months:
            entry["transfer_price"] = round(transfer_price, 4)
            entry["total_payment"] = round(monthly_rent + transfer_price, 4)
            entry["type"] = "rent+transfer"
        schedule.append(entry)

    return json.dumps({
        "model_type": "ijara",
        "asset_value": round(asset_value, 4),
        "lease_term_months": lease_term_months,
        "monthly_rent": round(monthly_rent, 4),
        "transfer_price": round(transfer_price, 4),
        "total_rental": round(total_rental, 4),
        "maintenance": round(maintenance, 4),
        "net_rental": round(net_rental, 4),
        "net_yield": round(net_yield, 4),
        "effective_annual_rate": round(effective_annual_rate, 4),
        "schedule_length": len(schedule),
        "schedule": schedule[:12],  # first 12 months for brevity
    })


def compute_inflation_adjusted_valuation(
    db: Session,
    nominal_values: list,
    base_year: int,
    cpi_values: Optional[list] = None,
    inflation_rates: Optional[list] = None,
    metric_name: str = "value",
) -> str:
    """Deflate nominal values to real terms using CPI or inflation rates."""
    if not nominal_values:
        return json.dumps({"error": "nominal_values list is required"})
    if cpi_values is None and inflation_rates is None:
        return json.dumps({"error": "Either cpi_values or inflation_rates must be provided"})

    # Build CPI lookup
    cpi_lookup: dict = {}

    if cpi_values is not None:
        for entry in cpi_values:
            cpi_lookup[entry["year"]] = entry["cpi"]
    elif inflation_rates is not None:
        # Build CPI from inflation rates; base_year CPI = 100
        rates_by_year = {entry["year"]: entry["rate"] for entry in inflation_rates}
        all_years = sorted(set(
            [nv["year"] for nv in nominal_values] + list(rates_by_year.keys()) + [base_year]
        ))
        # Set base_year CPI = 100, then forward/backward from there
        cpi_lookup[base_year] = 100.0
        # Forward from base_year
        for i in range(len(all_years)):
            yr = all_years[i]
            if yr > base_year and yr not in cpi_lookup:
                prev_yr = yr - 1
                if prev_yr in cpi_lookup and prev_yr in rates_by_year:
                    cpi_lookup[yr] = cpi_lookup[prev_yr] * (1 + rates_by_year[prev_yr])
        # Backward from base_year
        for i in range(len(all_years) - 1, -1, -1):
            yr = all_years[i]
            if yr < base_year and yr not in cpi_lookup:
                next_yr = yr + 1
                if next_yr in cpi_lookup and yr in rates_by_year:
                    cpi_lookup[yr] = cpi_lookup[next_yr] / (1 + rates_by_year[yr])

    # Get base CPI
    if base_year not in cpi_lookup:
        return json.dumps({"error": f"CPI for base_year {base_year} not available"})
    base_cpi = cpi_lookup[base_year]

    # Deflate
    adjusted = []
    for nv in nominal_values:
        yr = nv["year"]
        nominal = nv["value"]
        if yr in cpi_lookup and cpi_lookup[yr] > 0:
            real_val = nominal * (base_cpi / cpi_lookup[yr])
            adjusted.append({
                "year": yr,
                "nominal": round(nominal, 4),
                "real": round(real_val, 4),
                "cpi": round(cpi_lookup[yr], 4),
            })
        else:
            adjusted.append({
                "year": yr,
                "nominal": round(nominal, 4),
                "real": None,
                "cpi": None,
            })

    # Compute CAGR for nominal and real
    valid = [a for a in adjusted if a["real"] is not None]
    nominal_cagr = None
    real_cagr = None
    if len(valid) >= 2:
        first = valid[0]
        last = valid[-1]
        n_years = last["year"] - first["year"]
        if n_years > 0:
            if first["nominal"] > 0 and last["nominal"] > 0:
                nominal_cagr = round((last["nominal"] / first["nominal"]) ** (1 / n_years) - 1, 4)
            if first["real"] > 0 and last["real"] > 0:
                real_cagr = round((last["real"] / first["real"]) ** (1 / n_years) - 1, 4)

    inflation_impact_pct = None
    if nominal_cagr is not None and real_cagr is not None and nominal_cagr > 0:
        inflation_impact_pct = round(1 - real_cagr / nominal_cagr, 4)

    return json.dumps({
        "model_type": "inflation_adjusted",
        "metric_name": metric_name,
        "base_year": base_year,
        "adjusted_values": adjusted,
        "nominal_cagr": nominal_cagr,
        "real_cagr": real_cagr,
        "inflation_impact_pct": inflation_impact_pct,
    })


def build_tehran_housing_model(
    db: Session,
    area_sqm: float,
    price_per_sqm: float,
    monthly_rent_per_sqm: float,
    annual_appreciation_pct: float,
    maintenance_cost_pct: float = 0.01,
    vacancy_months_per_year: float = 0,
    mortgage_amount: Optional[float] = None,
    mortgage_rate: Optional[float] = None,
    mortgage_term_months: Optional[int] = None,
) -> str:
    """Build a Tehran housing investment model with yield, appreciation, and mortgage analysis."""
    if area_sqm <= 0 or price_per_sqm <= 0:
        return json.dumps({"error": "area_sqm and price_per_sqm must be positive"})

    property_value = area_sqm * price_per_sqm
    annual_rent = monthly_rent_per_sqm * area_sqm * 12 * ((12 - vacancy_months_per_year) / 12)
    annual_maintenance = property_value * maintenance_cost_pct
    gross_yield = annual_rent / property_value if property_value > 0 else 0.0
    net_yield = (annual_rent - annual_maintenance) / property_value if property_value > 0 else 0.0

    result: dict = {
        "model_type": "tehran_housing",
        "area_sqm": round(area_sqm, 4),
        "property_value": round(property_value, 4),
        "annual_rent": round(annual_rent, 4),
        "annual_maintenance": round(annual_maintenance, 4),
        "gross_yield": round(gross_yield, 4),
        "net_yield": round(net_yield, 4),
    }

    # Mortgage analysis
    monthly_payment = None
    total_interest = None
    if (mortgage_amount is not None and mortgage_rate is not None
            and mortgage_term_months is not None and mortgage_term_months > 0):
        monthly_rate = mortgage_rate / 12
        monthly_payment = _pmt(monthly_rate, mortgage_term_months, mortgage_amount)
        total_paid = monthly_payment * mortgage_term_months
        total_interest = total_paid - mortgage_amount
        result["mortgage_amount"] = round(mortgage_amount, 4)
        result["mortgage_rate"] = round(mortgage_rate, 4)
        result["mortgage_term_months"] = mortgage_term_months
        result["monthly_payment"] = round(monthly_payment, 4)
        result["total_interest"] = round(total_interest, 4)

    # 5-year total return projection
    appreciation_5yr = property_value * ((1 + annual_appreciation_pct) ** 5 - 1)
    rental_income_5yr = annual_rent * 5
    maintenance_5yr = annual_maintenance * 5
    mortgage_interest_5yr = 0.0
    if monthly_payment is not None:
        # Total mortgage payments over 5 years (or mortgage term if shorter)
        months_5yr = min(60, mortgage_term_months)
        mortgage_interest_5yr = monthly_payment * months_5yr - mortgage_amount * (months_5yr / mortgage_term_months)

    total_return_5yr = appreciation_5yr + rental_income_5yr - maintenance_5yr - mortgage_interest_5yr
    result["appreciation_5yr"] = round(appreciation_5yr, 4)
    result["rental_income_5yr"] = round(rental_income_5yr, 4)
    result["maintenance_5yr"] = round(maintenance_5yr, 4)
    result["total_return_5yr"] = round(total_return_5yr, 4)
    result["total_return_5yr_pct"] = round(total_return_5yr / property_value, 4) if property_value > 0 else 0.0

    # Buy vs rent breakeven (simplified)
    # Each year: ownership cost = maintenance + mortgage_interest_annual (if any) - appreciation
    # vs rent paid = annual_rent
    # Find year where cumulative ownership advantage turns positive
    breakeven_years = None
    cumulative_ownership_cost = 0.0
    cumulative_rent = 0.0
    for yr in range(1, 31):
        annual_appreciation = property_value * ((1 + annual_appreciation_pct) ** yr
                                                - (1 + annual_appreciation_pct) ** (yr - 1))
        annual_mortgage_cost = 0.0
        if monthly_payment is not None and yr * 12 <= mortgage_term_months:
            annual_mortgage_cost = monthly_payment * 12 - mortgage_amount / (mortgage_term_months / 12)

        ownership_year_cost = annual_maintenance + annual_mortgage_cost - annual_appreciation
        cumulative_ownership_cost += ownership_year_cost
        cumulative_rent += annual_rent

        if cumulative_rent > cumulative_ownership_cost and breakeven_years is None:
            breakeven_years = yr
            break

    result["buy_vs_rent_breakeven_years"] = breakeven_years

    return json.dumps(result)


# ── Phase 10: Excel Formula Gap Coverage ────────────────────────────────────


def compute_dupont(
    db: Session,
    net_income: float,
    sales: float,
    total_assets: float,
    total_equity: float,
    ebit: Optional[float] = None,
    ebt: Optional[float] = None,
    tax: Optional[float] = None,
) -> str:
    """DuPont analysis: 3-factor (basic) or 5-factor (with EBIT/EBT breakdown)."""
    if sales <= 0:
        return json.dumps({"error": "sales must be > 0"})
    if total_assets <= 0:
        return json.dumps({"error": "total_assets must be > 0"})
    if total_equity <= 0:
        return json.dumps({"error": "total_equity must be > 0"})

    asset_turnover = sales / total_assets
    equity_multiplier = total_assets / total_equity

    if (ebit is None) != (ebt is None):
        return json.dumps({"error": "For 5-factor DuPont, both ebit and ebt must be provided"})

    if ebit is not None and ebt is not None:
        # 5-factor DuPont
        if ebit == 0:
            return json.dumps({"error": "ebit must not be zero for 5-factor DuPont"})
        if ebt == 0:
            return json.dumps({"error": "ebt must not be zero for 5-factor DuPont"})

        tax_burden = net_income / ebt
        interest_burden = ebt / ebit
        ebit_margin = ebit / sales

        roe = tax_burden * interest_burden * ebit_margin * asset_turnover * equity_multiplier

        return json.dumps({
            "model_type": "dupont",
            "mode": "5_factor",
            "roe": round(roe, 4),
            "roe_pct": round(roe * 100, 4),
            "components": {
                "tax_burden": round(tax_burden, 4),
                "interest_burden": round(interest_burden, 4),
                "ebit_margin": round(ebit_margin, 4),
                "asset_turnover": round(asset_turnover, 4),
                "equity_multiplier": round(equity_multiplier, 4),
            },
        })
    else:
        # 3-factor DuPont
        profit_margin = net_income / sales
        roe = profit_margin * asset_turnover * equity_multiplier

        return json.dumps({
            "model_type": "dupont",
            "mode": "3_factor",
            "roe": round(roe, 4),
            "roe_pct": round(roe * 100, 4),
            "components": {
                "profit_margin": round(profit_margin, 4),
                "asset_turnover": round(asset_turnover, 4),
                "equity_multiplier": round(equity_multiplier, 4),
            },
        })


def compute_brinson_attribution(
    db: Session,
    sectors: list,
) -> str:
    """Brinson-Fachler performance attribution by sector."""
    if not sectors or len(sectors) == 0:
        return json.dumps({"error": "sectors list must be non-empty"})

    total_benchmark_return = sum(
        s["benchmark_weight"] * s["benchmark_return"] for s in sectors
    )

    wp_sum = sum(s["portfolio_weight"] for s in sectors)
    wb_sum = sum(s["benchmark_weight"] for s in sectors)

    weight_warning = None
    if abs(wp_sum - 1.0) > 0.02 or abs(wb_sum - 1.0) > 0.02:
        weight_warning = (
            f"Weights do not sum to 1: portfolio={round(wp_sum, 4)}, "
            f"benchmark={round(wb_sum, 4)}"
        )

    sector_results = []
    allocation_total = 0.0
    selection_total = 0.0
    interaction_total = 0.0

    for s in sectors:
        wp = s["portfolio_weight"]
        wb = s["benchmark_weight"]
        rp = s["portfolio_return"]
        rb = s["benchmark_return"]

        allocation = (wp - wb) * (rb - total_benchmark_return)
        selection = wb * (rp - rb)
        interaction = (wp - wb) * (rp - rb)

        allocation_total += allocation
        selection_total += selection
        interaction_total += interaction

        sector_results.append({
            "name": s["name"],
            "wp": round(wp, 4),
            "wb": round(wb, 4),
            "rp": round(rp, 4),
            "rb": round(rb, 4),
            "allocation": round(allocation, 4),
            "selection": round(selection, 4),
            "interaction": round(interaction, 4),
        })

    active_return = allocation_total + selection_total + interaction_total

    return json.dumps({
        "model_type": "brinson_attribution",
        "sectors": sector_results,
        "totals": {
            "allocation": round(allocation_total, 4),
            "selection": round(selection_total, 4),
            "interaction": round(interaction_total, 4),
            "active_return": round(active_return, 4),
        },
        "weight_warning": weight_warning,
    })


def compute_black_litterman(
    db: Session,
    market_caps: list,
    covariance_matrix: list,
    risk_aversion: float,
    tau: float,
    views: list,
    view_confidences: list,
    risk_free_rate: float = 0.0,
) -> str:
    """Black-Litterman asset allocation model."""
    n = len(market_caps)

    if n != len(covariance_matrix):
        return json.dumps({"error": "market_caps length must match covariance_matrix dimension"})
    if n > 10:
        return json.dumps({"error": "Maximum 10 assets supported (matrix inversion limit)"})
    for row in covariance_matrix:
        if len(row) != n:
            return json.dumps({"error": "covariance_matrix must be NxN"})
    if len(views) != len(view_confidences):
        return json.dumps({"error": "views and view_confidences must have same length"})
    if len(views) == 0:
        return json.dumps({"error": "At least one view is required"})

    # ── Matrix helpers ──
    def mat_mult(a, b):
        """Multiply two 2D lists."""
        rows_a, cols_a = len(a), len(a[0])
        cols_b = len(b[0])
        result = [[0.0] * cols_b for _ in range(rows_a)]
        for i in range(rows_a):
            for j in range(cols_b):
                for k in range(cols_a):
                    result[i][j] += a[i][k] * b[k][j]
        return result

    def mat_transpose(a):
        """Transpose a 2D list."""
        rows, cols = len(a), len(a[0])
        return [[a[i][j] for i in range(rows)] for j in range(cols)]

    def mat_scale(a, s):
        """Scale a matrix by a scalar."""
        return [[a[i][j] * s for j in range(len(a[0]))] for i in range(len(a))]

    def mat_add(a, b):
        """Add two matrices."""
        return [[a[i][j] + b[i][j] for j in range(len(a[0]))] for i in range(len(a))]

    def mat_identity(size):
        """Create identity matrix."""
        return [[1.0 if i == j else 0.0 for j in range(size)] for i in range(size)]

    def mat_minor(a, i, j):
        """Get minor matrix by removing row i and column j."""
        return [
            [a[r][c] for c in range(len(a[0])) if c != j]
            for r in range(len(a)) if r != i
        ]

    def mat_det(a):
        """Determinant via cofactor expansion."""
        size = len(a)
        if size == 1:
            return a[0][0]
        if size == 2:
            return a[0][0] * a[1][1] - a[0][1] * a[1][0]
        det = 0.0
        for j in range(size):
            det += ((-1) ** j) * a[0][j] * mat_det(mat_minor(a, 0, j))
        return det

    def mat_inverse(a):
        """Inverse via cofactor method. Returns None if singular."""
        size = len(a)
        det = mat_det(a)
        if abs(det) < 1e-14:
            return None
        if size == 1:
            return [[1.0 / det]]
        # Cofactor matrix
        cofactors = [[0.0] * size for _ in range(size)]
        for i in range(size):
            for j in range(size):
                cofactors[i][j] = ((-1) ** (i + j)) * mat_det(mat_minor(a, i, j))
        # Adjugate = transpose of cofactor
        adjugate = mat_transpose(cofactors)
        return mat_scale(adjugate, 1.0 / det)

    def vec_to_col(v):
        """List to column vector [[v0],[v1],...]."""
        return [[x] for x in v]

    def col_to_vec(m):
        """Column vector [[v0],[v1],...] to list."""
        return [m[i][0] for i in range(len(m))]

    def mat_diag(vals):
        """Create diagonal matrix from list of values."""
        size = len(vals)
        return [[vals[i] if i == j else 0.0 for j in range(size)] for i in range(size)]

    # ── Market weights ──
    total_mc = sum(market_caps)
    if total_mc <= 0:
        return json.dumps({"error": "Total market cap must be > 0"})
    w = [mc / total_mc for mc in market_caps]

    sigma = covariance_matrix
    delta = risk_aversion

    # ── Implied equilibrium returns: pi = delta * Sigma * w ──
    w_col = vec_to_col(w)
    pi_col = mat_mult(mat_scale(sigma, delta), w_col)
    pi = col_to_vec(pi_col)

    # ── Build P matrix (K x N) and Q vector (K x 1) ──
    k = len(views)
    p_matrix = [[0.0] * n for _ in range(k)]
    q_vector = []

    for vi, view in enumerate(views):
        for ai, asset_idx in enumerate(view["assets"]):
            if 0 <= asset_idx < n:
                p_matrix[vi][asset_idx] = view["weights"][ai]
        q_vector.append(view["expected_return"])

    # ── Omega = diag(p_i' * (tau*Sigma) * p_i / confidence_i) ──
    tau_sigma = mat_scale(sigma, tau)
    omega_diag = []
    for vi in range(k):
        p_row = [p_matrix[vi]]  # 1 x N
        p_row_t = vec_to_col(p_matrix[vi])  # N x 1
        var_view = mat_mult(mat_mult(p_row, tau_sigma), p_row_t)  # 1x1
        conf = view_confidences[vi] if view_confidences[vi] > 1e-10 else 1e-10
        omega_diag.append(var_view[0][0] / conf)

    omega = mat_diag(omega_diag)

    # ── BL returns: E[R] = inv(inv(tau*Sigma) + P'.inv(Omega).P) * (inv(tau*Sigma)*pi + P'.inv(Omega)*Q) ──
    tau_sigma_inv = mat_inverse(tau_sigma)
    if tau_sigma_inv is None:
        return json.dumps({"error": "tau*Sigma matrix is singular"})

    omega_inv = mat_inverse(omega)
    if omega_inv is None:
        return json.dumps({"error": "Omega matrix is singular"})

    p_t = mat_transpose(p_matrix)  # N x K

    # inv(tau*Sigma) + P' * inv(Omega) * P
    pt_omega_inv_p = mat_mult(mat_mult(p_t, omega_inv), p_matrix)  # N x N
    lhs_matrix = mat_add(tau_sigma_inv, pt_omega_inv_p)
    lhs_inv = mat_inverse(lhs_matrix)
    if lhs_inv is None:
        return json.dumps({"error": "BL posterior precision matrix is singular"})

    # inv(tau*Sigma) * pi + P' * inv(Omega) * Q
    term1 = mat_mult(tau_sigma_inv, vec_to_col(pi))  # N x 1
    q_col = vec_to_col(q_vector)  # K x 1
    term2 = mat_mult(mat_mult(p_t, omega_inv), q_col)  # N x 1
    rhs = [[term1[i][0] + term2[i][0]] for i in range(n)]

    bl_returns_col = mat_mult(lhs_inv, rhs)
    bl_returns = col_to_vec(bl_returns_col)

    # ── BL weights: w_bl = inv(delta * Sigma) * E[R], normalized ──
    delta_sigma = mat_scale(sigma, delta)
    delta_sigma_inv = mat_inverse(delta_sigma)
    if delta_sigma_inv is None:
        return json.dumps({"error": "delta*Sigma matrix is singular"})

    bl_weights_col = mat_mult(delta_sigma_inv, bl_returns_col)
    bl_weights_raw = col_to_vec(bl_weights_col)
    w_sum = sum(bl_weights_raw)
    if abs(w_sum) < 1e-14:
        bl_weights = bl_weights_raw
    else:
        bl_weights = [x / w_sum for x in bl_weights_raw]

    return json.dumps({
        "model_type": "black_litterman",
        "n_assets": n,
        "n_views": k,
        "market_weights": [round(x, 6) for x in w],
        "implied_returns": [round(x, 6) for x in pi],
        "bl_returns": [round(x, 6) for x in bl_returns],
        "bl_weights": [round(x, 6) for x in bl_weights],
        "risk_aversion": round(delta, 4),
        "tau": round(tau, 4),
    })


def compute_pe_fund_metrics(
    db: Session,
    contributions: list,
    distributions: list,
    nav: float,
    dates: Optional[list] = None,
) -> str:
    """Private equity fund performance metrics: TVPI, DPI, RVPI, and optionally IRR."""
    if not contributions or len(contributions) == 0:
        return json.dumps({"error": "contributions list must be non-empty"})
    if not distributions or len(distributions) == 0:
        return json.dumps({"error": "distributions list must be non-empty"})

    paid_in = sum(contributions)
    if paid_in <= 0:
        return json.dumps({"error": "Total paid-in capital must be > 0"})

    total_distributed = sum(distributions)
    tvpi = (total_distributed + nav) / paid_in
    dpi = total_distributed / paid_in
    rvpi = nav / paid_in

    mwr = None
    if dates is not None:
        # Build net cash flow series: negative for contributions, positive for distributions
        max_len = max(len(contributions), len(distributions))
        net_cash_flows = []
        for i in range(max_len):
            cf = 0.0
            if i < len(contributions):
                cf -= contributions[i]
            if i < len(distributions):
                cf += distributions[i]
            net_cash_flows.append(cf)
        # Final cash flow: add NAV
        net_cash_flows.append(nav)
        mwr = _irr(net_cash_flows)
        if mwr is not None:
            mwr = round(mwr, 6)

    vintage_summary = (
        f"TVPI {round(tvpi, 2)}x | DPI {round(dpi, 2)}x | RVPI {round(rvpi, 2)}x"
    )

    return json.dumps({
        "model_type": "pe_fund_metrics",
        "paid_in_capital": round(paid_in, 4),
        "total_distributed": round(total_distributed, 4),
        "nav": round(nav, 4),
        "tvpi": round(tvpi, 4),
        "dpi": round(dpi, 4),
        "rvpi": round(rvpi, 4),
        "mwr": mwr,
        "vintage_summary": vintage_summary,
    })


def compute_omega_ratio(
    db: Session,
    returns: list,
    threshold: float = 0.0,
) -> str:
    """Omega ratio and related downside risk metrics."""
    if len(returns) < 2:
        return json.dumps({"error": "At least 2 return observations required"})

    gains = sum(max(r - threshold, 0) for r in returns)
    losses = sum(max(threshold - r, 0) for r in returns)
    omega = gains / losses if losses > 1e-10 else 9999.99

    n = len(returns)
    pct_above = len([r for r in returns if r > threshold]) / n
    pct_below = 1.0 - pct_above
    mean_return = statistics.mean(returns)

    downside_returns = [min(r - threshold, 0) ** 2 for r in returns]
    downside_dev = math.sqrt(sum(downside_returns) / n)

    sortino_vs_threshold = (
        (mean_return - threshold) / downside_dev
        if downside_dev > 1e-10
        else 9999.99
    )

    upside_potential = statistics.mean([max(r - threshold, 0) for r in returns])
    upside_potential_ratio = (
        upside_potential / downside_dev
        if downside_dev > 1e-10
        else 9999.99
    )

    return json.dumps({
        "model_type": "omega_ratio",
        "omega": round(omega, 4),
        "threshold": round(threshold, 6),
        "n_periods": n,
        "mean_return": round(mean_return, 6),
        "pct_above": round(pct_above, 4),
        "pct_below": round(pct_below, 4),
        "upside_potential_ratio": round(upside_potential_ratio, 4),
        "sortino_vs_threshold": round(sortino_vs_threshold, 4),
        "gains_sum": round(gains, 6),
        "losses_sum": round(losses, 6),
    })


def compute_credit_risk(
    db: Session,
    ead: float,
    pd: float,
    lgd: float,
    asset_value: Optional[float] = None,
    debt_face: Optional[float] = None,
    asset_volatility: Optional[float] = None,
    time_horizon: Optional[float] = None,
    risk_free_rate: Optional[float] = None,
) -> str:
    """Credit risk metrics (EL, UL, Credit VaR) with optional Merton structural model."""
    if ead <= 0:
        return json.dumps({"error": "ead must be > 0"})
    if pd < 0 or pd > 1:
        return json.dumps({"error": "pd must be between 0 and 1"})
    if lgd < 0 or lgd > 1:
        return json.dumps({"error": "lgd must be between 0 and 1"})

    expected_loss = ead * pd * lgd
    unexpected_loss = ead * math.sqrt(pd * (1 - pd)) * lgd
    credit_var_99 = unexpected_loss * 2.326
    loss_rate = pd * lgd

    result = {
        "model_type": "credit_risk",
        "ead": round(ead, 4),
        "pd": round(pd, 6),
        "lgd": round(lgd, 4),
        "expected_loss": round(expected_loss, 4),
        "unexpected_loss": round(unexpected_loss, 4),
        "credit_var_99": round(credit_var_99, 4),
        "loss_rate": round(loss_rate, 6),
    }

    # Merton structural model (optional)
    if (asset_value is not None and debt_face is not None
            and asset_volatility is not None and time_horizon is not None):
        if asset_value <= 0:
            return json.dumps({"error": "asset_value must be > 0 for Merton model"})
        if debt_face <= 0:
            return json.dumps({"error": "debt_face must be > 0 for Merton model"})
        if asset_volatility <= 0:
            return json.dumps({"error": "asset_volatility must be > 0 for Merton model"})
        if time_horizon <= 0:
            return json.dumps({"error": "time_horizon must be > 0 for Merton model"})

        r = risk_free_rate if risk_free_rate is not None else 0.0

        d2 = (
            (math.log(asset_value / debt_face) + (r - asset_volatility ** 2 / 2) * time_horizon)
            / (asset_volatility * math.sqrt(time_horizon))
        )
        d1 = d2 + asset_volatility * math.sqrt(time_horizon)

        distance_to_default = d2
        pd_merton = 1 - _norm_cdf(d2)
        equity_value = (
            asset_value * _norm_cdf(d1)
            - debt_face * math.exp(-r * time_horizon) * _norm_cdf(d2)
        )

        result["distance_to_default"] = round(distance_to_default, 4)
        result["pd_merton"] = round(pd_merton, 6)
        result["equity_value"] = round(equity_value, 4)
        result["d1"] = round(d1, 4)
        result["d2"] = round(d2, 4)

    return json.dumps(result)


def compute_forward_rates(
    db: Session,
    spot_rates: Optional[list] = None,
    maturities: Optional[list] = None,
    par_rates: Optional[list] = None,
    cash_flows: Optional[list] = None,
    price: Optional[float] = None,
) -> str:
    """Forward rate computation, spot rate bootstrap, and z-spread calculation."""
    # ── Mode B: Bootstrap spots from par rates ──
    if par_rates is not None:
        n = len(par_rates)
        if n == 0:
            return json.dumps({"error": "par_rates list must be non-empty"})
        mats = maturities if maturities is not None else list(range(1, n + 1))
        if len(mats) != n:
            return json.dumps({"error": "maturities length must match par_rates length"})

        spots = [0.0] * n
        spots[0] = par_rates[0]

        for i in range(1, n):
            coupon = par_rates[i]
            # 1 = sum(c / (1+s_j)^t_j for j<i) + (1+c) / (1+s_i)^t_i
            pv_coupons = sum(
                coupon / (1 + spots[j]) ** mats[j] for j in range(i)
            )
            remaining = 1.0 - pv_coupons
            if remaining <= 0:
                return json.dumps({"error": f"Bootstrap failed at maturity {mats[i]}: non-positive residual"})
            # remaining = (1 + coupon) / (1 + s_i)^t_i
            # (1 + s_i)^t_i = (1 + coupon) / remaining
            spots[i] = ((1 + coupon) / remaining) ** (1.0 / mats[i]) - 1

        # Compute forwards from bootstrapped spots
        forwards = []
        for i in range(1, n):
            t1, t2 = mats[i - 1], mats[i]
            dt = t2 - t1
            if dt <= 0:
                forwards.append(None)
                continue
            fwd = ((1 + spots[i]) ** t2 / (1 + spots[i - 1]) ** t1) ** (1.0 / dt) - 1
            forwards.append(round(fwd, 6))

        discount_factors = [round(1.0 / (1 + spots[i]) ** mats[i], 6) for i in range(n)]

        return json.dumps({
            "model_type": "forward_rates",
            "mode": "bootstrap",
            "spot_rates": [round(s, 6) for s in spots],
            "forward_rates": forwards,
            "discount_factors": discount_factors,
            "z_spread": None,
            "maturities": mats,
        })

    # ── Require spot_rates for Mode A and Mode C ──
    if spot_rates is None or len(spot_rates) == 0:
        return json.dumps({"error": "spot_rates or par_rates must be provided"})

    n = len(spot_rates)
    mats = maturities if maturities is not None else list(range(1, n + 1))
    if len(mats) != n:
        return json.dumps({"error": "maturities length must match spot_rates length"})

    # ── Mode C: Z-Spread ──
    if cash_flows is not None and price is not None:
        if len(cash_flows) == 0:
            return json.dumps({"error": "cash_flows list must be non-empty"})
        # Use spot rates for each cash flow period; if fewer spots than CFs, reuse last
        def calc_price(z):
            pv = 0.0
            for i, cf in enumerate(cash_flows):
                t = mats[i] if i < len(mats) else mats[-1] + (i - len(mats) + 1)
                s = spot_rates[i] if i < n else spot_rates[-1]
                pv += cf / (1 + s + z) ** t
            return pv

        lo, hi = -0.5, 0.5
        z_spread = None
        for _ in range(200):
            mid = (lo + hi) / 2.0
            calc_p = calc_price(mid)
            if abs(calc_p - price) < 0.0001:
                z_spread = mid
                break
            if calc_p > price:
                lo = mid
            else:
                hi = mid
        else:
            z_spread = (lo + hi) / 2.0

        # Also compute forwards and discount factors
        forwards = []
        for i in range(1, n):
            t1, t2 = mats[i - 1], mats[i]
            dt = t2 - t1
            if dt <= 0:
                forwards.append(None)
                continue
            fwd = ((1 + spot_rates[i]) ** t2 / (1 + spot_rates[i - 1]) ** t1) ** (1.0 / dt) - 1
            forwards.append(round(fwd, 6))

        discount_factors = [round(1.0 / (1 + spot_rates[i]) ** mats[i], 6) for i in range(n)]

        return json.dumps({
            "model_type": "forward_rates",
            "mode": "z_spread",
            "spot_rates": [round(s, 6) for s in spot_rates],
            "forward_rates": forwards,
            "discount_factors": discount_factors,
            "z_spread": round(z_spread, 6) if z_spread is not None else None,
            "maturities": mats,
        })

    # ── Mode A: Forward rates from spot rates ──
    forwards = []
    for i in range(1, n):
        t1, t2 = mats[i - 1], mats[i]
        dt = t2 - t1
        if dt <= 0:
            forwards.append(None)
            continue
        fwd = ((1 + spot_rates[i]) ** t2 / (1 + spot_rates[i - 1]) ** t1) ** (1.0 / dt) - 1
        forwards.append(round(fwd, 6))

    discount_factors = [round(1.0 / (1 + spot_rates[i]) ** mats[i], 6) for i in range(n)]

    return json.dumps({
        "model_type": "forward_rates",
        "mode": "forward",
        "spot_rates": [round(s, 6) for s in spot_rates],
        "forward_rates": forwards,
        "discount_factors": discount_factors,
        "z_spread": None,
        "maturities": mats,
    })


# ── Tool Definitions ──────────────────────────────────────────────────────────

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "build_dcf_model",
            "description": (
                "Build a Discounted Cash Flow (DCF) valuation model. "
                "Calculates FCFF, enterprise value, equity value, and price per share. "
                "Creates an Excel spreadsheet with Assumptions, DCF Valuation, and Sensitivity tabs."
            ),
            "parameters": {
                "type": "object",
                "required": ["company_name", "projections", "wacc", "terminal_growth"],
                "properties": {
                    "company_name": {
                        "type": "string",
                        "description": "Company name (Persian or English)",
                    },
                    "projections": {
                        "type": "array",
                        "description": "Annual projections (1–10 years). Each item must include: ebit, tax_rate, da, capex, delta_wc (all in billion IRR).",
                        "items": {
                            "type": "object",
                            "required": ["ebit", "tax_rate", "da", "capex", "delta_wc"],
                            "properties": {
                                "ebit": {"type": "number", "description": "EBIT (billion IRR)"},
                                "tax_rate": {"type": "number", "description": "Tax rate decimal (e.g. 0.25 for 25%)"},
                                "da": {"type": "number", "description": "Depreciation & Amortization (billion IRR)"},
                                "capex": {"type": "number", "description": "Capital expenditure (billion IRR)"},
                                "delta_wc": {"type": "number", "description": "Change in working capital (billion IRR)"},
                            },
                        },
                    },
                    "wacc": {
                        "type": "number",
                        "description": "Weighted average cost of capital decimal (e.g. 0.22 for 22%). Typical TSE range: 0.20–0.28.",
                    },
                    "terminal_growth": {
                        "type": "number",
                        "description": "Terminal growth rate decimal (e.g. 0.03 for 3%). Must be less than WACC.",
                    },
                    "net_debt": {
                        "type": "number",
                        "description": "Net debt = total debt minus cash (billion IRR). Default 0.",
                    },
                    "shares_outstanding": {
                        "type": "number",
                        "description": "Shares outstanding in millions. Default 1000.",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_pl_model",
            "description": (
                "Build a multi-year P&L (Profit & Loss) projection. "
                "Projects Revenue → Gross Profit → EBITDA → EBIT → Net Income waterfall. "
                "Creates an Excel spreadsheet."
            ),
            "parameters": {
                "type": "object",
                "required": [
                    "company_name", "base_revenue", "revenue_growth_rates",
                    "gross_margin", "ebitda_margin", "da_pct",
                    "interest_expense", "tax_rate",
                ],
                "properties": {
                    "company_name": {"type": "string"},
                    "base_revenue": {
                        "type": "number",
                        "description": "Base year revenue (billion IRR)",
                    },
                    "revenue_growth_rates": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Annual revenue growth rates (decimals). One per projection year.",
                    },
                    "gross_margin": {
                        "type": "number",
                        "description": "Gross profit margin decimal (e.g. 0.40 for 40%)",
                    },
                    "ebitda_margin": {
                        "type": "number",
                        "description": "EBITDA margin decimal (e.g. 0.25 for 25%)",
                    },
                    "da_pct": {
                        "type": "number",
                        "description": "D&A as % of revenue decimal (e.g. 0.05 for 5%)",
                    },
                    "interest_expense": {
                        "type": "number",
                        "description": "Annual interest expense (billion IRR)",
                    },
                    "tax_rate": {
                        "type": "number",
                        "description": "Corporate tax rate decimal. Iran standard: 0.25",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_loan_amortization",
            "description": (
                "Build a loan amortization schedule. Supports three types: "
                "fully_amortizing (equal periodic payments), "
                "bullet (interest-only with principal at maturity), "
                "balloon (partial amortization then lump sum). "
                "Creates an Excel amortization table."
            ),
            "parameters": {
                "type": "object",
                "required": ["principal", "annual_rate", "term_months"],
                "properties": {
                    "principal": {
                        "type": "number",
                        "description": "Loan principal (million IRR)",
                    },
                    "annual_rate": {
                        "type": "number",
                        "description": "Annual interest rate decimal (e.g. 0.18 for 18%)",
                    },
                    "term_months": {
                        "type": "integer",
                        "description": "Loan term in months",
                    },
                    "loan_type": {
                        "type": "string",
                        "enum": ["fully_amortizing", "bullet", "balloon"],
                        "description": "Amortization type. Default: fully_amortizing",
                    },
                    "balloon_month": {
                        "type": "integer",
                        "description": "Month of balloon payment (balloon type only)",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_bond_model",
            "description": (
                "Price a bond and compute risk metrics: YTM (via IRR), "
                "Macaulay duration, and Modified duration. "
                "Creates an Excel cash flow schedule."
            ),
            "parameters": {
                "type": "object",
                "required": ["face_value", "coupon_rate", "periods", "ytm"],
                "properties": {
                    "face_value": {
                        "type": "number",
                        "description": "Face/par value of the bond (IRR)",
                    },
                    "coupon_rate": {
                        "type": "number",
                        "description": "Annual coupon rate decimal (e.g. 0.18 for 18%)",
                    },
                    "periods": {
                        "type": "integer",
                        "description": "Number of coupon periods until maturity",
                    },
                    "ytm": {
                        "type": "number",
                        "description": "Yield to maturity decimal (e.g. 0.20 for 20%)",
                    },
                    "frequency": {
                        "type": "integer",
                        "enum": [1, 2, 4],
                        "description": "Coupon payments per year: 1=annual, 2=semi-annual, 4=quarterly. Default: 1",
                    },
                },
            },
        },
    },
]

TOOL_DEFINITIONS += [
    {
        "type": "function",
        "function": {
            "name": "compute_wacc",
            "description": (
                "Compute WACC (Weighted Average Cost of Capital). "
                "Formula: WACC = (E/V)×Ke + (D/V)×Kd×(1-T). "
                "Chain with compute_capm to derive Ke first, then feed WACC into build_dcf_model."
            ),
            "parameters": {
                "type": "object",
                "required": ["equity_value", "debt_value", "cost_of_equity", "cost_of_debt", "tax_rate"],
                "properties": {
                    "equity_value": {"type": "number", "description": "Market value of equity"},
                    "debt_value": {"type": "number", "description": "Market value of debt (same unit as equity_value)"},
                    "cost_of_equity": {"type": "number", "description": "Cost of equity decimal (e.g. 0.20). Use compute_capm output."},
                    "cost_of_debt": {"type": "number", "description": "Pre-tax cost of debt decimal (e.g. 0.18)"},
                    "tax_rate": {"type": "number", "description": "Corporate tax rate decimal. Iran standard: 0.25"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_capm",
            "description": (
                "Compute cost of equity (Ke) using CAPM: Ke = Rf + β×ERP + size_premium + specific_premium. "
                "Iranian market defaults: Rf ≈ 20%, ERP ≈ 5-8%."
            ),
            "parameters": {
                "type": "object",
                "required": ["risk_free_rate", "beta", "equity_risk_premium"],
                "properties": {
                    "risk_free_rate": {"type": "number", "description": "Risk-free rate decimal (Iran ~0.20)"},
                    "beta": {"type": "number", "description": "Equity beta. β=1 = market risk, β<1 = defensive, β>1 = aggressive"},
                    "equity_risk_premium": {"type": "number", "description": "ERP decimal (Iran ~0.05–0.08)"},
                    "size_premium": {"type": "number", "description": "Small-cap premium decimal. Default 0."},
                    "specific_premium": {"type": "number", "description": "Company-specific risk premium decimal. Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_ddm_model",
            "description": (
                "Dividend Discount Model (DDM) valuation. Three types: "
                "'gordon' (P₀=D₁/(ke−g)), 'h_model' (linearly declining growth), "
                "'multistage' (explicit years + Gordon terminal). CFA Level 2."
            ),
            "parameters": {
                "type": "object",
                "required": ["current_dividend", "discount_rate"],
                "properties": {
                    "current_dividend": {"type": "number", "description": "Most recent annual dividend per share (D₀)"},
                    "discount_rate": {"type": "number", "description": "Required return on equity (ke) decimal"},
                    "model_type": {"type": "string", "enum": ["gordon", "h_model", "multistage"], "description": "DDM variant. Default: gordon"},
                    "growth_rate": {"type": "number", "description": "Perpetual growth rate for Gordon model"},
                    "short_term_growth": {"type": "number", "description": "Near-term growth rate for H-model"},
                    "long_term_growth": {"type": "number", "description": "Long-term sustainable growth for H-model"},
                    "half_life": {"type": "number", "description": "Half the high-growth period length in years (H-model)"},
                    "stage_growth_rates": {"type": "array", "items": {"type": "number"}, "description": "Per-year growth rates for multistage explicit period"},
                    "terminal_growth": {"type": "number", "description": "Long-run growth rate after explicit period (multistage)"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_residual_income_model",
            "description": (
                "Residual Income (RI) valuation. V₀ = B₀ + PV(explicit RIs) + PV(continuing RI). "
                "RIₜ = EPSₜ − ke × BVPSₜ₋₁. Positive RI creates premium to book. CFA Level 2."
            ),
            "parameters": {
                "type": "object",
                "required": ["book_value_per_share", "earnings_per_share_list", "cost_of_equity"],
                "properties": {
                    "book_value_per_share": {"type": "number", "description": "Book value per share at t=0"},
                    "earnings_per_share_list": {"type": "array", "items": {"type": "number"}, "description": "EPS forecast for each explicit year [EPS₁, EPS₂, ...]"},
                    "cost_of_equity": {"type": "number", "description": "Required return on equity (ke) decimal"},
                    "persistence_factor": {"type": "number", "description": "RI persistence beyond explicit period: 0=fades immediately, 1=perpetuity. Default 1."},
                    "continuing_ri_growth": {"type": "number", "description": "Growth rate of continuing RI (decimal). Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_multiples_model",
            "description": (
                "Peer comparables (multiples) valuation. "
                "Applies EV/EBITDA, P/E, P/B, P/S to estimate implied share price range."
            ),
            "parameters": {
                "type": "object",
                "required": ["ebitda", "net_income", "book_value", "revenue", "shares_outstanding",
                             "net_debt", "peer_ev_ebitda", "peer_pe", "peer_pb", "peer_ps"],
                "properties": {
                    "ebitda": {"type": "number", "description": "EBITDA"},
                    "net_income": {"type": "number", "description": "Net income (same unit as ebitda)"},
                    "book_value": {"type": "number", "description": "Total book value of equity"},
                    "revenue": {"type": "number", "description": "Total revenue"},
                    "shares_outstanding": {"type": "number", "description": "Shares in millions"},
                    "net_debt": {"type": "number", "description": "Net debt = total debt − cash"},
                    "peer_ev_ebitda": {"type": "number", "description": "Peer median EV/EBITDA multiple"},
                    "peer_pe": {"type": "number", "description": "Peer median P/E multiple"},
                    "peer_pb": {"type": "number", "description": "Peer median P/Book multiple"},
                    "peer_ps": {"type": "number", "description": "Peer median P/Sales multiple"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_fcfe",
            "description": (
                "Compute Free Cash Flow to Equity (FCFE). "
                "Direct: FCFE = NI + D&A − CapEx − ΔNWC + Net Borrowing. "
                "From FCFF: FCFE = FCFF − Interest × (1-T) + Net Borrowing."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "net_income": {"type": "number", "description": "Net income (direct path)"},
                    "da": {"type": "number", "description": "Depreciation & Amortization (direct path)"},
                    "capex": {"type": "number", "description": "Capital expenditure (direct path)"},
                    "delta_wc": {"type": "number", "description": "Change in net working capital (direct path)"},
                    "net_borrowing": {"type": "number", "description": "Net new debt issued (both paths)"},
                    "fcff": {"type": "number", "description": "FCFF value (FCFF path)"},
                    "interest_expense": {"type": "number", "description": "Interest expense (FCFF path)"},
                    "tax_rate": {"type": "number", "description": "Tax rate decimal (FCFF path)"},
                },
            },
        },
    },
]

TOOL_DEFINITIONS += [
    {
        "type": "function",
        "function": {
            "name": "build_revenue_model",
            "description": (
                "Build a multi-year revenue projection. Three approaches: "
                "'growth_rates' (compound by annual rates), "
                "'top_down' (market_size × market_share), "
                "'bottom_up' (units × price). "
                "Output feeds into build_pl_model or build_dcf_model."
            ),
            "parameters": {
                "type": "object",
                "required": ["base_revenue", "years"],
                "properties": {
                    "base_revenue": {"type": "number", "description": "Year-0 revenue (billion IRR). Used for growth_rates approach."},
                    "years": {"type": "integer", "description": "Number of forecast years"},
                    "approach": {"type": "string", "enum": ["growth_rates", "top_down", "bottom_up"], "description": "Projection method. Default: growth_rates"},
                    "growth_rates": {"type": "array", "items": {"type": "number"}, "description": "Per-year growth rates. Length must equal years."},
                    "market_size": {"type": "number", "description": "Total addressable market (top_down)"},
                    "market_share_pct": {"type": "number", "description": "Market share decimal (top_down)"},
                    "market_growth_rate": {"type": "number", "description": "Annual market growth decimal (top_down). Default 0."},
                    "units_sold": {"type": "number", "description": "Base units sold (bottom_up)"},
                    "price_per_unit": {"type": "number", "description": "Base price per unit in IRR (bottom_up)"},
                    "volume_growth_rate": {"type": "number", "description": "Annual volume growth decimal (bottom_up). Default 0."},
                    "price_growth_rate": {"type": "number", "description": "Annual price growth decimal (bottom_up). Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_wc_model",
            "description": (
                "Build a Working Capital model. "
                "AR = DSO/365 × Revenue; Inventory = DIO/365 × COGS; AP = DPO/365 × COGS. "
                "ΔWC feeds into build_dcf_model projections. CCC = DSO + DIO - DPO."
            ),
            "parameters": {
                "type": "object",
                "required": ["revenue_list", "cogs_pct", "dso", "dio", "dpo"],
                "properties": {
                    "revenue_list": {"type": "array", "items": {"type": "number"}, "description": "Revenue per year (from build_revenue_model)"},
                    "cogs_pct": {"type": "number", "description": "COGS as fraction of revenue (e.g. 0.60 for 60%)"},
                    "dso": {"type": "number", "description": "Days Sales Outstanding (e.g. 30)"},
                    "dio": {"type": "number", "description": "Days Inventory Outstanding (e.g. 45)"},
                    "dpo": {"type": "number", "description": "Days Payable Outstanding (e.g. 20)"},
                    "opening_nwc": {"type": "number", "description": "NWC at t=0 for year-1 ΔWC. Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_capex_schedule",
            "description": (
                "Build a PP&E roll-forward and depreciation schedule. "
                "DA = Gross PP&E(t-1) / useful_life (straight-line). "
                "capex and da arrays feed into build_dcf_model projections."
            ),
            "parameters": {
                "type": "object",
                "required": ["opening_gross_ppe", "opening_accum_dep", "useful_life", "years"],
                "properties": {
                    "opening_gross_ppe": {"type": "number", "description": "Gross PP&E at t=0 (billion IRR)"},
                    "opening_accum_dep": {"type": "number", "description": "Accumulated depreciation at t=0"},
                    "useful_life": {"type": "number", "description": "Asset useful life in years (straight-line)"},
                    "years": {"type": "integer", "description": "Number of forecast years"},
                    "capex_list": {"type": "array", "items": {"type": "number"}, "description": "Explicit CapEx per year"},
                    "capex_pct_revenue": {"type": "number", "description": "CapEx as % of revenue. Requires revenue_list."},
                    "revenue_list": {"type": "array", "items": {"type": "number"}, "description": "Revenue per year (needed if using capex_pct_revenue)"},
                    "disposals_list": {"type": "array", "items": {"type": "number"}, "description": "Asset disposals per year. Default zeros."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_debt_schedule",
            "description": (
                "Build a multi-tranche debt schedule. "
                "Ending = Opening - Opening×amort_pct. "
                "Interest = avg(Opening, Ending) × rate. "
                "interest_expense feeds into build_pl_model; net_debt into build_dcf_model equity bridge."
            ),
            "parameters": {
                "type": "object",
                "required": ["tranches", "years"],
                "properties": {
                    "tranches": {
                        "type": "array",
                        "description": "Debt tranches: [{name, opening_balance, annual_rate, amortization_pct}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "opening_balance": {"type": "number"},
                                "annual_rate": {"type": "number"},
                                "amortization_pct": {"type": "number"},
                            },
                        },
                    },
                    "years": {"type": "integer", "description": "Forecast years"},
                    "cash_list": {"type": "array", "items": {"type": "number"}, "description": "Cash per year for net_debt. Default zeros."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_three_statement_model",
            "description": (
                "Build linked Income Statement, Cash Flow Statement, and Balance Sheet. "
                "Takes outputs from Phase 2 tools (build_pl_model, build_capex_schedule, "
                "build_debt_schedule, build_wc_model) and links all three statements. "
                "Validates balance check (Assets = L+E) per year. CFA guide section 18."
            ),
            "parameters": {
                "type": "object",
                "required": [
                    "revenue_list", "ebit_list", "interest_expense_list", "tax_rate",
                    "da_list", "capex_list", "total_debt_list", "net_borrowing_list", "opening_bs"
                ],
                "properties": {
                    "revenue_list": {"type": "array", "items": {"type": "number"}, "description": "Revenue per year"},
                    "ebit_list": {"type": "array", "items": {"type": "number"}, "description": "EBIT per year (from build_pl_model)"},
                    "interest_expense_list": {"type": "array", "items": {"type": "number"}, "description": "Interest expense per year (from build_debt_schedule)"},
                    "tax_rate": {"type": "number", "description": "Corporate tax rate decimal (e.g. 0.25)"},
                    "da_list": {"type": "array", "items": {"type": "number"}, "description": "D&A per year (from build_capex_schedule)"},
                    "capex_list": {"type": "array", "items": {"type": "number"}, "description": "CapEx per year (from build_capex_schedule)"},
                    "total_debt_list": {"type": "array", "items": {"type": "number"}, "description": "Total debt balance per year (from build_debt_schedule)"},
                    "net_borrowing_list": {"type": "array", "items": {"type": "number"}, "description": "Net new debt per year: positive=drawdown, negative=repayment"},
                    "opening_bs": {
                        "type": "object",
                        "description": "Opening balance sheet. Required: cash, ppe_net, other_assets, other_liabilities, equity. Optional: opening_ar, opening_inventory, opening_ap.",
                    },
                    "ar_list": {"type": "array", "items": {"type": "number"}, "description": "AR per year (from build_wc_model). Optional."},
                    "inventory_list": {"type": "array", "items": {"type": "number"}, "description": "Inventory per year (from build_wc_model). Optional."},
                    "ap_list": {"type": "array", "items": {"type": "number"}, "description": "AP per year (from build_wc_model). Optional."},
                    "dividend_payout_ratio": {"type": "number", "description": "Fraction of NI paid as dividends. Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_beta",
            "description": (
                "Compute unlevered and re-levered beta using the Hamada equation. "
                "β_U = β_L / [1 + (1−T)×(D/E)]. Also computes Bloomberg adjusted beta = 2/3×β + 1/3. "
                "Feed re_levered_beta into compute_capm."
            ),
            "parameters": {
                "type": "object",
                "required": ["levered_beta", "debt_to_equity", "tax_rate"],
                "properties": {
                    "levered_beta": {"type": "number", "description": "Observed equity beta"},
                    "debt_to_equity": {"type": "number", "description": "Current D/E ratio"},
                    "tax_rate": {"type": "number", "description": "Corporate tax rate decimal"},
                    "target_debt_to_equity": {"type": "number", "description": "D/E for re-levering. Defaults to current."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_scenario_model",
            "description": (
                "Apply bear/base/bull scenarios to any model output dict. "
                "bear_value = base × (1 + bear_pct). bull_value = base × (1 + bull_pct). "
                "Per-metric overrides supported."
            ),
            "parameters": {
                "type": "object",
                "required": ["base_results", "bear_pct", "bull_pct"],
                "properties": {
                    "base_results": {"type": "object", "description": "Base-case outputs e.g. {price_per_share: 100}"},
                    "bear_pct": {"type": "number", "description": "Global bear discount decimal (e.g. -0.30)"},
                    "bull_pct": {"type": "number", "description": "Global bull premium decimal (e.g. 0.25)"},
                    "bear_overrides": {"type": "object", "description": "Per-metric bear overrides"},
                    "bull_overrides": {"type": "object", "description": "Per-metric bull overrides"},
                    "scenario_labels": {"type": "object", "description": "Optional {bear, base, bull} labels."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_operating_leverage",
            "description": (
                "Compute DOL, contribution margin, and operating breakeven. "
                "DOL = CM / EBIT. Breakeven Revenue = Fixed Costs / CM Ratio."
            ),
            "parameters": {
                "type": "object",
                "required": ["revenue", "variable_costs", "fixed_costs"],
                "properties": {
                    "revenue": {"type": "number", "description": "Total revenue"},
                    "variable_costs": {"type": "number", "description": "Total variable costs"},
                    "fixed_costs": {"type": "number", "description": "Total fixed operating costs"},
                    "units_sold": {"type": "number", "description": "Units sold for per-unit breakeven. Optional."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_pvgo",
            "description": (
                "Compute PVGO = Intrinsic Value − E₁/ke, and justified P/E ratios. "
                "Justified Leading P/E = (1−b)/(ke−g). CFA L2 equity valuation."
            ),
            "parameters": {
                "type": "object",
                "required": ["intrinsic_value", "earnings_per_share", "cost_of_equity", "growth_rate", "payout_ratio"],
                "properties": {
                    "intrinsic_value": {"type": "number", "description": "Per-share value from DDM or DCF"},
                    "earnings_per_share": {"type": "number", "description": "Forward EPS (E₁)"},
                    "cost_of_equity": {"type": "number", "description": "Required return decimal (ke)"},
                    "growth_rate": {"type": "number", "description": "Long-term growth rate decimal (g)"},
                    "payout_ratio": {"type": "number", "description": "Dividend payout ratio decimal (b)"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_eva",
            "description": (
                "Compute EVA = (ROIC − WACC) × Invested Capital, and optionally MVA. "
                "Positive EVA = value creation. ROIC = NOPAT / IC. CFA L2."
            ),
            "parameters": {
                "type": "object",
                "required": ["ebit", "tax_rate", "wacc", "invested_capital"],
                "properties": {
                    "ebit": {"type": "number", "description": "EBIT (billion IRR)"},
                    "tax_rate": {"type": "number", "description": "Corporate tax rate decimal"},
                    "wacc": {"type": "number", "description": "WACC decimal (from compute_wacc)"},
                    "invested_capital": {"type": "number", "description": "Total Debt + Equity (book value)"},
                    "market_value_of_firm": {"type": "number", "description": "Market cap + market debt. Optional — enables MVA."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "lookup_industry_benchmarks",
            "description": (
                "Search the web for industry financial benchmarks for a given business type. "
                "Returns revenue range, gross margin, EBITDA margin, and cost structure benchmarks. "
                "ALWAYS call this FIRST when a user mentions a real business (coffee shop, restaurant, "
                "trading exchange, pharmacy, gym, manufacturing, etc.) before building any model."
            ),
            "parameters": {
                "type": "object",
                "required": ["business_type"],
                "properties": {
                    "business_type": {
                        "type": "string",
                        "description": "Business type in English or Persian (e.g. 'coffee roastery', 'online trading exchange', 'قهوه‌برشته‌کاری')",
                    },
                    "country": {
                        "type": "string",
                        "description": "Country for localized benchmarks. Default: 'Iran'",
                    },
                },
            },
        },
    },
    # ── Phase 5: Advanced Wall Street ──────────────────────────────────────────
    {
        "type": "function",
        "function": {
            "name": "build_lbo_model",
            "description": (
                "Build a Leveraged Buyout (LBO) model. Computes MOIC and IRR from entry equity, "
                "debt paydown via cash sweep, and exit equity. Returns attribution of value creation "
                "to EBITDA growth, multiple expansion, and debt paydown."
            ),
            "parameters": {
                "type": "object",
                "required": ["entry_ebitda", "entry_multiple", "debt_pct", "interest_rate",
                             "ebitda_growth_rates", "exit_multiple"],
                "properties": {
                    "entry_ebitda": {"type": "number", "description": "EBITDA at acquisition (billion IRR)"},
                    "entry_multiple": {"type": "number", "description": "EV/EBITDA at entry (e.g. 8.0)"},
                    "debt_pct": {"type": "number", "description": "Debt as fraction of EV at entry (e.g. 0.65)"},
                    "interest_rate": {"type": "number", "description": "Blended interest rate decimal (e.g. 0.10)"},
                    "ebitda_growth_rates": {"type": "array", "items": {"type": "number"}, "description": "Annual EBITDA growth rates. Length = hold period."},
                    "exit_multiple": {"type": "number", "description": "EV/EBITDA at exit"},
                    "tax_rate": {"type": "number", "description": "Tax rate decimal. Default 0.25"},
                    "da_pct_ebitda": {"type": "number", "description": "D&A as fraction of EBITDA. Default 0.15"},
                    "capex_pct_ebitda": {"type": "number", "description": "CapEx as fraction of EBITDA. Default 0.12"},
                    "mandatory_amort_pct": {"type": "number", "description": "Mandatory debt amort as fraction of original debt per year. Default 0.05"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_ma_model",
            "description": (
                "Build an M&A Accretion/Dilution model. Determines if a merger is accretive "
                "(Pro Forma EPS > Standalone EPS) or dilutive. Accounts for synergies, "
                "financing costs (interest on cash/debt vs. share dilution), and goodwill."
            ),
            "parameters": {
                "type": "object",
                "required": ["acquirer_net_income", "acquirer_shares", "acquirer_share_price",
                             "target_net_income", "deal_value", "cash_pct", "tax_rate"],
                "properties": {
                    "acquirer_net_income": {"type": "number", "description": "Acquirer standalone net income"},
                    "acquirer_shares": {"type": "number", "description": "Acquirer diluted shares outstanding (millions)"},
                    "acquirer_share_price": {"type": "number", "description": "Acquirer current share price (IRR)"},
                    "target_net_income": {"type": "number", "description": "Target net income"},
                    "deal_value": {"type": "number", "description": "Total deal consideration (same unit as net incomes)"},
                    "cash_pct": {"type": "number", "description": "Fraction of deal paid in cash (vs. stock). 1.0 = all cash."},
                    "tax_rate": {"type": "number", "description": "Tax rate decimal for financing cost adjustment"},
                    "cost_synergies": {"type": "number", "description": "Annual after-tax cost synergies. Default 0."},
                    "revenue_synergies": {"type": "number", "description": "Annual after-tax revenue synergies. Default 0."},
                    "financing_rate": {"type": "number", "description": "Interest rate on new debt for cash portion. Default 0.08."},
                    "target_book_equity": {"type": "number", "description": "Target book equity for goodwill calc. Default = deal_value × 0.6."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_credit_metrics",
            "description": (
                "Compute key credit and leverage metrics for debt analysis. "
                "Returns leverage ratio, interest coverage, DSCR, and max debt capacity. "
                "Useful for evaluating LBO feasibility, covenant headroom, and credit ratings."
            ),
            "parameters": {
                "type": "object",
                "required": ["ebitda", "interest_expense", "total_debt", "cash"],
                "properties": {
                    "ebitda": {"type": "number", "description": "EBITDA (billion IRR)"},
                    "interest_expense": {"type": "number", "description": "Annual interest expense"},
                    "total_debt": {"type": "number", "description": "Total debt"},
                    "cash": {"type": "number", "description": "Cash and equivalents"},
                    "capex": {"type": "number", "description": "Capital expenditure. Default 0."},
                    "mandatory_amortization": {"type": "number", "description": "Annual mandatory debt repayment. Default 0."},
                    "revenue": {"type": "number", "description": "Revenue for debt/revenue ratio. Optional."},
                    "max_leverage_multiple": {"type": "number", "description": "Assumed max leverage for debt capacity. Default 4.0×."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_liquidation_value",
            "description": (
                "Compute liquidation value and creditor recovery in a distressed scenario. "
                "Applies recovery rates by asset class, deducts admin costs, then distributes "
                "proceeds through the claims waterfall (secured → unsecured → equity)."
            ),
            "parameters": {
                "type": "object",
                "required": ["assets", "claims"],
                "properties": {
                    "assets": {
                        "type": "array",
                        "description": "Asset list: [{name, book_value, recovery_rate}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "book_value": {"type": "number"},
                                "recovery_rate": {"type": "number", "description": "Fraction recoverable (0–1)"},
                            },
                        },
                    },
                    "claims": {
                        "type": "array",
                        "description": "Creditor claims in priority order: [{name, amount, priority}] where priority 1=senior secured",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "amount": {"type": "number"},
                                "priority": {"type": "integer"},
                            },
                        },
                    },
                    "admin_costs_pct": {"type": "number", "description": "Bankruptcy admin costs as fraction of liquidation proceeds. Default 0.05."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_ipo_pricing",
            "description": (
                "Analyze IPO economics: gross proceeds, net proceeds after underwriting, "
                "first-day underpricing, money left on the table, and greenshoe option. "
                "Useful for pre-IPO valuation analysis."
            ),
            "parameters": {
                "type": "object",
                "required": ["shares_offered", "offer_price", "pre_ipo_shares"],
                "properties": {
                    "shares_offered": {"type": "number", "description": "Shares offered in IPO (millions)"},
                    "offer_price": {"type": "number", "description": "IPO offer price (IRR)"},
                    "pre_ipo_shares": {"type": "number", "description": "Existing shares before IPO (millions)"},
                    "underwriting_discount_pct": {"type": "number", "description": "Underwriting fee as fraction of gross proceeds. Default 0.05."},
                    "first_day_close": {"type": "number", "description": "First-day closing price. Optional — for underpricing calc."},
                    "greenshoe_pct": {"type": "number", "description": "Greenshoe overallotment fraction. Default 0.15."},
                },
            },
        },
    },
    # ── Phase 6: Earnings Quality & FP&A ───────────────────────────────────────
    {
        "type": "function",
        "function": {
            "name": "compute_altman_z",
            "description": (
                "Compute Altman Z-Score for bankruptcy prediction. "
                "Public: Z = 1.2X1 + 1.4X2 + 3.3X3 + 0.6X4 + 1.0X5. "
                "Private Z': uses book equity instead of market cap in X4. "
                "Z > 2.99 = Safe; 1.81–2.99 = Grey; < 1.81 = Distress."
            ),
            "parameters": {
                "type": "object",
                "required": ["working_capital", "total_assets", "retained_earnings",
                             "ebit", "total_liabilities", "revenue"],
                "properties": {
                    "working_capital": {"type": "number", "description": "Current Assets - Current Liabilities"},
                    "total_assets": {"type": "number", "description": "Total assets"},
                    "retained_earnings": {"type": "number", "description": "Cumulative retained earnings"},
                    "ebit": {"type": "number", "description": "Earnings before interest and taxes"},
                    "total_liabilities": {"type": "number", "description": "Total debt + current liabilities"},
                    "revenue": {"type": "number", "description": "Total revenue"},
                    "market_cap": {"type": "number", "description": "Market capitalization. Use for public company. Omit for private."},
                    "book_equity": {"type": "number", "description": "Book value of equity. Used for private Z'-Score when market_cap not provided."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_beneish_score",
            "description": (
                "Compute Beneish M-Score for earnings manipulation detection. "
                "M > -1.78 suggests likely manipulation. Requires two periods of financial data "
                "to compute the 8 index ratios."
            ),
            "parameters": {
                "type": "object",
                "required": ["ar_current", "sales_current", "ar_prior", "sales_prior",
                             "gross_profit_current", "gross_profit_prior",
                             "current_assets_current", "ppe_current", "total_assets_current",
                             "current_assets_prior", "ppe_prior", "total_assets_prior",
                             "dep_current", "dep_prior", "sga_current", "sga_prior",
                             "net_income", "cfo", "ltd_current", "current_liab_current",
                             "ltd_prior", "current_liab_prior"],
                "properties": {
                    "ar_current": {"type": "number", "description": "Accounts Receivable (current year)"},
                    "sales_current": {"type": "number", "description": "Net Sales (current year)"},
                    "ar_prior": {"type": "number", "description": "Accounts Receivable (prior year)"},
                    "sales_prior": {"type": "number", "description": "Net Sales (prior year)"},
                    "gross_profit_current": {"type": "number"},
                    "gross_profit_prior": {"type": "number"},
                    "current_assets_current": {"type": "number"},
                    "ppe_current": {"type": "number", "description": "Net PP&E (current year)"},
                    "total_assets_current": {"type": "number"},
                    "current_assets_prior": {"type": "number"},
                    "ppe_prior": {"type": "number", "description": "Net PP&E (prior year)"},
                    "total_assets_prior": {"type": "number"},
                    "dep_current": {"type": "number", "description": "Depreciation expense (current year)"},
                    "dep_prior": {"type": "number", "description": "Depreciation expense (prior year)"},
                    "sga_current": {"type": "number", "description": "SG&A expense (current year)"},
                    "sga_prior": {"type": "number", "description": "SG&A expense (prior year)"},
                    "net_income": {"type": "number", "description": "Net income (current year)"},
                    "cfo": {"type": "number", "description": "Cash flow from operations (current year)"},
                    "ltd_current": {"type": "number", "description": "Long-term debt (current year)"},
                    "current_liab_current": {"type": "number"},
                    "ltd_prior": {"type": "number"},
                    "current_liab_prior": {"type": "number"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_accrual_ratios",
            "description": (
                "Compute accrual-based earnings quality metrics. "
                "High accruals relative to assets indicate lower earnings quality. "
                "Two methods: Balance Sheet (ΔNon-Cash Assets - ΔNon-Debt Liabilities) / Avg Assets, "
                "and Cash Flow (NI - CFO - CFI) / Avg Assets."
            ),
            "parameters": {
                "type": "object",
                "required": ["net_income", "cfo", "cfi",
                             "total_assets_current", "total_assets_prior"],
                "properties": {
                    "net_income": {"type": "number", "description": "Net income"},
                    "cfo": {"type": "number", "description": "Cash flow from operations"},
                    "cfi": {"type": "number", "description": "Cash flow from investing"},
                    "total_assets_current": {"type": "number"},
                    "total_assets_prior": {"type": "number"},
                    "cash_current": {"type": "number", "description": "Cash (current). For BS accrual."},
                    "cash_prior": {"type": "number"},
                    "total_debt_current": {"type": "number"},
                    "total_debt_prior": {"type": "number"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_variance_analysis",
            "description": (
                "Compute budget vs. actual variance analysis for FP&A. "
                "Splits total revenue variance into Price Variance and Volume Variance. "
                "Splits labor cost variance into Spending (Rate) Variance and Efficiency Variance."
            ),
            "parameters": {
                "type": "object",
                "required": ["actual_price", "budget_price", "actual_volume", "budget_volume"],
                "properties": {
                    "actual_price": {"type": "number", "description": "Actual selling price per unit"},
                    "budget_price": {"type": "number", "description": "Budgeted selling price per unit"},
                    "actual_volume": {"type": "number", "description": "Actual units sold"},
                    "budget_volume": {"type": "number", "description": "Budgeted units sold"},
                    "actual_hours": {"type": "number", "description": "Actual labor hours. Optional."},
                    "budget_hours": {"type": "number", "description": "Budgeted labor hours. Optional."},
                    "actual_rate": {"type": "number", "description": "Actual labor rate per hour. Optional."},
                    "budget_rate": {"type": "number", "description": "Budgeted labor rate per hour. Optional."},
                },
            },
        },
    },
    # ── Phase 7: Portfolio & Risk Analytics ────────────────────────────────────
    {
        "type": "function",
        "function": {
            "name": "compute_portfolio_stats",
            "description": (
                "Compute portfolio return, volatility, and diversification ratio. "
                "E(Rp) = Σ wᵢ×E(Rᵢ), σp = √(w'Σw), diversification ratio = (Σwᵢσᵢ)/σp. "
                "Validates weights sum to ~1.0."
            ),
            "parameters": {
                "type": "object",
                "required": ["assets", "correlation_matrix"],
                "properties": {
                    "assets": {
                        "type": "array",
                        "description": "Asset list: [{name, weight, expected_return, volatility}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "weight": {"type": "number", "description": "Portfolio weight (0–1)"},
                                "expected_return": {"type": "number", "description": "Expected annual return (decimal)"},
                                "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                            },
                        },
                    },
                    "correlation_matrix": {
                        "type": "array",
                        "description": "NxN correlation matrix as list of lists",
                        "items": {"type": "array", "items": {"type": "number"}},
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_risk_metrics",
            "description": (
                "Compute risk-adjusted performance metrics: annualized return, volatility, "
                "Sharpe ratio, Sortino ratio, max drawdown, Calmar ratio. "
                "If benchmark provided: beta, Treynor, information ratio, tracking error."
            ),
            "parameters": {
                "type": "object",
                "required": ["returns"],
                "properties": {
                    "returns": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Periodic returns (e.g. monthly)",
                    },
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal). Default 0.0."},
                    "benchmark_returns": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Benchmark periodic returns. Optional.",
                    },
                    "periods_per_year": {"type": "integer", "description": "Periods per year (12=monthly, 252=daily). Default 12."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_var",
            "description": (
                "Compute Value at Risk (VaR) using parametric, historical, or Monte Carlo method. "
                "Parametric: z-score approach with normal distribution. "
                "Historical: percentile of actual returns. "
                "Monte Carlo: simulate random paths and take percentile."
            ),
            "parameters": {
                "type": "object",
                "required": ["portfolio_value"],
                "properties": {
                    "portfolio_value": {"type": "number", "description": "Total portfolio value"},
                    "confidence_level": {"type": "number", "description": "Confidence level (0.90, 0.95, 0.975, 0.99). Default 0.95."},
                    "method": {"type": "string", "enum": ["parametric", "historical", "monte_carlo"], "description": "VaR method. Default: parametric."},
                    "horizon_days": {"type": "integer", "description": "Time horizon in days. Default 1."},
                    "expected_return": {"type": "number", "description": "Annual expected return (for parametric/MC)."},
                    "volatility": {"type": "number", "description": "Annual volatility (for parametric/MC)."},
                    "returns": {"type": "array", "items": {"type": "number"}, "description": "Historical returns (for historical method)."},
                    "num_simulations": {"type": "integer", "description": "Number of MC simulations. Default 10000."},
                    "seed": {"type": "integer", "description": "Random seed for reproducibility."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_cvar",
            "description": (
                "Compute Conditional VaR (CVaR / Expected Shortfall). "
                "CVaR = E[Loss | Loss > VaR]. "
                "Auto-selects: historical if returns provided, else parametric."
            ),
            "parameters": {
                "type": "object",
                "required": ["portfolio_value"],
                "properties": {
                    "portfolio_value": {"type": "number", "description": "Total portfolio value"},
                    "confidence_level": {"type": "number", "description": "Confidence level. Default 0.95."},
                    "horizon_days": {"type": "integer", "description": "Time horizon in days. Default 1."},
                    "expected_return": {"type": "number", "description": "Annual expected return (for parametric)."},
                    "volatility": {"type": "number", "description": "Annual volatility (for parametric)."},
                    "returns": {"type": "array", "items": {"type": "number"}, "description": "Historical returns (for historical method)."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_monte_carlo",
            "description": (
                "Run Monte Carlo simulation using Geometric Brownian Motion (GBM). "
                "S(t+dt) = S(t) × exp((μ − σ²/2)dt + σ√dt × Z). "
                "Returns terminal value statistics, percentile paths, and probability of loss."
            ),
            "parameters": {
                "type": "object",
                "required": ["initial_value", "expected_return", "volatility"],
                "properties": {
                    "initial_value": {"type": "number", "description": "Initial portfolio/asset value"},
                    "expected_return": {"type": "number", "description": "Annual expected return (decimal)"},
                    "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                    "horizon_years": {"type": "number", "description": "Simulation horizon in years. Default 1.0."},
                    "num_simulations": {"type": "integer", "description": "Number of simulation paths. Default 10000."},
                    "num_steps": {"type": "integer", "description": "Steps per path (e.g. 252 trading days). Default 252."},
                    "seed": {"type": "integer", "description": "Random seed for reproducibility."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "optimize_portfolio",
            "description": (
                "Optimize portfolio weights via Monte Carlo random search. "
                "Objectives: min_variance, max_sharpe, or target_return. "
                "Supports weight constraints (min/max per asset)."
            ),
            "parameters": {
                "type": "object",
                "required": ["assets", "correlation_matrix"],
                "properties": {
                    "assets": {
                        "type": "array",
                        "description": "Asset list: [{name, expected_return, volatility}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "expected_return": {"type": "number"},
                                "volatility": {"type": "number"},
                            },
                        },
                    },
                    "correlation_matrix": {
                        "type": "array",
                        "description": "NxN correlation matrix",
                        "items": {"type": "array", "items": {"type": "number"}},
                    },
                    "risk_free_rate": {"type": "number", "description": "Risk-free rate (decimal). Default 0.0."},
                    "objective": {"type": "string", "enum": ["min_variance", "max_sharpe", "target_return"], "description": "Optimization objective. Default: max_sharpe."},
                    "target_return": {"type": "number", "description": "Required when objective=target_return."},
                    "min_weight": {"type": "number", "description": "Minimum weight per asset. Default 0.0."},
                    "max_weight": {"type": "number", "description": "Maximum weight per asset. Default 1.0."},
                    "num_portfolios": {"type": "integer", "description": "Number of random portfolios. Default 10000."},
                    "seed": {"type": "integer", "description": "Random seed."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_efficient_frontier",
            "description": (
                "Generate efficient frontier points by finding minimum-variance portfolios "
                "at different target return levels. Also identifies the tangent (max Sharpe) "
                "and minimum variance portfolios."
            ),
            "parameters": {
                "type": "object",
                "required": ["assets", "correlation_matrix"],
                "properties": {
                    "assets": {
                        "type": "array",
                        "description": "Asset list: [{name, expected_return, volatility}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "expected_return": {"type": "number"},
                                "volatility": {"type": "number"},
                            },
                        },
                    },
                    "correlation_matrix": {
                        "type": "array",
                        "description": "NxN correlation matrix",
                        "items": {"type": "array", "items": {"type": "number"}},
                    },
                    "risk_free_rate": {"type": "number", "description": "Risk-free rate. Default 0.0."},
                    "num_points": {"type": "integer", "description": "Number of frontier points. Default 50."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_risk_parity",
            "description": (
                "Compute risk parity portfolio weights where each asset contributes "
                "equally to total portfolio risk. Uses iterative inverse-volatility "
                "weighting to equalize risk contributions."
            ),
            "parameters": {
                "type": "object",
                "required": ["assets", "correlation_matrix"],
                "properties": {
                    "assets": {
                        "type": "array",
                        "description": "Asset list: [{name, volatility}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                            },
                        },
                    },
                    "correlation_matrix": {
                        "type": "array",
                        "description": "NxN correlation matrix",
                        "items": {"type": "array", "items": {"type": "number"}},
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_factor_model",
            "description": (
                "Compute single-factor (CAPM) or Fama-French three-factor regression. "
                "Single: Rᵢ−Rf = α + β(Rm−Rf). "
                "Three-factor: add SMB and HML. Solved via OLS normal equations."
            ),
            "parameters": {
                "type": "object",
                "required": ["asset_returns", "market_returns"],
                "properties": {
                    "asset_returns": {"type": "array", "items": {"type": "number"}, "description": "Asset periodic returns"},
                    "market_returns": {"type": "array", "items": {"type": "number"}, "description": "Market periodic returns"},
                    "risk_free_rate": {"type": "number", "description": "Periodic risk-free rate. Default 0.0."},
                    "smb_returns": {"type": "array", "items": {"type": "number"}, "description": "SMB factor returns (for 3-factor)."},
                    "hml_returns": {"type": "array", "items": {"type": "number"}, "description": "HML factor returns (for 3-factor)."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_stress_test",
            "description": (
                "Run stress test scenarios on a portfolio. "
                "Each scenario defines percentage shocks per asset. "
                "Assets not mentioned in a scenario default to 0% shock."
            ),
            "parameters": {
                "type": "object",
                "required": ["portfolio", "scenarios"],
                "properties": {
                    "portfolio": {
                        "type": "array",
                        "description": "Portfolio holdings: [{asset, weight, current_value}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "asset": {"type": "string"},
                                "weight": {"type": "number"},
                                "current_value": {"type": "number"},
                            },
                        },
                    },
                    "scenarios": {
                        "type": "array",
                        "description": "Stress scenarios: [{name, shocks: {asset: pct_change}}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "shocks": {
                                    "type": "object",
                                    "description": "Map of asset name → percentage change (decimal, e.g. -0.20 for -20%)",
                                },
                            },
                        },
                    },
                },
            },
        },
    },
]

TOOL_DEFINITIONS += [
    # ── Phase 8: Derivatives & Options ────────────────────────────────
    {
        "type": "function",
        "function": {
            "name": "price_option_bsm",
            "description": (
                "Price a European option using the Black-Scholes-Merton model. "
                "Returns theoretical price, d1/d2 values, intrinsic and time value. "
                "Supports dividend yield via continuous dividend adjustment."
            ),
            "parameters": {
                "type": "object",
                "required": ["spot", "strike", "time_to_expiry", "risk_free_rate", "volatility"],
                "properties": {
                    "spot": {"type": "number", "description": "Current price of the underlying asset"},
                    "strike": {"type": "number", "description": "Option strike price"},
                    "time_to_expiry": {"type": "number", "description": "Time to expiry in years"},
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal)"},
                    "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                    "option_type": {"type": "string", "enum": ["call", "put"], "description": "Option type. Default: call."},
                    "dividend_yield": {"type": "number", "description": "Continuous dividend yield (decimal). Default 0.0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "price_option_binomial",
            "description": (
                "Price an option using the Cox-Ross-Rubinstein binomial tree model. "
                "Supports both European and American exercise styles. "
                "American options allow early exercise at each node."
            ),
            "parameters": {
                "type": "object",
                "required": ["spot", "strike", "time_to_expiry", "risk_free_rate", "volatility"],
                "properties": {
                    "spot": {"type": "number", "description": "Current price of the underlying asset"},
                    "strike": {"type": "number", "description": "Option strike price"},
                    "time_to_expiry": {"type": "number", "description": "Time to expiry in years"},
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal)"},
                    "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                    "option_type": {"type": "string", "enum": ["call", "put"], "description": "Option type. Default: call."},
                    "dividend_yield": {"type": "number", "description": "Continuous dividend yield (decimal). Default 0.0."},
                    "steps": {"type": "integer", "description": "Number of tree steps. Default 100."},
                    "exercise": {"type": "string", "enum": ["european", "american"], "description": "Exercise style. Default: european."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_greeks",
            "description": (
                "Compute option Greeks using the Black-Scholes-Merton model. "
                "Returns delta, gamma, vega (per 1% vol), theta (per day), rho (per 1% rate)."
            ),
            "parameters": {
                "type": "object",
                "required": ["spot", "strike", "time_to_expiry", "risk_free_rate", "volatility"],
                "properties": {
                    "spot": {"type": "number", "description": "Current price of the underlying asset"},
                    "strike": {"type": "number", "description": "Option strike price"},
                    "time_to_expiry": {"type": "number", "description": "Time to expiry in years"},
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal)"},
                    "volatility": {"type": "number", "description": "Annual volatility (decimal)"},
                    "option_type": {"type": "string", "enum": ["call", "put"], "description": "Option type. Default: call."},
                    "dividend_yield": {"type": "number", "description": "Continuous dividend yield (decimal). Default 0.0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_implied_volatility",
            "description": (
                "Compute implied volatility from a market option price using Newton-Raphson "
                "with bisection fallback. Returns the implied vol, iteration count, and "
                "convergence status."
            ),
            "parameters": {
                "type": "object",
                "required": ["market_price", "spot", "strike", "time_to_expiry", "risk_free_rate"],
                "properties": {
                    "market_price": {"type": "number", "description": "Observed market price of the option"},
                    "spot": {"type": "number", "description": "Current price of the underlying asset"},
                    "strike": {"type": "number", "description": "Option strike price"},
                    "time_to_expiry": {"type": "number", "description": "Time to expiry in years"},
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal)"},
                    "option_type": {"type": "string", "enum": ["call", "put"], "description": "Option type. Default: call."},
                    "dividend_yield": {"type": "number", "description": "Continuous dividend yield (decimal). Default 0.0."},
                    "max_iterations": {"type": "integer", "description": "Maximum iterations. Default 100."},
                    "tolerance": {"type": "number", "description": "Convergence tolerance. Default 1e-6."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "check_put_call_parity",
            "description": (
                "Check European put-call parity: C - P = S·e^(-qT) - K·e^(-rT). "
                "Computes deviation and identifies arbitrage opportunities with strategy."
            ),
            "parameters": {
                "type": "object",
                "required": ["call_price", "put_price", "spot", "strike", "time_to_expiry", "risk_free_rate"],
                "properties": {
                    "call_price": {"type": "number", "description": "Market call price"},
                    "put_price": {"type": "number", "description": "Market put price"},
                    "spot": {"type": "number", "description": "Current price of the underlying asset"},
                    "strike": {"type": "number", "description": "Option strike price"},
                    "time_to_expiry": {"type": "number", "description": "Time to expiry in years"},
                    "risk_free_rate": {"type": "number", "description": "Annual risk-free rate (decimal)"},
                    "dividend_yield": {"type": "number", "description": "Continuous dividend yield (decimal). Default 0.0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_option_strategy",
            "description": (
                "Build and analyze a multi-leg option strategy. "
                "Computes payoff/profit at each spot price in range, finds breakevens, "
                "max profit, and max loss. Supports calls, puts, and stock legs."
            ),
            "parameters": {
                "type": "object",
                "required": ["legs", "underlying_price"],
                "properties": {
                    "legs": {
                        "type": "array",
                        "description": "Strategy legs: [{type, position, strike, premium}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "type": {"type": "string", "enum": ["call", "put", "stock"], "description": "Leg type"},
                                "position": {"type": "string", "enum": ["long", "short"], "description": "Long or short"},
                                "strike": {"type": "number", "description": "Strike price (or purchase price for stock)"},
                                "premium": {"type": "number", "description": "Premium paid (positive) or received (positive, position determines sign)"},
                            },
                        },
                    },
                    "underlying_price": {"type": "number", "description": "Current underlying asset price"},
                    "spot_min": {"type": "number", "description": "Min spot price for payoff range. Default: underlying × 0.7."},
                    "spot_max": {"type": "number", "description": "Max spot price for payoff range. Default: underlying × 1.3."},
                    "spot_steps": {"type": "integer", "description": "Number of spot price points. Default 50."},
                },
            },
        },
    },
]

TOOL_DEFINITIONS += [
    # ── Phase 9: Iranian Market & Real Estate ─────────────────────────
    {
        "type": "function",
        "function": {
            "name": "compute_real_estate_noi",
            "description": (
                "Compute Net Operating Income (NOI) for a real estate property. "
                "Calculates effective gross income, NOI, cap rate, GRM, DSCR, "
                "and cash-on-cash return depending on inputs provided."
            ),
            "parameters": {
                "type": "object",
                "required": ["gross_rental_income", "vacancy_rate", "operating_expenses"],
                "properties": {
                    "gross_rental_income": {"type": "number", "description": "Annual gross rental income"},
                    "vacancy_rate": {"type": "number", "description": "Vacancy rate as decimal (e.g. 0.05 for 5%)"},
                    "operating_expenses": {"type": "number", "description": "Annual operating expenses"},
                    "property_value": {"type": "number", "description": "Property market value (to compute cap rate)"},
                    "cap_rate": {"type": "number", "description": "Cap rate (to compute implied value when property_value not given)"},
                    "debt_service": {"type": "number", "description": "Annual debt service payment"},
                    "equity_invested": {"type": "number", "description": "Total equity invested (for cash-on-cash return)"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_development_proforma",
            "description": (
                "Build a real estate development proforma with land cost, construction, "
                "financing, and sale projections. Calculates profit, margin, equity multiple, "
                "IRR, and breakeven sale price per square meter."
            ),
            "parameters": {
                "type": "object",
                "required": [
                    "land_cost", "construction_cost_per_sqm", "total_sqm",
                    "sellable_pct", "sale_price_per_sqm", "construction_months",
                ],
                "properties": {
                    "land_cost": {"type": "number", "description": "Total land acquisition cost"},
                    "construction_cost_per_sqm": {"type": "number", "description": "Construction cost per square meter"},
                    "total_sqm": {"type": "number", "description": "Total built area in square meters"},
                    "sellable_pct": {"type": "number", "description": "Sellable percentage (e.g. 0.85 for 85%)"},
                    "sale_price_per_sqm": {"type": "number", "description": "Expected sale price per sqm"},
                    "construction_months": {"type": "integer", "description": "Construction duration in months"},
                    "absorption_months": {"type": "integer", "description": "Sales absorption period in months. Default 6."},
                    "financing_rate": {"type": "number", "description": "Annual financing rate (decimal). Default 0.20."},
                    "equity_pct": {"type": "number", "description": "Equity percentage of hard costs. Default 0.40."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_sukuk_model",
            "description": (
                "Build an Islamic bond (sukuk) pricing model. Supports ijara (lease-based), "
                "murabaha (cost-plus), and musharaka (diminishing partnership) structures. "
                "Computes cash flows, pricing at market yield, and Macaulay duration."
            ),
            "parameters": {
                "type": "object",
                "required": ["face_value", "profit_rate", "periods"],
                "properties": {
                    "face_value": {"type": "number", "description": "Face/par value of the sukuk"},
                    "profit_rate": {"type": "number", "description": "Periodic profit rate (decimal)"},
                    "periods": {"type": "integer", "description": "Number of payment periods"},
                    "sukuk_type": {"type": "string", "enum": ["ijara", "murabaha", "musharaka"], "description": "Sukuk structure type. Default: ijara."},
                    "market_yield": {"type": "number", "description": "Market yield for pricing (decimal). If provided, computes fair price."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_murabaha_schedule",
            "description": (
                "Build a Murabaha (cost-plus financing) payment schedule. "
                "Computes markup, total price, installment amount, and full "
                "payment schedule with optional grace period."
            ),
            "parameters": {
                "type": "object",
                "required": ["cost_price", "markup_rate", "installments"],
                "properties": {
                    "cost_price": {"type": "number", "description": "Original cost/purchase price of the asset"},
                    "markup_rate": {"type": "number", "description": "Markup rate (decimal, e.g. 0.20 for 20%)"},
                    "installments": {"type": "integer", "description": "Number of installment payments"},
                    "grace_months": {"type": "integer", "description": "Grace period months (no payments). Default 0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_ijara_model",
            "description": (
                "Build an Ijara (Islamic lease) model. Computes total rental, maintenance, "
                "net yield, effective annual rate, and a payment schedule including "
                "the final ownership transfer payment."
            ),
            "parameters": {
                "type": "object",
                "required": ["asset_value", "lease_term_months", "monthly_rent", "transfer_price"],
                "properties": {
                    "asset_value": {"type": "number", "description": "Original asset value"},
                    "lease_term_months": {"type": "integer", "description": "Lease duration in months"},
                    "monthly_rent": {"type": "number", "description": "Monthly rental payment"},
                    "transfer_price": {"type": "number", "description": "Final ownership transfer price"},
                    "maintenance_pct": {"type": "number", "description": "Annual maintenance as % of asset value. Default 0.01."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_inflation_adjusted_valuation",
            "description": (
                "Deflate nominal values to real terms using CPI data or inflation rates. "
                "Computes nominal and real CAGR, and measures inflation's impact on returns. "
                "Useful for analyzing Iranian asset values in real terms."
            ),
            "parameters": {
                "type": "object",
                "required": ["nominal_values", "base_year"],
                "properties": {
                    "nominal_values": {
                        "type": "array",
                        "description": "List of {year, value} objects with nominal values",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": "integer"},
                                "value": {"type": "number"},
                            },
                        },
                    },
                    "base_year": {"type": "integer", "description": "Base year for deflation (real values expressed in this year's prices)"},
                    "cpi_values": {
                        "type": "array",
                        "description": "List of {year, cpi} objects. Provide either this or inflation_rates.",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": "integer"},
                                "cpi": {"type": "number"},
                            },
                        },
                    },
                    "inflation_rates": {
                        "type": "array",
                        "description": "List of {year, rate} objects (e.g. 0.30 for 30%). Provide either this or cpi_values.",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": "integer"},
                                "rate": {"type": "number"},
                            },
                        },
                    },
                    "metric_name": {"type": "string", "description": "Name of the metric being adjusted. Default: 'value'."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "build_tehran_housing_model",
            "description": (
                "Build a Tehran housing investment analysis model. Calculates gross/net yield, "
                "5-year total return with appreciation, optional mortgage analysis, "
                "and buy-vs-rent breakeven estimation."
            ),
            "parameters": {
                "type": "object",
                "required": ["area_sqm", "price_per_sqm", "monthly_rent_per_sqm", "annual_appreciation_pct"],
                "properties": {
                    "area_sqm": {"type": "number", "description": "Property area in square meters"},
                    "price_per_sqm": {"type": "number", "description": "Purchase price per square meter"},
                    "monthly_rent_per_sqm": {"type": "number", "description": "Monthly rent per square meter"},
                    "annual_appreciation_pct": {"type": "number", "description": "Expected annual price appreciation (decimal, e.g. 0.25 for 25%)"},
                    "maintenance_cost_pct": {"type": "number", "description": "Annual maintenance as % of property value. Default 0.01."},
                    "vacancy_months_per_year": {"type": "number", "description": "Expected vacancy months per year. Default 0."},
                    "mortgage_amount": {"type": "number", "description": "Mortgage loan amount (optional)"},
                    "mortgage_rate": {"type": "number", "description": "Annual mortgage rate (decimal, optional)"},
                    "mortgage_term_months": {"type": "integer", "description": "Mortgage term in months (optional)"},
                },
            },
        },
    },
]

# ── Phase 10 — Excel Formula Gaps ──────────────────────────────────────
TOOL_DEFINITIONS += [
    {
        "type": "function",
        "function": {
            "name": "compute_dupont",
            "description": (
                "Perform DuPont ROE decomposition. Supports 3-factor (profit margin × "
                "asset turnover × equity multiplier) and 5-factor (adds tax burden and "
                "interest burden) analysis."
            ),
            "parameters": {
                "type": "object",
                "required": ["net_income", "sales", "total_assets", "total_equity"],
                "properties": {
                    "net_income": {"type": "number", "description": "Net income"},
                    "sales": {"type": "number", "description": "Total sales / revenue"},
                    "total_assets": {"type": "number", "description": "Total assets"},
                    "total_equity": {"type": "number", "description": "Total shareholders' equity"},
                    "ebit": {"type": "number", "description": "EBIT (for 5-factor mode, optional)"},
                    "ebt": {"type": "number", "description": "Earnings before tax (for 5-factor mode, optional)"},
                    "tax": {"type": "number", "description": "Tax expense (optional)"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_brinson_attribution",
            "description": (
                "Brinson-Fachler performance attribution. Decomposes active return into "
                "allocation, selection, and interaction effects across sectors."
            ),
            "parameters": {
                "type": "object",
                "required": ["sectors"],
                "properties": {
                    "sectors": {
                        "type": "array",
                        "description": "List of sector data for attribution analysis.",
                        "items": {
                            "type": "object",
                            "required": ["name", "portfolio_weight", "benchmark_weight", "portfolio_return", "benchmark_return"],
                            "properties": {
                                "name": {"type": "string", "description": "Sector name"},
                                "portfolio_weight": {"type": "number", "description": "Portfolio weight in this sector (decimal)"},
                                "benchmark_weight": {"type": "number", "description": "Benchmark weight in this sector (decimal)"},
                                "portfolio_return": {"type": "number", "description": "Portfolio return in this sector (decimal)"},
                                "benchmark_return": {"type": "number", "description": "Benchmark return in this sector (decimal)"},
                            },
                        },
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_black_litterman",
            "description": (
                "Black-Litterman asset allocation model. Combines market equilibrium with "
                "investor views to produce posterior expected returns and optimal weights. "
                "Limited to 10 assets maximum."
            ),
            "parameters": {
                "type": "object",
                "required": ["market_caps", "covariance_matrix", "risk_aversion", "tau", "views", "view_confidences"],
                "properties": {
                    "market_caps": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Market capitalizations for each asset",
                    },
                    "covariance_matrix": {
                        "type": "array",
                        "description": "NxN covariance matrix (list of lists)",
                        "items": {"type": "array", "items": {"type": "number"}},
                    },
                    "risk_aversion": {"type": "number", "description": "Risk aversion coefficient (delta, typically 2-4)"},
                    "tau": {"type": "number", "description": "Scaling factor for uncertainty in equilibrium (typically 0.025-0.05)"},
                    "views": {
                        "type": "array",
                        "description": "Investor views. Each view: {assets: [indices], weights: [floats], expected_return: float}",
                        "items": {
                            "type": "object",
                            "required": ["assets", "weights", "expected_return"],
                            "properties": {
                                "assets": {"type": "array", "items": {"type": "integer"}, "description": "Asset indices in the view"},
                                "weights": {"type": "array", "items": {"type": "number"}, "description": "View weights (sum to 0 for relative, 1 for absolute)"},
                                "expected_return": {"type": "number", "description": "Expected return for this view (decimal)"},
                            },
                        },
                    },
                    "view_confidences": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Confidence in each view (0 to 1, higher = more confident)",
                    },
                    "risk_free_rate": {"type": "number", "description": "Risk-free rate (decimal). Default 0.0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_pe_fund_metrics",
            "description": (
                "Compute private equity fund performance metrics: TVPI, DPI, RVPI, "
                "and optionally money-weighted return (IRR)."
            ),
            "parameters": {
                "type": "object",
                "required": ["contributions", "distributions", "nav"],
                "properties": {
                    "contributions": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "List of capital contributions (positive amounts)",
                    },
                    "distributions": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "List of distributions to LPs (positive amounts)",
                    },
                    "nav": {"type": "number", "description": "Current net asset value (residual value)"},
                    "dates": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "ISO date strings for each cash flow period (optional, triggers IRR calculation)",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_omega_ratio",
            "description": (
                "Compute the Omega ratio and related downside risk metrics. "
                "Measures the probability-weighted ratio of gains to losses relative to a threshold."
            ),
            "parameters": {
                "type": "object",
                "required": ["returns"],
                "properties": {
                    "returns": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "List of periodic returns (decimals, e.g. 0.05 for 5%)",
                    },
                    "threshold": {"type": "number", "description": "Return threshold (decimal). Default 0.0."},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_credit_risk",
            "description": (
                "Compute credit risk metrics: expected loss, unexpected loss, Credit VaR. "
                "Optionally runs Merton structural model for distance-to-default and "
                "market-implied probability of default."
            ),
            "parameters": {
                "type": "object",
                "required": ["ead", "pd", "lgd"],
                "properties": {
                    "ead": {"type": "number", "description": "Exposure at default"},
                    "pd": {"type": "number", "description": "Probability of default (0 to 1)"},
                    "lgd": {"type": "number", "description": "Loss given default (0 to 1)"},
                    "asset_value": {"type": "number", "description": "Firm asset value for Merton model (optional)"},
                    "debt_face": {"type": "number", "description": "Face value of debt for Merton model (optional)"},
                    "asset_volatility": {"type": "number", "description": "Asset volatility for Merton model (decimal, optional)"},
                    "time_horizon": {"type": "number", "description": "Time horizon in years for Merton model (optional)"},
                    "risk_free_rate": {"type": "number", "description": "Risk-free rate for Merton model (decimal, optional)"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compute_forward_rates",
            "description": (
                "Compute forward rates from spot rates, bootstrap spot rates from par rates, "
                "or solve for Z-spread given bond cash flows and market price."
            ),
            "parameters": {
                "type": "object",
                "required": [],
                "properties": {
                    "spot_rates": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Spot rates at each maturity (decimals)",
                    },
                    "maturities": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Maturities in years. Default [1, 2, 3, ...].",
                    },
                    "par_rates": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Par coupon rates for bootstrap mode (decimals)",
                    },
                    "cash_flows": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "Bond cash flows for Z-spread calculation",
                    },
                    "price": {"type": "number", "description": "Bond market price for Z-spread calculation"},
                },
            },
        },
    },
]

TOOL_DISPATCH = {
    "build_dcf_model": build_dcf_model,
    "build_pl_model": build_pl_model,
    "build_loan_amortization": build_loan_amortization,
    "build_bond_model": build_bond_model,
    "compute_wacc": compute_wacc,
    "compute_capm": compute_capm,
    "build_ddm_model": build_ddm_model,
    "build_residual_income_model": build_residual_income_model,
    "build_multiples_model": build_multiples_model,
    "compute_fcfe": compute_fcfe,
    "build_revenue_model": build_revenue_model,
    "build_wc_model": build_wc_model,
    "build_capex_schedule": build_capex_schedule,
    "build_debt_schedule": build_debt_schedule,
    "build_three_statement_model": build_three_statement_model,
    "compute_beta": compute_beta,
    "build_scenario_model": build_scenario_model,
    "compute_operating_leverage": compute_operating_leverage,
    "compute_pvgo": compute_pvgo,
    "compute_eva": compute_eva,
    "lookup_industry_benchmarks": lookup_industry_benchmarks,
    # Phase 5
    "build_lbo_model": build_lbo_model,
    "build_ma_model": build_ma_model,
    "compute_credit_metrics": compute_credit_metrics,
    "compute_liquidation_value": compute_liquidation_value,
    "compute_ipo_pricing": compute_ipo_pricing,
    # Phase 6
    "compute_altman_z": compute_altman_z,
    "compute_beneish_score": compute_beneish_score,
    "compute_accrual_ratios": compute_accrual_ratios,
    "compute_variance_analysis": compute_variance_analysis,
    # Phase 7 — Portfolio & Risk
    "compute_portfolio_stats": compute_portfolio_stats,
    "compute_risk_metrics": compute_risk_metrics,
    "compute_var": compute_var,
    "compute_cvar": compute_cvar,
    "run_monte_carlo": run_monte_carlo,
    "optimize_portfolio": optimize_portfolio,
    "compute_efficient_frontier": compute_efficient_frontier,
    "compute_risk_parity": compute_risk_parity,
    "compute_factor_model": compute_factor_model,
    "run_stress_test": run_stress_test,
    # Phase 8 — Derivatives & Options
    "price_option_bsm": price_option_bsm,
    "price_option_binomial": price_option_binomial,
    "compute_greeks": compute_greeks,
    "compute_implied_volatility": compute_implied_volatility,
    "check_put_call_parity": check_put_call_parity,
    "build_option_strategy": build_option_strategy,
    # Phase 9 — Iranian Market & Real Estate
    "compute_real_estate_noi": compute_real_estate_noi,
    "build_development_proforma": build_development_proforma,
    "build_sukuk_model": build_sukuk_model,
    "build_murabaha_schedule": build_murabaha_schedule,
    "build_ijara_model": build_ijara_model,
    "compute_inflation_adjusted_valuation": compute_inflation_adjusted_valuation,
    "build_tehran_housing_model": build_tehran_housing_model,
    # Phase 10 — Excel Formula Gaps
    "compute_dupont": compute_dupont,
    "compute_brinson_attribution": compute_brinson_attribution,
    "compute_black_litterman": compute_black_litterman,
    "compute_pe_fund_metrics": compute_pe_fund_metrics,
    "compute_omega_ratio": compute_omega_ratio,
    "compute_credit_risk": compute_credit_risk,
    "compute_forward_rates": compute_forward_rates,
}
